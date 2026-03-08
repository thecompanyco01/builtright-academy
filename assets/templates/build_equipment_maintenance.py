#!/usr/bin/env python3
"""Build professional Equipment Maintenance Log Excel template."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
NAVY = "1A1A2E"
ORANGE = "F97316"
WHITE = "FFFFFF"
ALT_ROW = "F3F4F6"
LIGHT_ORANGE = "FFF7ED"
BORDER_COLOR = "D1D5DB"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

orange_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")

alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
no_fill = PatternFill(fill_type=None)

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

data_font = Font(name="Calibri", size=11)
data_align = Alignment(vertical="center")
center_align = Alignment(horizontal="center", vertical="center")
currency_fmt = '#,##0.00'
date_fmt = 'YYYY-MM-DD'
number_fmt = '#,##0'

# Title row style
title_font = Font(name="Calibri", bold=True, color=NAVY, size=14)
subtitle_font = Font(name="Calibri", italic=True, color="6B7280", size=10)


def style_header(ws, row, cols, use_orange=False):
    fill = orange_fill if use_orange else header_fill
    font = orange_font if use_orange else header_font
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, cols, fmt_map=None):
    """Apply alternating row colors and borders. fmt_map: {col_index: format_string}"""
    fmt_map = fmt_map or {}
    for r in range(start_row, end_row + 1):
        fill = alt_fill if (r - start_row) % 2 == 0 else no_fill
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center_align if c == 1 else data_align
            if c in fmt_map:
                cell.number_format = fmt_map[c]


def add_title(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = title_font
    t.alignment = Alignment(vertical="center")
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = subtitle_font
    s.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18


# ═══════════════════════════════════════════════════════════════════════
# SHEET 1: Equipment List
# ═══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Equipment List"
ws1.sheet_properties.tabColor = NAVY

headers1 = ["ID#", "Description", "Make / Model", "Serial #", "Purchase Date",
            "Purchase Price", "Warranty Exp.", "Location", "Status"]
widths1 = [10, 30, 22, 20, 15, 16, 15, 20, 14]

add_title(ws1, "📋 Equipment Inventory", "Track all equipment and vehicles in your fleet", len(headers1))

HR = 4  # header row
for i, h in enumerate(headers1, 1):
    ws1.cell(row=HR, column=i, value=h)
ws1.row_dimensions[HR].height = 28
style_header(ws1, HR, len(headers1))

DATA_START = HR + 1
DATA_END = DATA_START + 29  # 30 rows

# Pre-fill ID column with EQ-001 etc.
for r in range(DATA_START, DATA_END + 1):
    ws1.cell(row=r, column=1, value=f"EQ-{r - DATA_START + 1:03d}")

style_data_rows(ws1, DATA_START, DATA_END, len(headers1),
                fmt_map={5: date_fmt, 6: currency_fmt, 7: date_fmt})

# Status dropdown
dv_status = DataValidation(type="list", formula1='"Active,In Repair,Retired"', allow_blank=True)
dv_status.error = "Please select Active, In Repair, or Retired"
dv_status.errorTitle = "Invalid Status"
ws1.add_data_validation(dv_status)
dv_status.add(f"I{DATA_START}:I{DATA_END}")

for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws1.freeze_panes = f"A{DATA_START}"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2: Maintenance Log
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Maintenance Log")
ws2.sheet_properties.tabColor = ORANGE

headers2 = ["Date", "Equipment ID", "Type", "Description", "Parts Used",
            "Labor Hours", "Total Cost", "Next Service Date", "Technician"]
widths2 = [14, 14, 16, 35, 25, 13, 14, 16, 18]

add_title(ws2, "🔧 Maintenance Log", "Record every maintenance activity — costs auto-total below", len(headers2))

HR2 = 4
for i, h in enumerate(headers2, 1):
    ws2.cell(row=HR2, column=i, value=h)
ws2.row_dimensions[HR2].height = 28
style_header(ws2, HR2, len(headers2))

D2_START = HR2 + 1
D2_END = D2_START + 49  # 50 rows

style_data_rows(ws2, D2_START, D2_END, len(headers2),
                fmt_map={1: date_fmt, 6: '#,##0.0', 7: currency_fmt, 8: date_fmt})

# Type dropdown
dv_type = DataValidation(type="list", formula1='"Preventive,Repair,Inspection"', allow_blank=True)
dv_type.error = "Please select Preventive, Repair, or Inspection"
ws2.add_data_validation(dv_type)
dv_type.add(f"C{D2_START}:C{D2_END}")

# Equipment ID dropdown referencing Sheet 1
eq_range = f"'Equipment List'!$A${DATA_START}:$A${DATA_END}"
dv_eq = DataValidation(type="list", formula1=eq_range, allow_blank=True)
ws2.add_data_validation(dv_eq)
dv_eq.add(f"B{D2_START}:B{D2_END}")

# Running cost total row
total_row = D2_END + 1
ws2.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
tc = ws2.cell(row=total_row, column=1, value="TOTAL MAINTENANCE COST")
tc.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
tc.fill = orange_fill
tc.alignment = Alignment(horizontal="right", vertical="center")
for c in range(2, 6):
    ws2.cell(row=total_row, column=c).fill = orange_fill
    ws2.cell(row=total_row, column=c).border = thin_border

# Hours total
ht = ws2.cell(row=total_row, column=6, value=f"=SUM(F{D2_START}:F{D2_END})")
ht.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
ht.fill = orange_fill
ht.number_format = '#,##0.0'
ht.border = thin_border
ht.alignment = center_align

# Cost total
ct = ws2.cell(row=total_row, column=7, value=f"=SUM(G{D2_START}:G{D2_END})")
ct.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
ct.fill = orange_fill
ct.number_format = currency_fmt
ct.border = thin_border
ct.alignment = center_align

for c in [8, 9]:
    ws2.cell(row=total_row, column=c).fill = orange_fill
    ws2.cell(row=total_row, column=c).border = thin_border

for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = f"A{D2_START}"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3: Service Schedule
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Service Schedule")
ws3.sheet_properties.tabColor = "22C55E"

headers3 = ["Equipment ID", "Description", "Service Interval (days)",
            "Last Service Date", "Next Due Date", "Days Until Due", "Status"]
widths3 = [14, 30, 20, 18, 18, 16, 14]

add_title(ws3, "📅 Service Schedule", "Auto-calculates due dates — red=overdue, yellow=due soon, green=OK", len(headers3))

HR3 = 4
for i, h in enumerate(headers3, 1):
    ws3.cell(row=HR3, column=i, value=h)
ws3.row_dimensions[HR3].height = 28
style_header(ws3, HR3, len(headers3))

D3_START = HR3 + 1
D3_END = D3_START + 29  # 30 rows for equipment

# Pre-fill equipment IDs and formulas
for r in range(D3_START, D3_END + 1):
    idx = r - D3_START + 1
    ws3.cell(row=r, column=1, value=f"EQ-{idx:03d}")
    # Link description from Equipment List
    ws3.cell(row=r, column=2, value=f"=IF('Equipment List'!B{DATA_START + idx - 1}<>\"\", 'Equipment List'!B{DATA_START + idx - 1}, \"\")")
    # Next Due Date = Last Service Date + Interval
    ws3.cell(row=r, column=5, value=f'=IF(AND(D{r}<>"",C{r}<>""), D{r}+C{r}, "")')
    # Days Until Due = Next Due Date - TODAY()
    ws3.cell(row=r, column=6, value=f'=IF(E{r}<>"", E{r}-TODAY(), "")')
    # Status text
    ws3.cell(row=r, column=7, value=f'=IF(F{r}="","",IF(F{r}<0,"OVERDUE",IF(F{r}<=7,"DUE SOON","OK")))')

style_data_rows(ws3, D3_START, D3_END, len(headers3),
                fmt_map={3: number_fmt, 4: date_fmt, 5: date_fmt, 6: number_fmt})

# Conditional formatting on Days Until Due (column F) and Status (column G)
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
red_font = Font(color="DC2626", bold=True)
yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
yellow_font = Font(color="A16207", bold=True)
green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
green_font = Font(color="16A34A", bold=True)

# Apply to Status column G
range_g = f"G{D3_START}:G{D3_END}"
ws3.conditional_formatting.add(range_g, CellIsRule(operator="equal", formula=['"OVERDUE"'], fill=red_fill, font=red_font))
ws3.conditional_formatting.add(range_g, CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=yellow_fill, font=yellow_font))
ws3.conditional_formatting.add(range_g, CellIsRule(operator="equal", formula=['"OK"'], fill=green_fill, font=green_font))

# Also color the Days Until Due column F
range_f = f"F{D3_START}:F{D3_END}"
ws3.conditional_formatting.add(range_f, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font))
ws3.conditional_formatting.add(range_f, FormulaRule(formula=[f"AND(F{D3_START}>=0, F{D3_START}<=7)"], fill=yellow_fill, font=yellow_font))
ws3.conditional_formatting.add(range_f, CellIsRule(operator="greaterThan", formula=["7"], fill=green_fill, font=green_font))

for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = f"A{D3_START}"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4: Cost Summary
# ═══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Cost Summary")
ws4.sheet_properties.tabColor = "8B5CF6"

headers4 = ["Equipment ID", "Description", "Total Maintenance Cost",
            "Number of Services", "Avg Cost per Service"]
widths4 = [14, 30, 22, 18, 22]

add_title(ws4, "💰 Cost Summary", "Auto-calculated from Maintenance Log — no manual entry needed", len(headers4))

HR4 = 4
for i, h in enumerate(headers4, 1):
    ws4.cell(row=HR4, column=i, value=h)
ws4.row_dimensions[HR4].height = 28
style_header(ws4, HR4, len(headers4))

D4_START = HR4 + 1
D4_END = D4_START + 29  # 30 rows

log_eq_range = f"'Maintenance Log'!$B${D2_START}:$B${D2_END}"
log_cost_range = f"'Maintenance Log'!$G${D2_START}:$G${D2_END}"

for r in range(D4_START, D4_END + 1):
    idx = r - D4_START + 1
    eq_id = f"EQ-{idx:03d}"
    ws4.cell(row=r, column=1, value=eq_id)
    # Description from Equipment List
    ws4.cell(row=r, column=2, value=f"=IF('Equipment List'!B{DATA_START + idx - 1}<>\"\", 'Equipment List'!B{DATA_START + idx - 1}, \"\")")
    # Total Maintenance Cost (SUMIF)
    ws4.cell(row=r, column=3, value=f'=SUMIF({log_eq_range}, A{r}, {log_cost_range})')
    # Number of Services (COUNTIF)
    ws4.cell(row=r, column=4, value=f'=COUNTIF({log_eq_range}, A{r})')
    # Avg Cost per Service
    ws4.cell(row=r, column=5, value=f'=IF(D{r}>0, C{r}/D{r}, 0)')

style_data_rows(ws4, D4_START, D4_END, len(headers4),
                fmt_map={3: currency_fmt, 4: number_fmt, 5: currency_fmt})

# Grand total row
gt_row = D4_END + 1
ws4.merge_cells(start_row=gt_row, start_column=1, end_row=gt_row, end_column=2)
gt = ws4.cell(row=gt_row, column=1, value="GRAND TOTAL")
gt.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
gt.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
gt.alignment = Alignment(horizontal="right", vertical="center")
ws4.cell(row=gt_row, column=2).fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")

purple_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
gt_cost = ws4.cell(row=gt_row, column=3, value=f"=SUM(C{D4_START}:C{D4_END})")
gt_cost.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
gt_cost.fill = purple_fill
gt_cost.number_format = currency_fmt
gt_cost.border = thin_border

gt_svc = ws4.cell(row=gt_row, column=4, value=f"=SUM(D{D4_START}:D{D4_END})")
gt_svc.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
gt_svc.fill = purple_fill
gt_svc.number_format = number_fmt
gt_svc.border = thin_border

gt_avg = ws4.cell(row=gt_row, column=5, value=f'=IF(D{gt_row}>0, C{gt_row}/D{gt_row}, 0)')
gt_avg.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
gt_avg.fill = purple_fill
gt_avg.number_format = currency_fmt
gt_avg.border = thin_border

for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.freeze_panes = f"A{D4_START}"

# ── Save ────────────────────────────────────────────────────────────────
out_path = "/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/equipment-maintenance-log.xlsx"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"✅ Saved: {out_path}")
print(f"   Size: {os.path.getsize(out_path):,} bytes")
