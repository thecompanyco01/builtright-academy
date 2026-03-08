#!/usr/bin/env python3
"""Template 4: Equipment Maintenance Log - Equipment and vehicle tracker."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

NAVY = "1A1A2E"
ORANGE = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
DARK_GRAY = "6B7280"
MEDIUM_GRAY = "D1D5DB"
BLACK = "000000"
RED = "EF4444"
YELLOW = "F59E0B"
GREEN = "22C55E"

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

title_font = Font(name="Calibri", size=18, bold=True, color=WHITE)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=14, bold=True, color=ORANGE)
label_font = Font(name="Calibri", size=11, bold=True, color=BLACK)
normal_font = Font(name="Calibri", size=11, color=BLACK)
hint_font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
total_font = Font(name="Calibri", size=11, bold=True, color=NAVY)

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)
CURRENCY_FMT = '#,##0.00'
DATE_FMT = 'MM/DD/YYYY'

wb = openpyxl.Workbook()

def style_cell(ws, row, col, value=None, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def make_title_bar(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    style_cell(ws, row, 1, text, title_font, navy_fill, Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 40

def make_table_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        col = start_col + i
        style_cell(ws, row, col, h, header_font, navy_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
    ws.row_dimensions[row].height = 28

# =============================================
# SHEET 1: Equipment List
# =============================================
ws1 = wb.active
ws1.title = "Equipment List"
ws1.sheet_properties.tabColor = NAVY

cols = {'A': 10, 'B': 28, 'C': 18, 'D': 16, 'E': 14, 'F': 14, 'G': 16, 'H': 14, 'I': 12}
for c, w in cols.items():
    ws1.column_dimensions[c].width = w

make_title_bar(ws1, 1, "EQUIPMENT & VEHICLE INVENTORY", 9)

r = 3
headers = ["ID #", "Description", "Make / Model", "Serial #", "Purchase Date", "Purchase Price", "Warranty Exp.", "Location", "Status"]
make_table_header(ws1, r, headers)

# Status dropdown
status_dv = DataValidation(type="list", formula1='"Active,In Repair,Retired"', allow_blank=True)
status_dv.error = "Select: Active, In Repair, or Retired"
ws1.add_data_validation(status_dv)

# Sample equipment data
equipment = [
    ("EQ-001", "Ford F-250 Super Duty", "Ford / F-250", "1FTBF2B6XKEE12345", "01/15/2025", 45000, "01/15/2028", "Main Yard", "Active"),
    ("EQ-002", "Kubota Mini Excavator", "Kubota / KX040-4", "KX040-12345", "03/01/2025", 55000, "03/01/2028", "Main Yard", "Active"),
    ("EQ-003", "DeWalt Table Saw (10\")", "DeWalt / DWE7491RS", "DW-98765432", "01/20/2025", 650, "01/20/2028", "Shop", "Active"),
    ("EQ-004", "Cat Skid Steer Loader", "CAT / 262D3", "CAT-55678901", "06/15/2025", 42000, "06/15/2028", "Job Site A", "Active"),
    ("EQ-005", "Hilti Rotary Hammer", "Hilti / TE 70-ATC", "HI-11223344", "02/01/2025", 1800, "02/01/2028", "Truck 1", "Active"),
    ("EQ-006", "Honda Generator 7000W", "Honda / EU7000iS", "HN-44556677", "01/10/2025", 4500, "01/10/2028", "Trailer", "Active"),
    ("EQ-007", "Enclosed Cargo Trailer", "Big Tex / 70CH-16", "BT-78901234", "02/15/2025", 8500, "N/A", "Main Yard", "Active"),
    ("EQ-008", "Bosch Laser Level", "Bosch / GLL3-330CG", "BO-33445566", "03/10/2025", 450, "03/10/2028", "Truck 1", "Active"),
    ("EQ-009", "Milwaukee Impact Driver Set", "Milwaukee / 2960-22", "MW-99887766", "01/25/2025", 380, "01/25/2028", "Shop", "Active"),
    ("EQ-010", "Husqvarna Concrete Saw", "Husqvarna / K770", "HQ-55443322", "04/01/2025", 1200, "04/01/2027", "Shop", "Active"),
]

for i, (eid, desc, make, serial, pdate, price, warr, loc, status) in enumerate(equipment):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    style_cell(ws1, r, 1, eid, label_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws1, r, 2, desc, normal_font, fill, Alignment(horizontal="left"), thin_border)
    style_cell(ws1, r, 3, make, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws1, r, 4, serial, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws1, r, 5, pdate, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws1, r, 6, price, normal_font, fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
    style_cell(ws1, r, 7, warr, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws1, r, 8, loc, normal_font, fill, Alignment(horizontal="center"), thin_border)
    cell = style_cell(ws1, r, 9, status, normal_font, fill, Alignment(horizontal="center"), thin_border)
    status_dv.add(cell)

equip_data_end = r

# Add 10 more blank rows for user to fill
for i in range(10):
    r += 1
    fill = gray_fill if (i + len(equipment)) % 2 == 0 else white_fill
    for c in range(1, 10):
        fmt = CURRENCY_FMT if c == 6 else None
        style_cell(ws1, r, c, None, normal_font, fill, Alignment(horizontal="center"), thin_border, fmt)
    status_dv.add(ws1.cell(row=r, column=9))

equip_list_end = r

# Total row
r += 1
style_cell(ws1, r, 1, None, normal_font, navy_fill, border=thin_border)
for c in range(2, 6):
    style_cell(ws1, r, c, None, normal_font, navy_fill, border=thin_border)
style_cell(ws1, r, 5, "TOTAL VALUE:", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="right"), thin_border)
style_cell(ws1, r, 6, f"=SUM(F4:F{equip_list_end})", Font(name="Calibri", size=12, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
for c in range(7, 10):
    style_cell(ws1, r, c, None, normal_font, navy_fill, border=thin_border)

# =============================================
# SHEET 2: Maintenance Log
# =============================================
ws2 = wb.create_sheet("Maintenance Log")
ws2.sheet_properties.tabColor = ORANGE

cols2 = {'A': 14, 'B': 12, 'C': 14, 'D': 35, 'E': 20, 'F': 12, 'G': 14, 'H': 14, 'I': 16}
for c, w in cols2.items():
    ws2.column_dimensions[c].width = w

make_title_bar(ws2, 1, "MAINTENANCE LOG", 9)

r = 3
headers = ["Date", "Equipment ID", "Type", "Description", "Parts Used", "Labor Hrs", "Total Cost", "Next Service", "Technician"]
make_table_header(ws2, r, headers)

# Equipment ID dropdown (references sheet 1)
equip_dv = DataValidation(type="list", formula1=f"='Equipment List'!$A$4:$A${equip_list_end}", allow_blank=True)
equip_dv.error = "Select an Equipment ID from the Equipment List"
ws2.add_data_validation(equip_dv)

# Type dropdown
type_dv = DataValidation(type="list", formula1='"Preventive,Repair,Inspection,Emergency,Overhaul"', allow_blank=True)
ws2.add_data_validation(type_dv)

# Sample maintenance data
maint_data = [
    ("01/20/2025", "EQ-001", "Preventive", "Oil change and tire rotation", "Oil filter, 8qt oil", 1.5, 125.00, "04/20/2025", "Mike J."),
    ("01/25/2025", "EQ-003", "Inspection", "Annual safety inspection", "None", 0.5, 0, "01/25/2026", "Tom R."),
    ("02/01/2025", "EQ-002", "Preventive", "Hydraulic fluid change, track tension check", "Hydraulic fluid (5gal)", 2.0, 280.00, "05/01/2025", "Mike J."),
    ("02/10/2025", "EQ-005", "Repair", "Replace trigger mechanism", "Trigger assembly", 1.0, 185.00, "N/A", "Hilti Service"),
    ("02/15/2025", "EQ-006", "Preventive", "Oil change, air filter replacement", "Oil, air filter", 1.0, 95.00, "05/15/2025", "Tom R."),
    ("03/01/2025", "EQ-001", "Repair", "Replace brake pads, rotor resurface", "Brake pads (4), rotors (2)", 3.0, 450.00, "03/01/2026", "ABC Auto"),
    ("03/05/2025", "EQ-004", "Preventive", "Grease all fittings, check hydraulics", "Grease cartridges (4)", 1.5, 65.00, "06/05/2025", "Mike J."),
    ("03/10/2025", "EQ-009", "Inspection", "Battery check and calibration", "None", 0.5, 0, "09/10/2025", "Tom R."),
]

for i, (date, eid, mtype, desc, parts, hrs, cost, next_svc, tech) in enumerate(maint_data):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    style_cell(ws2, r, 1, date, normal_font, fill, Alignment(horizontal="center"), thin_border)
    cell_eid = style_cell(ws2, r, 2, eid, label_font, fill, Alignment(horizontal="center"), thin_border)
    equip_dv.add(cell_eid)
    cell_type = style_cell(ws2, r, 3, mtype, normal_font, fill, Alignment(horizontal="center"), thin_border)
    type_dv.add(cell_type)
    style_cell(ws2, r, 4, desc, normal_font, fill, Alignment(horizontal="left", wrap_text=True), thin_border)
    style_cell(ws2, r, 5, parts, normal_font, fill, Alignment(horizontal="left", wrap_text=True), thin_border)
    style_cell(ws2, r, 6, hrs, normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0')
    style_cell(ws2, r, 7, cost, normal_font, fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
    style_cell(ws2, r, 8, next_svc, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws2, r, 9, tech, normal_font, fill, Alignment(horizontal="center"), thin_border)

maint_data_end = r

# 20 more blank rows
for i in range(20):
    r += 1
    fill = gray_fill if (i + len(maint_data)) % 2 == 0 else white_fill
    for c in range(1, 10):
        fmt = CURRENCY_FMT if c == 7 else ('0.0' if c == 6 else None)
        style_cell(ws2, r, c, None, normal_font, fill, Alignment(horizontal="center"), thin_border, fmt)
    equip_dv.add(ws2.cell(row=r, column=2))
    type_dv.add(ws2.cell(row=r, column=3))

maint_log_end = r

# Totals
r += 1
for c in range(1, 6):
    style_cell(ws2, r, c, None, normal_font, navy_fill, border=thin_border)
style_cell(ws2, r, 5, "TOTALS:", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="right"), thin_border)
style_cell(ws2, r, 6, f"=SUM(F4:F{maint_log_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, '0.0')
style_cell(ws2, r, 7, f"=SUM(G4:G{maint_log_end})", Font(name="Calibri", size=12, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
for c in [8, 9]:
    style_cell(ws2, r, c, None, normal_font, navy_fill, border=thin_border)

# =============================================
# SHEET 3: Service Schedule
# =============================================
ws3 = wb.create_sheet("Service Schedule")
ws3.sheet_properties.tabColor = ORANGE

cols3 = {'A': 10, 'B': 28, 'C': 18, 'D': 16, 'E': 16, 'F': 16, 'G': 14, 'H': 14}
for c, w in cols3.items():
    ws3.column_dimensions[c].width = w

make_title_bar(ws3, 1, "SERVICE SCHEDULE", 8)

ws3.merge_cells("A2:H2")
style_cell(ws3, 2, 1, "Track upcoming maintenance — Conditional formatting: 🔴 Overdue  🟡 Due within 7 days  🟢 OK", 
           Font(name="Calibri", size=10, italic=True, color=WHITE), navy_fill, Alignment(horizontal="center"))

r = 4
headers = ["ID #", "Equipment", "Service Type", "Interval (Days)", "Last Service", "Next Due", "Days Until Due", "Status"]
make_table_header(ws3, r, headers)

# Service schedule items
schedule_items = [
    ("EQ-001", "Ford F-250 Super Duty", "Oil Change", 90, "03/01/2025", None, None, None),
    ("EQ-001", "Ford F-250 Super Duty", "Tire Rotation", 180, "01/20/2025", None, None, None),
    ("EQ-001", "Ford F-250 Super Duty", "Brake Inspection", 365, "03/01/2025", None, None, None),
    ("EQ-002", "Kubota Mini Excavator", "Hydraulic Service", 90, "02/01/2025", None, None, None),
    ("EQ-002", "Kubota Mini Excavator", "Track Tension", 30, "02/01/2025", None, None, None),
    ("EQ-003", "DeWalt Table Saw", "Safety Inspection", 365, "01/25/2025", None, None, None),
    ("EQ-003", "DeWalt Table Saw", "Blade Replacement", 90, "01/25/2025", None, None, None),
    ("EQ-004", "Cat Skid Steer", "Grease Fittings", 90, "03/05/2025", None, None, None),
    ("EQ-004", "Cat Skid Steer", "Hydraulic Check", 180, "03/05/2025", None, None, None),
    ("EQ-005", "Hilti Rotary Hammer", "Calibration", 180, "02/10/2025", None, None, None),
    ("EQ-006", "Honda Generator", "Oil Change", 90, "02/15/2025", None, None, None),
    ("EQ-006", "Honda Generator", "Air Filter", 180, "02/15/2025", None, None, None),
    ("EQ-009", "Milwaukee Impact Driver", "Battery Check", 180, "03/10/2025", None, None, None),
    ("EQ-010", "Husqvarna Concrete Saw", "Blade Inspection", 30, "03/01/2025", None, None, None),
    ("EQ-010", "Husqvarna Concrete Saw", "Full Service", 180, "03/01/2025", None, None, None),
]

for i, (eid, desc, stype, interval, last_svc, _, _, _) in enumerate(schedule_items):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    style_cell(ws3, r, 1, eid, label_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws3, r, 2, desc, normal_font, fill, Alignment(horizontal="left"), thin_border)
    style_cell(ws3, r, 3, stype, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws3, r, 4, interval, normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws3, r, 5, last_svc, normal_font, fill, Alignment(horizontal="center"), thin_border)
    # Next Due = Last Service + Interval
    style_cell(ws3, r, 6, f'=IF(E{r}="","",E{r}+D{r})', normal_font, fill, Alignment(horizontal="center"), thin_border, DATE_FMT)
    # Days Until Due = Next Due - TODAY()
    style_cell(ws3, r, 7, f'=IF(F{r}="","",F{r}-TODAY())', normal_font, fill, Alignment(horizontal="center"), thin_border, '0')
    # Status formula
    style_cell(ws3, r, 8, f'=IF(G{r}="","",IF(G{r}<0,"OVERDUE",IF(G{r}<=7,"DUE SOON","OK")))', 
               Font(name="Calibri", size=11, bold=True), fill, Alignment(horizontal="center"), thin_border)

sched_data_end = r

# Add blank rows
for i in range(5):
    r += 1
    fill = gray_fill if (i + len(schedule_items)) % 2 == 0 else white_fill
    for c in range(1, 9):
        style_cell(ws3, r, c, None, normal_font, fill, Alignment(horizontal="center"), thin_border)
    # Formulas for blank rows too
    style_cell(ws3, r, 6, f'=IF(E{r}="","",E{r}+D{r})', normal_font, fill, Alignment(horizontal="center"), thin_border, DATE_FMT)
    style_cell(ws3, r, 7, f'=IF(F{r}="","",F{r}-TODAY())', normal_font, fill, Alignment(horizontal="center"), thin_border, '0')
    style_cell(ws3, r, 8, f'=IF(G{r}="","",IF(G{r}<0,"OVERDUE",IF(G{r}<=7,"DUE SOON","OK")))',
               Font(name="Calibri", size=11, bold=True), fill, Alignment(horizontal="center"), thin_border)

sched_end = r

# Conditional formatting
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
red_font = Font(color="DC2626", bold=True)
yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
yellow_font = Font(color="D97706", bold=True)
green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
green_font = Font(color="16A34A", bold=True)

# Apply to Status column (H)
ws3.conditional_formatting.add(f"H5:H{sched_end}",
    CellIsRule(operator="equal", formula=['"OVERDUE"'], fill=red_fill, font=red_font))
ws3.conditional_formatting.add(f"H5:H{sched_end}",
    CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=yellow_fill, font=yellow_font))
ws3.conditional_formatting.add(f"H5:H{sched_end}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=green_fill, font=green_font))

# Also color the Days column
ws3.conditional_formatting.add(f"G5:G{sched_end}",
    CellIsRule(operator="lessThan", formula=['0'], fill=red_fill, font=red_font))
ws3.conditional_formatting.add(f"G5:G{sched_end}",
    FormulaRule(formula=[f'AND(G5>=0,G5<=7)'], fill=yellow_fill, font=yellow_font))
ws3.conditional_formatting.add(f"G5:G{sched_end}",
    CellIsRule(operator="greaterThan", formula=['7'], fill=green_fill, font=green_font))

# =============================================
# SHEET 4: Cost Summary
# =============================================
ws4 = wb.create_sheet("Cost Summary")
ws4.sheet_properties.tabColor = ORANGE

cols4 = {'A': 28, 'B': 16, 'C': 14, 'D': 16, 'E': 16, 'F': 14}
for c, w in cols4.items():
    ws4.column_dimensions[c].width = w

make_title_bar(ws4, 1, "MAINTENANCE COST SUMMARY", 6)

# === Cost by Equipment ===
r = 3
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
style_cell(ws4, r, 1, "◆ TOTAL COST BY EQUIPMENT", section_font, white_fill, Alignment(horizontal="left"))
for c in range(1, 7):
    ws4.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

r += 1
make_table_header(ws4, r, ["Equipment", "Equipment ID", "# Services", "Total Labor Hrs", "Total Cost", "% of Total"])

for i, (eid, desc, _, _, _, _, _, _, _) in enumerate(equipment):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    style_cell(ws4, r, 1, desc, normal_font, fill, Alignment(horizontal="left"), thin_border)
    style_cell(ws4, r, 2, eid, label_font, fill, Alignment(horizontal="center"), thin_border)
    # COUNTIF for services
    style_cell(ws4, r, 3, f'=COUNTIF(\'Maintenance Log\'!B:B,B{r})', normal_font, fill, Alignment(horizontal="center"), thin_border)
    # SUMIF for labor hours
    style_cell(ws4, r, 4, f'=SUMIF(\'Maintenance Log\'!B:B,B{r},\'Maintenance Log\'!F:F)', normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0')
    # SUMIF for cost
    style_cell(ws4, r, 5, f'=SUMIF(\'Maintenance Log\'!B:B,B{r},\'Maintenance Log\'!G:G)', total_font, fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
    # % of total
    style_cell(ws4, r, 6, None, normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0%')

equip_cost_end = r
# Totals
r += 1
equip_total_row = r
style_cell(ws4, r, 1, "TOTAL", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="right"), thin_border)
style_cell(ws4, r, 2, None, normal_font, navy_fill, border=thin_border)
style_cell(ws4, r, 3, f"=SUM(C5:C{equip_cost_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border)
style_cell(ws4, r, 4, f"=SUM(D5:D{equip_cost_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, '0.0')
style_cell(ws4, r, 5, f"=SUM(E5:E{equip_cost_end})", Font(name="Calibri", size=12, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
style_cell(ws4, r, 6, "100%", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border)

# Now fill in % formulas
for row_i in range(5, equip_cost_end + 1):
    ws4.cell(row=row_i, column=6).value = f"=IF(E{equip_total_row}=0,0,E{row_i}/E{equip_total_row})"

# === Cost by Maintenance Type ===
r += 2
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
style_cell(ws4, r, 1, "◆ COST BY MAINTENANCE TYPE", section_font, white_fill, Alignment(horizontal="left"))
for c in range(1, 7):
    ws4.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

r += 1
make_table_header(ws4, r, ["Type", "", "# Services", "Total Labor Hrs", "Total Cost", "% of Total"])

types = ["Preventive", "Repair", "Inspection", "Emergency", "Overhaul"]
type_start = r + 1
for i, mtype in enumerate(types):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    style_cell(ws4, r, 1, mtype, label_font, fill, Alignment(horizontal="left"), thin_border)
    style_cell(ws4, r, 3, f'=COUNTIF(\'Maintenance Log\'!C:C,A{r})', normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws4, r, 4, f'=SUMIF(\'Maintenance Log\'!C:C,A{r},\'Maintenance Log\'!F:F)', normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0')
    style_cell(ws4, r, 5, f'=SUMIF(\'Maintenance Log\'!C:C,A{r},\'Maintenance Log\'!G:G)', total_font, fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
    style_cell(ws4, r, 6, None, normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0%')

type_end = r
r += 1
type_total_row = r
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
style_cell(ws4, r, 1, "TOTAL", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="right"), thin_border)
style_cell(ws4, r, 3, f"=SUM(C{type_start}:C{type_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border)
style_cell(ws4, r, 4, f"=SUM(D{type_start}:D{type_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, '0.0')
style_cell(ws4, r, 5, f"=SUM(E{type_start}:E{type_end})", Font(name="Calibri", size=12, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
style_cell(ws4, r, 6, "100%", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border)

# Fill % formulas for type
for row_i in range(type_start, type_end + 1):
    ws4.cell(row=row_i, column=6).value = f"=IF(E{type_total_row}=0,0,E{row_i}/E{type_total_row})"

# === Cost by Month ===
r += 2
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
style_cell(ws4, r, 1, "◆ MONTHLY MAINTENANCE COSTS", section_font, white_fill, Alignment(horizontal="left"))
for c in range(1, 7):
    ws4.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

r += 1
make_table_header(ws4, r, ["Month", "", "# Services", "Total Labor Hrs", "Total Cost", "% of Total"])

months = [
    ("January 2025", "01/01/2025", "01/31/2025"),
    ("February 2025", "02/01/2025", "02/28/2025"),
    ("March 2025", "03/01/2025", "03/31/2025"),
    ("April 2025", "04/01/2025", "04/30/2025"),
    ("May 2025", "05/01/2025", "05/31/2025"),
    ("June 2025", "06/01/2025", "06/30/2025"),
    ("July 2025", "07/01/2025", "07/31/2025"),
    ("August 2025", "08/01/2025", "08/31/2025"),
    ("September 2025", "09/01/2025", "09/30/2025"),
    ("October 2025", "10/01/2025", "10/31/2025"),
    ("November 2025", "11/01/2025", "11/30/2025"),
    ("December 2025", "12/01/2025", "12/31/2025"),
]

month_start = r + 1
for i, (month_name, start_date, end_date) in enumerate(months):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    style_cell(ws4, r, 1, month_name, label_font, fill, Alignment(horizontal="left"), thin_border)
    # COUNTIFS with date range
    style_cell(ws4, r, 3, f'=COUNTIFS(\'Maintenance Log\'!A:A,">="&"{start_date}",\'Maintenance Log\'!A:A,"<="&"{end_date}")', 
               normal_font, fill, Alignment(horizontal="center"), thin_border)
    style_cell(ws4, r, 4, f'=SUMIFS(\'Maintenance Log\'!F:F,\'Maintenance Log\'!A:A,">="&"{start_date}",\'Maintenance Log\'!A:A,"<="&"{end_date}")',
               normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0')
    style_cell(ws4, r, 5, f'=SUMIFS(\'Maintenance Log\'!G:G,\'Maintenance Log\'!A:A,">="&"{start_date}",\'Maintenance Log\'!A:A,"<="&"{end_date}")',
               total_font, fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
    style_cell(ws4, r, 6, None, normal_font, fill, Alignment(horizontal="center"), thin_border, '0.0%')

month_end = r
r += 1
month_total_row = r
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
style_cell(ws4, r, 1, "TOTAL", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="right"), thin_border)
style_cell(ws4, r, 3, f"=SUM(C{month_start}:C{month_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border)
style_cell(ws4, r, 4, f"=SUM(D{month_start}:D{month_end})", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, '0.0')
style_cell(ws4, r, 5, f"=SUM(E{month_start}:E{month_end})", Font(name="Calibri", size=12, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border, CURRENCY_FMT)
style_cell(ws4, r, 6, "100%", Font(name="Calibri", size=11, bold=True, color=WHITE), navy_fill, Alignment(horizontal="center"), thin_border)

for row_i in range(month_start, month_end + 1):
    ws4.cell(row=row_i, column=6).value = f"=IF(E{month_total_row}=0,0,E{row_i}/E{month_total_row})"

# === Key Insights Box ===
r += 2
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
style_cell(ws4, r, 1, "◆ KEY INSIGHTS", section_font, white_fill, Alignment(horizontal="left"))
for c in range(1, 7):
    ws4.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

insights = [
    ("Most Expensive Equipment:", f'=INDEX(A5:A{equip_cost_end},MATCH(MAX(E5:E{equip_cost_end}),E5:E{equip_cost_end},0))'),
    ("Highest Cost Equipment Amount:", f'=MAX(E5:E{equip_cost_end})'),
    ("Average Cost per Service:", f'=IF(C{equip_total_row}=0,0,E{equip_total_row}/C{equip_total_row})'),
    ("Total Labor Hours YTD:", f'=D{equip_total_row}'),
    ("Total Maintenance Spend YTD:", f'=E{equip_total_row}'),
]
for i, (label, formula) in enumerate(insights):
    r += 1
    fill = gray_fill if i % 2 == 0 else white_fill
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_cell(ws4, r, 1, label, label_font, fill, Alignment(horizontal="right"), thin_border)
    ws4.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    fmt = CURRENCY_FMT if "Cost" in label or "Spend" in label or "Amount" in label else ('0.0' if "Hours" in label else None)
    style_cell(ws4, r, 4, formula, Font(name="Calibri", size=12, bold=True, color=ORANGE), 
               PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"), Alignment(horizontal="center"), thin_border, fmt)

# Save
output = "/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/equipment-maintenance-log.xlsx"
wb.save(output)
print(f"✅ Saved: {output}")
