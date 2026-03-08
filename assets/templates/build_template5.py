#!/usr/bin/env python3
"""Template 5: Weekly Timesheet Template"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

wb = openpyxl.Workbook()

# Colors
NAVY = "1A1A2E"
ORANGE = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
DARK_GRAY = "6B7280"
MEDIUM_GRAY = "D1D5DB"

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

white_font = Font(color=WHITE, bold=True, size=10)
navy_font = Font(color=NAVY, bold=True, size=10)
orange_font = Font(color=ORANGE, bold=True, size=10)
normal_font = Font(size=9, color="333333")

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)
bottom_border = Border(bottom=Side(style='thin', color=MEDIUM_GRAY))

locked = Protection(locked=True)
unlocked = Protection(locked=False)

# ==========================================
# SHEET 1: TIMESHEET
# ==========================================
ws = wb.active
ws.title = "Timesheet"

# Column widths
col_widths = {'A': 3, 'B': 14, 'C': 20, 'D': 12, 'E': 12, 'F': 10, 'G': 14, 'H': 14, 'I': 22}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Page setup
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.6)
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

# === HEADER ===
for col in range(1, 10):
    for row in range(1, 4):
        ws.cell(row=row, column=col).fill = navy_fill

ws.merge_cells('B1:E3')
ws.cell(row=1, column=2, value="WEEKLY TIMESHEET").font = Font(color=WHITE, bold=True, size=20)
ws.cell(row=1, column=2).alignment = Alignment(vertical='center')

ws.merge_cells('F1:I1')
ws.cell(row=1, column=6, value="[COMPANY NAME]").font = Font(color=WHITE, bold=True, size=12)
ws.cell(row=1, column=6).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=1, column=6).fill = navy_fill
ws.merge_cells('F2:I2')
ws.cell(row=2, column=6, value="[Company Address]").font = Font(color=MEDIUM_GRAY, size=9)
ws.cell(row=2, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=2, column=6).fill = navy_fill

# Employee info section
ws.cell(row=5, column=2, value="Employee Name:").font = navy_font
ws.merge_cells('C5:E5')
ws.cell(row=5, column=3).border = bottom_border
ws.cell(row=5, column=3).protection = unlocked

ws.cell(row=5, column=6, value="Employee ID:").font = navy_font
ws.cell(row=5, column=6).alignment = Alignment(horizontal='right')
ws.merge_cells('G5:H5')
ws.cell(row=5, column=7).border = bottom_border
ws.cell(row=5, column=7).protection = unlocked

ws.cell(row=6, column=2, value="Department:").font = navy_font
ws.merge_cells('C6:E6')
ws.cell(row=6, column=3).border = bottom_border
ws.cell(row=6, column=3).protection = unlocked

ws.cell(row=6, column=6, value="Week Starting:").font = navy_font
ws.cell(row=6, column=6).alignment = Alignment(horizontal='right')
ws.merge_cells('G6:H6')
ws.cell(row=6, column=7).border = bottom_border
ws.cell(row=6, column=7).number_format = 'MM/DD/YYYY'
ws.cell(row=6, column=7).protection = unlocked

# Orange divider
for col in range(1, 10):
    ws.cell(row=8, column=col).fill = orange_fill

# === DAILY TIME TABLE ===
header_row = 9
headers = ["", "Day / Date", "Job Name / #", "Start Time", "End Time", "Break (hrs)", "Regular Hrs", "Overtime Hrs", "Notes"]
for col_idx, header in enumerate(headers, 1):
    if col_idx == 1:
        continue
    cell = ws.cell(row=header_row, column=col_idx, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[header_row].height = 28

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

current_row = 10
day_rows = {}  # track which rows belong to each day for daily totals

for day_idx, day in enumerate(days):
    day_start_row = current_row
    
    # Each day gets 3 entry rows (can work multiple jobs per day)
    for entry in range(3):
        r = current_row
        fill = light_gray_fill if day_idx % 2 == 0 else white_fill
        
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(size=9, color="333333")
        
        if entry == 0:
            # Day name in first entry row
            ws.cell(row=r, column=2, value=day).font = Font(color=NAVY, bold=True, size=10)
            # Date formula (if week starting date is filled)
            ws.cell(row=r, column=2).value = day
        else:
            ws.cell(row=r, column=2).value = ""
        
        # Job Name (user input)
        ws.cell(row=r, column=3).protection = unlocked
        
        # Start/End Time
        ws.cell(row=r, column=4).number_format = 'h:mm AM/PM'
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).protection = unlocked
        ws.cell(row=r, column=5).number_format = 'h:mm AM/PM'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).protection = unlocked
        
        # Break hours
        ws.cell(row=r, column=6).number_format = '0.00'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=6).protection = unlocked
        
        # Regular Hours formula: MIN(8, total_worked - break)
        # Total worked = (End - Start) * 24
        ws.cell(row=r, column=7).value = f'=IF(AND(D{r}<>"",E{r}<>""),MIN(8,(E{r}-D{r})*24-IF(F{r}<>"",F{r},0)),"")'
        ws.cell(row=r, column=7).number_format = '0.00'
        ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=7).font = Font(size=9, color=NAVY, bold=True)
        ws.cell(row=r, column=7).protection = locked
        
        # Overtime Hours formula: MAX(0, total_worked - break - 8)
        ws.cell(row=r, column=8).value = f'=IF(AND(D{r}<>"",E{r}<>""),MAX(0,(E{r}-D{r})*24-IF(F{r}<>"",F{r},0)-8),"")'
        ws.cell(row=r, column=8).number_format = '0.00'
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=8).font = Font(size=9, color=ORANGE, bold=True)
        ws.cell(row=r, column=8).protection = locked
        
        # Notes
        ws.cell(row=r, column=9).protection = unlocked
        
        current_row += 1
    
    day_rows[day] = (day_start_row, current_row - 1)
    
    # Daily subtotal row
    r = current_row
    for col in range(1, 10):
        ws.cell(row=r, column=col).fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        ws.cell(row=r, column=col).border = thin_border
    
    ws.cell(row=r, column=5, value=f"{day} Total:").font = Font(color=NAVY, bold=True, size=8)
    ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
    
    start_r = day_start_row
    end_r = current_row - 1
    ws.cell(row=r, column=7).value = f'=SUM(G{start_r}:G{end_r})'
    ws.cell(row=r, column=7).number_format = '0.00'
    ws.cell(row=r, column=7).font = Font(color=NAVY, bold=True, size=9)
    ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=7).protection = locked
    
    ws.cell(row=r, column=8).value = f'=SUM(H{start_r}:H{end_r})'
    ws.cell(row=r, column=8).number_format = '0.00'
    ws.cell(row=r, column=8).font = Font(color=ORANGE, bold=True, size=9)
    ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=8).protection = locked
    
    current_row += 1

# === WEEKLY TOTALS ===
r = current_row + 1
for col in range(1, 10):
    ws.cell(row=r, column=col).fill = navy_fill
    ws.cell(row=r, column=col).border = thin_border

ws.merge_cells(f'B{r}:F{r}')
ws.cell(row=r, column=2, value="WEEKLY TOTALS").font = Font(color=WHITE, bold=True, size=12)
ws.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=r, column=2).fill = navy_fill

# Sum all daily subtotal rows for regular and OT
daily_total_rows = [day_rows[d][1] + 1 for d in days]
reg_formula = '=' + '+'.join([f'G{r}' for r in daily_total_rows])
ot_formula = '=' + '+'.join([f'H{r}' for r in daily_total_rows])

ws.cell(row=r, column=7).value = reg_formula
ws.cell(row=r, column=7).font = Font(color=WHITE, bold=True, size=12)
ws.cell(row=r, column=7).number_format = '0.00'
ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=7).fill = navy_fill
ws.cell(row=r, column=7).protection = locked

ws.cell(row=r, column=8).value = ot_formula
ws.cell(row=r, column=8).font = Font(color=ORANGE, bold=True, size=12)
ws.cell(row=r, column=8).number_format = '0.00'
ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=8).fill = navy_fill
ws.cell(row=r, column=8).protection = locked

weekly_totals_row = r

# Total hours (Reg + OT)
r += 1
ws.cell(row=r, column=6, value="Total Hours:").font = Font(color=NAVY, bold=True, size=11)
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=7).value = f'=G{weekly_totals_row}+H{weekly_totals_row}'
ws.cell(row=r, column=7).font = Font(color=NAVY, bold=True, size=14)
ws.cell(row=r, column=7).number_format = '0.00'
ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=7).protection = locked
total_hours_row = r

# === PAY CALCULATION ===
r += 2
for col in range(2, 10):
    ws.cell(row=r, column=col).fill = orange_fill
ws.merge_cells(f'B{r}:I{r}')
ws.cell(row=r, column=2, value="PAY CALCULATION").font = Font(color=WHITE, bold=True, size=12)
ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=2).fill = orange_fill
ws.row_dimensions[r].height = 25
r += 1

# Regular Rate
ws.cell(row=r, column=3, value="Regular Hourly Rate:").font = navy_font
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
ws.merge_cells(f'D{r}:E{r}')
ws.cell(row=r, column=4).number_format = '$#,##0.00'
ws.cell(row=r, column=4).font = Font(size=12, color=NAVY, bold=True)
ws.cell(row=r, column=4).border = bottom_border
ws.cell(row=r, column=4).protection = unlocked
reg_rate_row = r
r += 1

# OT Rate (auto-calculated 1.5x)
ws.cell(row=r, column=3, value="Overtime Rate (1.5x):").font = navy_font
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
ws.merge_cells(f'D{r}:E{r}')
ws.cell(row=r, column=4).value = f'=D{reg_rate_row}*1.5'
ws.cell(row=r, column=4).number_format = '$#,##0.00'
ws.cell(row=r, column=4).font = Font(size=12, color=ORANGE, bold=True)
ws.cell(row=r, column=4).protection = locked
ot_rate_row = r
r += 1

# Blank separator
r += 1

# Pay breakdown
ws.cell(row=r, column=3, value="Regular Pay:").font = navy_font
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=7, value="Hours").font = Font(size=8, color=DARK_GRAY)
ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=8, value="Rate").font = Font(size=8, color=DARK_GRAY)
ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')
ws.merge_cells(f'D{r}:E{r}')
ws.cell(row=r, column=4).value = f'=G{weekly_totals_row}*D{reg_rate_row}'
ws.cell(row=r, column=4).number_format = '$#,##0.00'
ws.cell(row=r, column=4).font = Font(size=11, color=NAVY, bold=True)
ws.cell(row=r, column=4).protection = locked
ws.cell(row=r, column=7).value = f'=G{weekly_totals_row}'
ws.cell(row=r, column=7).number_format = '0.00'
ws.cell(row=r, column=8).value = f'=D{reg_rate_row}'
ws.cell(row=r, column=8).number_format = '$#,##0.00'
reg_pay_row = r
r += 1

ws.cell(row=r, column=3, value="Overtime Pay:").font = Font(color=ORANGE, bold=True, size=10)
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
ws.merge_cells(f'D{r}:E{r}')
ws.cell(row=r, column=4).value = f'=H{weekly_totals_row}*D{ot_rate_row}'
ws.cell(row=r, column=4).number_format = '$#,##0.00'
ws.cell(row=r, column=4).font = Font(size=11, color=ORANGE, bold=True)
ws.cell(row=r, column=4).protection = locked
ws.cell(row=r, column=7).value = f'=H{weekly_totals_row}'
ws.cell(row=r, column=7).number_format = '0.00'
ws.cell(row=r, column=8).value = f'=D{ot_rate_row}'
ws.cell(row=r, column=8).number_format = '$#,##0.00'
ot_pay_row = r
r += 1

# Total Pay
for col in range(3, 6):
    ws.cell(row=r, column=col).fill = navy_fill
    ws.cell(row=r, column=col).border = thin_border
ws.cell(row=r, column=3, value="TOTAL PAY:").font = Font(color=WHITE, bold=True, size=12)
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=r, column=3).fill = navy_fill
ws.merge_cells(f'D{r}:E{r}')
ws.cell(row=r, column=4).value = f'=D{reg_pay_row}+D{ot_pay_row}'
ws.cell(row=r, column=4).number_format = '$#,##0.00'
ws.cell(row=r, column=4).font = Font(color=WHITE, bold=True, size=14)
ws.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=r, column=4).fill = navy_fill
ws.cell(row=r, column=4).protection = locked
ws.row_dimensions[r].height = 30
total_pay_row = r

# === SIGNATURES ===
r += 3
ws.cell(row=r, column=2, value="CERTIFICATION & APPROVAL").font = orange_font
r += 1
ws.cell(row=r, column=2, value="I certify that the hours reported above are accurate and complete.").font = Font(size=9, color="333333")
ws.merge_cells(f'B{r}:I{r}')
r += 2

# Employee signature
ws.cell(row=r, column=2, value="Employee Signature:").font = navy_font
ws.merge_cells(f'C{r}:E{r}')
ws.cell(row=r, column=3).border = Border(bottom=Side(style='thin', color=NAVY))
ws.cell(row=r, column=6, value="Date:").font = navy_font
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=7).border = bottom_border
ws.cell(row=r, column=7).number_format = 'MM/DD/YYYY'
r += 2

# Supervisor signature
ws.cell(row=r, column=2, value="Supervisor Signature:").font = navy_font
ws.merge_cells(f'C{r}:E{r}')
ws.cell(row=r, column=3).border = Border(bottom=Side(style='thin', color=NAVY))
ws.cell(row=r, column=6, value="Date:").font = navy_font
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=7).border = bottom_border
ws.cell(row=r, column=7).number_format = 'MM/DD/YYYY'

last_row = r + 1

# Print area
ws.print_area = f'A1:I{last_row}'
ws.oddFooter.center.text = "Page &P of &N"
ws.oddFooter.right.text = "Confidential — Employee Timesheet"

# Enable sheet protection (formulas locked, input cells unlocked)
ws.protection.sheet = True
ws.protection.password = ''  # No password needed, just protects formulas
ws.protection.enable()

# ==========================================
# SHEET 2: MONTHLY SUMMARY
# ==========================================
ws2 = wb.create_sheet("Summary")

for col in range(1, 10):
    ws2.cell(row=1, column=col).fill = navy_fill
ws2.merge_cells('A1:I1')
ws2.cell(row=1, column=1, value="MONTHLY TIMESHEET SUMMARY").font = Font(color=WHITE, bold=True, size=16)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=1, column=1).fill = navy_fill
ws2.row_dimensions[1].height = 35

# Employee info
ws2.cell(row=3, column=1, value="Employee:").font = navy_font
ws2.merge_cells('B3:D3')
ws2.cell(row=3, column=2).border = bottom_border
ws2.cell(row=3, column=5, value="Month/Year:").font = navy_font
ws2.merge_cells('F3:G3')
ws2.cell(row=3, column=6).border = bottom_border

# Weekly summary headers
headers = ["Week #", "Week Starting", "Regular Hrs", "Overtime Hrs", "Total Hrs", "Regular Pay", "OT Pay", "Total Pay", "Notes"]
widths = [10, 16, 14, 14, 12, 14, 14, 14, 25]
for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
    cell = ws2.cell(row=5, column=col_idx, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

# 5 weeks
for i in range(5):
    r = 6 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    ws2.cell(row=r, column=1, value=i+1).font = Font(size=10, color=NAVY, bold=True)
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws2.cell(row=r, column=1).fill = fill
    ws2.cell(row=r, column=1).border = thin_border
    
    for col in range(2, 10):
        cell = ws2.cell(row=r, column=col)
        cell.fill = fill
        cell.border = thin_border
        cell.font = Font(size=9, color="333333")
    
    ws2.cell(row=r, column=2).number_format = 'MM/DD/YYYY'
    ws2.cell(row=r, column=3).number_format = '0.00'
    ws2.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws2.cell(row=r, column=4).number_format = '0.00'
    ws2.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    # Total hours
    ws2.cell(row=r, column=5).value = f'=IF(OR(C{r}<>"",D{r}<>""),C{r}+D{r},"")'
    ws2.cell(row=r, column=5).number_format = '0.00'
    ws2.cell(row=r, column=5).font = Font(size=9, color=NAVY, bold=True)
    ws2.cell(row=r, column=5).alignment = Alignment(horizontal='center')
    ws2.cell(row=r, column=6).number_format = '$#,##0.00'
    ws2.cell(row=r, column=7).number_format = '$#,##0.00'
    # Total pay
    ws2.cell(row=r, column=8).value = f'=IF(OR(F{r}<>"",G{r}<>""),F{r}+G{r},"")'
    ws2.cell(row=r, column=8).number_format = '$#,##0.00'
    ws2.cell(row=r, column=8).font = Font(size=9, color=NAVY, bold=True)

# Monthly totals
r = 11
for col in range(1, 10):
    ws2.cell(row=r, column=col).fill = navy_fill
    ws2.cell(row=r, column=col).border = thin_border
ws2.cell(row=r, column=1, value="MONTHLY TOTALS").font = Font(color=WHITE, bold=True, size=10)
ws2.cell(row=r, column=1).fill = navy_fill
ws2.merge_cells(f'A{r}:B{r}')

for col in [3, 4, 5, 6, 7, 8]:
    ws2.cell(row=r, column=col).value = f'=SUM({get_column_letter(col)}6:{get_column_letter(col)}10)'
    ws2.cell(row=r, column=col).font = Font(color=WHITE, bold=True, size=11)
    ws2.cell(row=r, column=col).fill = navy_fill
    ws2.cell(row=r, column=col).alignment = Alignment(horizontal='center')
    if col <= 5:
        ws2.cell(row=r, column=col).number_format = '0.00'
    else:
        ws2.cell(row=r, column=col).number_format = '$#,##0.00'

ws2.print_area = 'A1:I11'
ws2.page_setup.orientation = 'landscape'

output_path = "/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/weekly-timesheet-template.xlsx"
wb.save(output_path)
print(f"✅ Template 5 saved: {output_path}")
