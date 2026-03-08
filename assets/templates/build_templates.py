#!/usr/bin/env python3
"""Build 3 professional contractor Excel templates."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from copy import copy

# ── Shared Style Constants ──
NAVY = "1A1A2E"
ORANGE = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
MEDIUM_GRAY = "E5E7EB"
DARK_TEXT = "1F2937"
GREEN = "22C55E"
RED = "EF4444"
LIGHT_GREEN = "DCFCE7"
LIGHT_RED = "FEE2E2"

navy_fill = PatternFill("solid", fgColor=NAVY)
orange_fill = PatternFill("solid", fgColor=ORANGE)
light_gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
white_fill = PatternFill("solid", fgColor=WHITE)
light_green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
light_red_fill = PatternFill("solid", fgColor=LIGHT_RED)

white_bold = Font(name="Calibri", bold=True, color=WHITE, size=11)
white_bold_14 = Font(name="Calibri", bold=True, color=WHITE, size=14)
white_bold_12 = Font(name="Calibri", bold=True, color=WHITE, size=12)
navy_bold_20 = Font(name="Calibri", bold=True, color=NAVY, size=20)
navy_bold_16 = Font(name="Calibri", bold=True, color=NAVY, size=16)
navy_bold_14 = Font(name="Calibri", bold=True, color=NAVY, size=14)
navy_bold_12 = Font(name="Calibri", bold=True, color=NAVY, size=12)
navy_bold = Font(name="Calibri", bold=True, color=NAVY, size=11)
orange_bold_14 = Font(name="Calibri", bold=True, color=ORANGE, size=14)
orange_bold_12 = Font(name="Calibri", bold=True, color=ORANGE, size=12)
orange_bold = Font(name="Calibri", bold=True, color=ORANGE, size=11)
dark_font = Font(name="Calibri", color=DARK_TEXT, size=11)
dark_bold = Font(name="Calibri", bold=True, color=DARK_TEXT, size=11)
dark_font_10 = Font(name="Calibri", color=DARK_TEXT, size=10)
label_font = Font(name="Calibri", color="6B7280", size=10)

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)
bottom_border = Border(bottom=Side(style="thin", color=MEDIUM_GRAY))
thick_bottom = Border(bottom=Side(style="medium", color=NAVY))

center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrapText=True)

USD = '"$"#,##0.00'
PCT = '0.0%'
DATE_FMT = 'MM/DD/YYYY'


def apply_table_header(ws, row, cols, font=None):
    """Apply navy header style to a row of cells."""
    f = font or white_bold
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = navy_fill
        cell.font = f
        cell.alignment = center
        cell.border = thin_border


def apply_alt_rows(ws, start_row, end_row, col_start, col_end):
    """Apply alternating row fills."""
    for r in range(start_row, end_row + 1):
        fill = light_gray_fill if (r - start_row) % 2 == 0 else white_fill
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
            cell.alignment = left_align


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════
# TEMPLATE 1: CONTRACTOR INVOICE
# ═══════════════════════════════════════════════════════════════
def build_invoice():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Page setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    set_col_widths(ws, {"A": 3, "B": 40, "C": 12, "D": 14, "E": 16, "F": 18, "G": 3})

    # ── Orange accent bar at top ──
    for c in range(1, 8):
        cell = ws.cell(row=1, column=c)
        cell.fill = orange_fill
        cell.value = ""
    ws.row_dimensions[1].height = 6

    # ── Company branding ──
    ws.merge_cells("B2:F2")
    ws.cell(row=2, column=2, value="YOUR COMPANY NAME").font = navy_bold_20
    ws.cell(row=2, column=2).alignment = left_align
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:D3")
    ws.cell(row=3, column=2, value="123 Business Street, City, State ZIP").font = label_font
    ws.cell(row=3, column=2).alignment = left_align

    ws.merge_cells("B4:D4")
    ws.cell(row=4, column=2, value="Phone: (555) 000-0000  |  Email: info@yourcompany.com").font = label_font
    ws.cell(row=4, column=2).alignment = left_align

    ws.merge_cells("B5:D5")
    ws.cell(row=5, column=2, value="License #: _______________  |  Tax ID: _______________").font = label_font
    ws.cell(row=5, column=2).alignment = left_align

    # ── INVOICE title ──
    ws.merge_cells("E3:F3")
    ws.cell(row=3, column=5, value="INVOICE").font = Font(name="Calibri", bold=True, color=NAVY, size=28)
    ws.cell(row=3, column=5).alignment = Alignment(horizontal="right", vertical="center")

    # Invoice details
    for r, (lbl, val) in enumerate([
        ("Invoice #:", "INV-001"),
        ("Date:", ""),
        ("Due Date:", ""),
    ], start=4):
        ws.cell(row=r, column=5, value=lbl).font = dark_bold
        ws.cell(row=r, column=5).alignment = right_align
        ws.cell(row=r, column=6, value=val).font = dark_font
        ws.cell(row=r, column=6).alignment = left_align
        ws.cell(row=r, column=6).number_format = DATE_FMT

    # ── Separator ──
    for c in range(2, 7):
        ws.cell(row=7, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))
    ws.row_dimensions[7].height = 8

    # ── Bill To / Project ──
    ws.merge_cells("B8:C8")
    ws.cell(row=8, column=2, value="BILL TO").font = orange_bold_12
    ws.cell(row=8, column=2).alignment = left_align
    ws.merge_cells("E8:F8")
    ws.cell(row=8, column=5, value="PROJECT DETAILS").font = orange_bold_12
    ws.cell(row=8, column=5).alignment = left_align

    bill_labels = ["Client Name:", "Company:", "Address:", "City, State ZIP:", "Phone:", "Email:"]
    proj_labels = ["Project Name:", "Location:", "PO Number:", "Start Date:", "End Date:", ""]
    for i, (bl, pl) in enumerate(zip(bill_labels, proj_labels)):
        r = 9 + i
        ws.cell(row=r, column=2, value=bl).font = label_font
        ws.cell(row=r, column=3).font = dark_font
        ws.cell(row=r, column=3).border = bottom_border
        if pl:
            ws.cell(row=r, column=5, value=pl).font = label_font
            ws.cell(row=r, column=6).font = dark_font
            ws.cell(row=r, column=6).border = bottom_border

    # ── Line Items Table ──
    header_row = 16
    headers = ["Description", "Quantity", "Unit", "Unit Price", "Amount"]
    cols = [2, 3, 4, 5, 6]
    for c, h in zip(cols, headers):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 28

    # Data rows (20 rows)
    data_start = header_row + 1
    data_end = data_start + 19
    for r in range(data_start, data_end + 1):
        fill = light_gray_fill if (r - data_start) % 2 == 0 else white_fill
        for c in cols:
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws.cell(row=r, column=2).alignment = wrap  # Description
        ws.cell(row=r, column=3).alignment = center  # Qty
        ws.cell(row=r, column=4).alignment = center  # Unit
        ws.cell(row=r, column=5).number_format = USD  # Unit Price
        ws.cell(row=r, column=5).alignment = right_align
        # Amount formula
        ws.cell(row=r, column=6).value = f'=IF(C{r}="","",C{r}*E{r})'
        ws.cell(row=r, column=6).number_format = USD
        ws.cell(row=r, column=6).alignment = right_align
        ws.row_dimensions[r].height = 22

    # Data validation for Unit column
    dv = DataValidation(
        type="list",
        formula1='"Hours,Days,Sq Ft,Linear Ft,Each,Lot,Job"',
        allow_blank=True
    )
    dv.error = "Please select a valid unit"
    dv.errorTitle = "Invalid Unit"
    dv.prompt = "Select unit type"
    dv.promptTitle = "Unit"
    ws.add_data_validation(dv)
    dv.add(f"D{data_start}:D{data_end}")

    # ── Totals section ──
    totals_start = data_end + 1
    # Subtotal
    for r_offset, (lbl, formula, is_orange) in enumerate([
        ("Subtotal", f'=SUM(F{data_start}:F{data_end})', False),
        ("Tax Rate", "", False),
        ("Tax Amount", f'=F{totals_start}*F{totals_start+1}', False),
    ]):
        r = totals_start + r_offset
        ws.merge_cells(f"D{r}:E{r}")
        ws.cell(row=r, column=4, value=lbl).font = dark_bold
        ws.cell(row=r, column=4).alignment = right_align
        if formula:
            ws.cell(row=r, column=6, value=formula).number_format = USD
        else:
            ws.cell(row=r, column=6, value=0).number_format = PCT
        ws.cell(row=r, column=6).font = dark_bold
        ws.cell(row=r, column=6).alignment = right_align
        ws.cell(row=r, column=6).border = thin_border

    # TOTAL row
    total_row = totals_start + 3
    ws.merge_cells(f"D{total_row}:E{total_row}")
    ws.cell(row=total_row, column=4, value="TOTAL DUE").font = white_bold_14
    ws.cell(row=total_row, column=4).fill = navy_fill
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=6).value = f'=F{totals_start}+F{totals_start+2}'
    ws.cell(row=total_row, column=6).number_format = USD
    ws.cell(row=total_row, column=6).font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws.cell(row=total_row, column=6).fill = navy_fill
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=6).border = thin_border
    ws.row_dimensions[total_row].height = 32

    # ── Payment Terms ──
    pay_row = total_row + 2
    ws.merge_cells(f"B{pay_row}:C{pay_row}")
    ws.cell(row=pay_row, column=2, value="PAYMENT TERMS").font = orange_bold_12
    ws.merge_cells(f"E{pay_row}:F{pay_row}")
    ws.cell(row=pay_row, column=5, value="PAYMENT METHODS").font = orange_bold_12

    terms = [
        "Payment due within 30 days of invoice date.",
        "Late payments subject to 1.5% monthly interest.",
        "Make checks payable to: YOUR COMPANY NAME",
    ]
    methods = [
        "☐ Check    ☐ Cash    ☐ Credit Card",
        "☐ Bank Transfer    ☐ Zelle    ☐ Venmo",
        "Account details provided upon request.",
    ]
    for i, (t, m) in enumerate(zip(terms, methods)):
        r = pay_row + 1 + i
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2, value=t).font = dark_font_10
        ws.merge_cells(f"E{r}:F{r}")
        ws.cell(row=r, column=5, value=m).font = dark_font_10

    # ── Thank You ──
    ty_row = pay_row + 5
    ws.merge_cells(f"B{ty_row}:F{ty_row}")
    ws.cell(row=ty_row, column=2, value="Thank you for your business!").font = Font(
        name="Calibri", italic=True, color=NAVY, size=12
    )
    ws.cell(row=ty_row, column=2).alignment = center

    # Bottom accent bar
    for c in range(1, 8):
        ws.cell(row=ty_row + 1, column=c).fill = orange_fill
    ws.row_dimensions[ty_row + 1].height = 6

    # Print area
    ws.print_area = f"A1:G{ty_row + 1}"

    # ═══ Payment Log Sheet ═══
    ws2 = wb.create_sheet("Payment Log")
    set_col_widths(ws2, {"A": 3, "B": 16, "C": 16, "D": 18, "E": 22, "F": 20, "G": 3})

    # Title
    ws2.merge_cells("B1:F1")
    ws2.cell(row=1, column=2).fill = orange_fill
    ws2.row_dimensions[1].height = 6

    ws2.merge_cells("B2:F2")
    ws2.cell(row=2, column=2, value="PAYMENT LOG").font = navy_bold_16
    ws2.cell(row=2, column=2).alignment = left_align
    ws2.row_dimensions[2].height = 30

    ws2.cell(row=3, column=2, value="Invoice #:").font = label_font
    ws2.cell(row=3, column=3).font = dark_font
    ws2.cell(row=3, column=3).border = bottom_border
    ws2.cell(row=3, column=4, value="Total Due:").font = label_font
    ws2.cell(row=3, column=5).font = dark_bold
    ws2.cell(row=3, column=5).number_format = USD
    ws2.cell(row=3, column=5).border = bottom_border

    # Headers
    pay_headers = ["Date", "Amount", "Method", "Check/Reference #", "Balance Remaining"]
    pay_cols = [2, 3, 4, 5, 6]
    hr = 5
    for c, h in zip(pay_cols, pay_headers):
        cell = ws2.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws2.row_dimensions[hr].height = 28

    for r in range(hr + 1, hr + 21):
        fill = light_gray_fill if (r - hr - 1) % 2 == 0 else white_fill
        for c in pay_cols:
            cell = ws2.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws2.cell(row=r, column=2).number_format = DATE_FMT
        ws2.cell(row=r, column=3).number_format = USD
        # Balance = previous balance - this amount (first row references Total Due)
        if r == hr + 1:
            ws2.cell(row=r, column=6).value = f'=IF(C{r}="","",E3-C{r})'
        else:
            ws2.cell(row=r, column=6).value = f'=IF(C{r}="","",F{r-1}-C{r})'
        ws2.cell(row=r, column=6).number_format = USD

    # Payment method validation
    dv2 = DataValidation(type="list", formula1='"Check,Cash,Credit Card,Bank Transfer,Zelle,Venmo,Other"', allow_blank=True)
    ws2.add_data_validation(dv2)
    dv2.add(f"D{hr+1}:D{hr+20}")

    wb.save("/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/contractor-invoice-template.xlsx")
    print("✅ Invoice template built")


# ═══════════════════════════════════════════════════════════════
# TEMPLATE 2: JOB ESTIMATE
# ═══════════════════════════════════════════════════════════════
def build_estimate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    set_col_widths(ws, {"A": 3, "B": 36, "C": 12, "D": 14, "E": 16, "F": 18, "G": 3})

    # ── Top accent ──
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = orange_fill
    ws.row_dimensions[1].height = 6

    # ── Company branding ──
    ws.merge_cells("B2:D2")
    ws.cell(row=2, column=2, value="YOUR COMPANY NAME").font = navy_bold_20
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:D3")
    ws.cell(row=3, column=2, value="123 Business Street, City, State ZIP").font = label_font

    ws.merge_cells("B4:D4")
    ws.cell(row=4, column=2, value="Phone: (555) 000-0000  |  Email: info@yourcompany.com").font = label_font

    ws.merge_cells("B5:D5")
    ws.cell(row=5, column=2, value="License #: _______________").font = label_font

    # Estimate title
    ws.merge_cells("E2:F2")
    ws.cell(row=2, column=5, value="ESTIMATE").font = Font(name="Calibri", bold=True, color=NAVY, size=28)
    ws.cell(row=2, column=5).alignment = Alignment(horizontal="right", vertical="center")

    est_info = [("Estimate #:", "EST-001"), ("Date:", ""), ("Valid Until:", "")]
    for i, (lbl, val) in enumerate(est_info):
        r = 3 + i
        ws.cell(row=r, column=5, value=lbl).font = dark_bold
        ws.cell(row=r, column=5).alignment = right_align
        ws.cell(row=r, column=6, value=val).font = dark_font
        ws.cell(row=r, column=6).number_format = DATE_FMT

    # Separator
    for c in range(2, 7):
        ws.cell(row=6, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))
    ws.row_dimensions[6].height = 8

    # Client / Project info
    ws.merge_cells("B7:C7")
    ws.cell(row=7, column=2, value="CLIENT INFORMATION").font = orange_bold_12

    ws.merge_cells("E7:F7")
    ws.cell(row=7, column=5, value="PROJECT INFORMATION").font = orange_bold_12

    client_labels = ["Client Name:", "Company:", "Address:", "Phone / Email:"]
    proj_labels = ["Project Name:", "Project Address:", "Scope of Work:", ""]
    for i, (cl, pl) in enumerate(zip(client_labels, proj_labels)):
        r = 8 + i
        ws.cell(row=r, column=2, value=cl).font = label_font
        ws.cell(row=r, column=3).border = bottom_border
        if pl:
            ws.cell(row=r, column=5, value=pl).font = label_font
            ws.cell(row=r, column=6).border = bottom_border

    # ── Cost Categories ──
    categories = ["LABOR", "MATERIALS", "EQUIPMENT", "OTHER COSTS"]
    cat_rows = 10  # rows per category
    headers = ["Description", "Qty", "Unit", "Unit Cost", "Total"]
    cols = [2, 3, 4, 5, 6]

    current_row = 13
    subtotal_cells = []

    for cat in categories:
        # Category header
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws.cell(row=current_row, column=2, value=cat).font = white_bold_12
        ws.cell(row=current_row, column=2).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=current_row, column=2).alignment = left_align
        for c in range(3, 7):
            ws.cell(row=current_row, column=c).fill = navy_fill
        ws.row_dimensions[current_row].height = 26
        current_row += 1

        # Column sub-headers
        for c, h in zip(cols, headers):
            cell = ws.cell(row=current_row, column=c, value=h)
            cell.fill = PatternFill("solid", fgColor="374151")
            cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
            cell.alignment = center
            cell.border = thin_border
        current_row += 1

        data_start = current_row
        for r in range(current_row, current_row + cat_rows):
            fill = light_gray_fill if (r - current_row) % 2 == 0 else white_fill
            for c in cols:
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin_border
                cell.font = dark_font
            ws.cell(row=r, column=3).alignment = center
            ws.cell(row=r, column=4).alignment = center
            ws.cell(row=r, column=5).number_format = USD
            ws.cell(row=r, column=5).alignment = right_align
            ws.cell(row=r, column=6).value = f'=IF(C{r}="","",C{r}*E{r})'
            ws.cell(row=r, column=6).number_format = USD
            ws.cell(row=r, column=6).alignment = right_align
        current_row += cat_rows

        # Subtotal row
        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws.cell(row=current_row, column=2, value=f"{cat} Subtotal").font = dark_bold
        ws.cell(row=current_row, column=2).alignment = right_align
        ws.cell(row=current_row, column=6).value = f'=SUM(F{data_start}:F{current_row - 1})'
        ws.cell(row=current_row, column=6).number_format = USD
        ws.cell(row=current_row, column=6).font = dark_bold
        ws.cell(row=current_row, column=6).alignment = right_align
        ws.cell(row=current_row, column=6).border = Border(top=Side(style="medium", color=NAVY), bottom=Side(style="thin", color=MEDIUM_GRAY))
        subtotal_cells.append(f"F{current_row}")
        current_row += 2  # blank row between categories

    # ── Summary Section ──
    ws.merge_cells(f"B{current_row}:F{current_row}")
    ws.cell(row=current_row, column=2, value="ESTIMATE SUMMARY").font = orange_bold_14
    for c in range(2, 7):
        ws.cell(row=current_row, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))
    current_row += 1

    sum_formula = "+".join(subtotal_cells)
    summary_items = [
        ("Direct Costs Subtotal", f"={sum_formula}"),
        ("Overhead", None),
        ("Overhead Amount", None),
        ("Contingency", None),
        ("Contingency Amount", None),
        ("Tax Rate", None),
        ("Tax Amount", None),
    ]

    summary_start = current_row
    for i, (lbl, formula) in enumerate(summary_items):
        r = current_row + i
        ws.merge_cells(f"D{r}:E{r}")
        ws.cell(row=r, column=4, value=lbl).font = dark_bold
        ws.cell(row=r, column=4).alignment = right_align
        ws.cell(row=r, column=6).border = thin_border
        ws.cell(row=r, column=6).alignment = right_align

    # Direct Costs
    ws.cell(row=summary_start, column=6).value = f"={sum_formula}"
    ws.cell(row=summary_start, column=6).number_format = USD
    ws.cell(row=summary_start, column=6).font = dark_bold

    # Overhead % (configurable)
    ws.cell(row=summary_start + 1, column=6, value=0.20).number_format = PCT
    ws.cell(row=summary_start + 1, column=6).font = orange_bold
    # Overhead Amount
    ws.cell(row=summary_start + 2, column=6).value = f'=F{summary_start}*F{summary_start+1}'
    ws.cell(row=summary_start + 2, column=6).number_format = USD

    # Contingency % (configurable)
    ws.cell(row=summary_start + 3, column=6, value=0.10).number_format = PCT
    ws.cell(row=summary_start + 3, column=6).font = orange_bold
    # Contingency Amount
    ws.cell(row=summary_start + 4, column=6).value = f'=F{summary_start}*F{summary_start+3}'
    ws.cell(row=summary_start + 4, column=6).number_format = USD

    # Tax Rate
    ws.cell(row=summary_start + 5, column=6, value=0).number_format = PCT
    ws.cell(row=summary_start + 5, column=6).font = orange_bold
    # Tax Amount
    ws.cell(row=summary_start + 6, column=6).value = f'=(F{summary_start}+F{summary_start+2}+F{summary_start+4})*F{summary_start+5}'
    ws.cell(row=summary_start + 6, column=6).number_format = USD

    # Grand Total
    gt_row = summary_start + 8
    ws.merge_cells(f"D{gt_row}:E{gt_row}")
    ws.cell(row=gt_row, column=4, value="GRAND TOTAL").font = white_bold_14
    ws.cell(row=gt_row, column=4).fill = navy_fill
    ws.cell(row=gt_row, column=4).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=gt_row, column=6).value = f'=F{summary_start}+F{summary_start+2}+F{summary_start+4}+F{summary_start+6}'
    ws.cell(row=gt_row, column=6).number_format = USD
    ws.cell(row=gt_row, column=6).font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws.cell(row=gt_row, column=6).fill = navy_fill
    ws.cell(row=gt_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[gt_row].height = 32

    # ── Timeline ──
    tl_row = gt_row + 2
    ws.merge_cells(f"B{tl_row}:F{tl_row}")
    ws.cell(row=tl_row, column=2, value="PROJECT TIMELINE").font = orange_bold_12
    for c in range(2, 7):
        ws.cell(row=tl_row, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

    tl_headers = ["Phase", "Description", "", "Start", "End"]
    tl_cols = [2, 3, 4, 5, 6]
    tl_row += 1
    for c, h in zip(tl_cols, tl_headers):
        cell = ws.cell(row=tl_row, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border

    for r in range(tl_row + 1, tl_row + 6):
        fill = light_gray_fill if (r - tl_row) % 2 == 0 else white_fill
        for c in tl_cols:
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws.cell(row=r, column=5).number_format = DATE_FMT
        ws.cell(row=r, column=6).number_format = DATE_FMT

    # ── Terms & Conditions ──
    terms_row = tl_row + 7
    ws.merge_cells(f"B{terms_row}:F{terms_row}")
    ws.cell(row=terms_row, column=2, value="TERMS & CONDITIONS").font = orange_bold_12
    for c in range(2, 7):
        ws.cell(row=terms_row, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

    terms = [
        "1. This estimate is valid for 30 days from the date issued.",
        "2. A 50% deposit is required before work begins.",
        "3. Final payment is due upon completion of work.",
        "4. Any changes to the scope of work may result in additional charges.",
        "5. All work is guaranteed for 1 year from completion date.",
    ]
    for i, t in enumerate(terms):
        r = terms_row + 1 + i
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=t).font = dark_font_10

    # ── Signature ──
    sig_row = terms_row + len(terms) + 2
    ws.merge_cells(f"B{sig_row}:C{sig_row}")
    ws.cell(row=sig_row, column=2, value="Contractor Signature:").font = dark_bold
    ws.cell(row=sig_row, column=2).alignment = left_align
    ws.merge_cells(f"E{sig_row}:F{sig_row}")
    ws.cell(row=sig_row, column=5, value="Client Signature:").font = dark_bold
    ws.cell(row=sig_row, column=5).alignment = left_align

    sig_row += 1
    for c_start, c_end in [(2, 3), (5, 6)]:
        ws.merge_cells(f"{get_column_letter(c_start)}{sig_row}:{get_column_letter(c_end)}{sig_row}")
        ws.cell(row=sig_row, column=c_start).border = Border(bottom=Side(style="thin", color=DARK_TEXT))
    ws.row_dimensions[sig_row].height = 30

    sig_row += 1
    ws.cell(row=sig_row, column=2, value="Date: _______________").font = label_font
    ws.cell(row=sig_row, column=5, value="Date: _______________").font = label_font

    # Bottom accent
    sig_row += 1
    for c in range(1, 8):
        ws.cell(row=sig_row, column=c).fill = orange_fill
    ws.row_dimensions[sig_row].height = 6

    ws.print_area = f"A1:G{sig_row}"

    # ═══ Estimate vs Actual Sheet ═══
    ws2 = wb.create_sheet("Estimate vs Actual")
    set_col_widths(ws2, {"A": 3, "B": 30, "C": 16, "D": 16, "E": 16, "F": 14, "G": 3})

    ws2.merge_cells("B1:F1")
    ws2.cell(row=1, column=2).fill = orange_fill
    ws2.row_dimensions[1].height = 6

    ws2.merge_cells("B2:F2")
    ws2.cell(row=2, column=2, value="ESTIMATE vs ACTUAL COMPARISON").font = navy_bold_16
    ws2.row_dimensions[2].height = 30

    ws2.cell(row=3, column=2, value="Project:").font = label_font
    ws2.cell(row=3, column=3).border = bottom_border
    ws2.cell(row=3, column=4, value="Estimate #:").font = label_font
    ws2.cell(row=3, column=5).border = bottom_border

    # Headers
    eva_headers = ["Category / Item", "Estimated", "Actual", "Variance ($)", "Variance (%)"]
    eva_cols = [2, 3, 4, 5, 6]
    hr = 5
    for c, h in zip(eva_cols, eva_headers):
        cell = ws2.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws2.row_dimensions[hr].height = 28

    current = hr + 1
    eva_subtotal_rows = []
    for cat in categories:
        # Category header row
        ws2.merge_cells(f"B{current}:F{current}")
        ws2.cell(row=current, column=2, value=cat).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        for c in eva_cols:
            ws2.cell(row=current, column=c).fill = PatternFill("solid", fgColor="374151")
        current += 1

        data_start = current
        for r in range(current, current + cat_rows):
            fill = light_gray_fill if (r - current) % 2 == 0 else white_fill
            for c in eva_cols:
                cell = ws2.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin_border
                cell.font = dark_font
            ws2.cell(row=r, column=3).number_format = USD
            ws2.cell(row=r, column=4).number_format = USD
            ws2.cell(row=r, column=5).value = f'=IF(C{r}="","",C{r}-D{r})'
            ws2.cell(row=r, column=5).number_format = USD
            ws2.cell(row=r, column=6).value = f'=IF(C{r}="","",IF(C{r}=0,"",(C{r}-D{r})/C{r}))'
            ws2.cell(row=r, column=6).number_format = PCT
        current += cat_rows

        # Subtotal
        ws2.cell(row=current, column=2, value=f"{cat} Subtotal").font = dark_bold
        ws2.cell(row=current, column=2).alignment = right_align
        ws2.cell(row=current, column=3).value = f'=SUM(C{data_start}:C{current-1})'
        ws2.cell(row=current, column=3).number_format = USD
        ws2.cell(row=current, column=3).font = dark_bold
        ws2.cell(row=current, column=4).value = f'=SUM(D{data_start}:D{current-1})'
        ws2.cell(row=current, column=4).number_format = USD
        ws2.cell(row=current, column=4).font = dark_bold
        ws2.cell(row=current, column=5).value = f'=C{current}-D{current}'
        ws2.cell(row=current, column=5).number_format = USD
        ws2.cell(row=current, column=5).font = dark_bold
        ws2.cell(row=current, column=6).value = f'=IF(C{current}=0,"",C{current}-D{current})/C{current}'
        ws2.cell(row=current, column=6).number_format = PCT
        ws2.cell(row=current, column=6).font = dark_bold
        for c in eva_cols:
            ws2.cell(row=current, column=c).border = Border(top=Side(style="medium", color=NAVY))
        eva_subtotal_rows.append(current)
        current += 2

    # Grand total
    ws2.merge_cells(f"B{current}:B{current}")
    ws2.cell(row=current, column=2, value="GRAND TOTAL").font = white_bold_12
    ws2.cell(row=current, column=2).fill = navy_fill
    ws2.cell(row=current, column=2).alignment = right_align
    for c_idx, c in enumerate([3, 4, 5, 6]):
        ws2.cell(row=current, column=c).fill = navy_fill
        ws2.cell(row=current, column=c).font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        ws2.cell(row=current, column=c).border = thin_border
    c3_refs = "+".join([f"C{r}" for r in eva_subtotal_rows])
    d3_refs = "+".join([f"D{r}" for r in eva_subtotal_rows])
    ws2.cell(row=current, column=3).value = f"={c3_refs}"
    ws2.cell(row=current, column=3).number_format = USD
    ws2.cell(row=current, column=4).value = f"={d3_refs}"
    ws2.cell(row=current, column=4).number_format = USD
    ws2.cell(row=current, column=5).value = f"=C{current}-D{current}"
    ws2.cell(row=current, column=5).number_format = USD
    ws2.cell(row=current, column=6).value = f'=IF(C{current}=0,"",C{current}-D{current})/C{current}'
    ws2.cell(row=current, column=6).number_format = PCT
    ws2.row_dimensions[current].height = 30

    # Conditional formatting for variance columns
    green_font = Font(color=GREEN)
    red_font = Font(color=RED)
    ws2.conditional_formatting.add(f"E6:E{current}",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="22C55E", bold=True), fill=light_green_fill))
    ws2.conditional_formatting.add(f"E6:E{current}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="EF4444", bold=True), fill=light_red_fill))

    wb.save("/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/job-estimate-template.xlsx")
    print("✅ Estimate template built")


# ═══════════════════════════════════════════════════════════════
# TEMPLATE 3: JOB COSTING TRACKER
# ═══════════════════════════════════════════════════════════════
def build_job_costing():
    wb = openpyxl.Workbook()

    # ═══ SHEET 1: DASHBOARD ═══
    ws = wb.active
    ws.title = "Dashboard"
    set_col_widths(ws, {"A": 3, "B": 22, "C": 18, "D": 18, "E": 18, "F": 16, "G": 18, "H": 3})

    # Top accent
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = orange_fill
    ws.row_dimensions[1].height = 6

    ws.merge_cells("B2:G2")
    ws.cell(row=2, column=2, value="JOB COSTING DASHBOARD").font = navy_bold_20
    ws.row_dimensions[2].height = 36

    # Job info
    info_fields = [
        ("Job Name:", "B3", "C3"), ("Job #:", "E3", "F3"),
        ("Client:", "B4", "C4"), ("Start Date:", "E4", "F4"),
        ("Address:", "B5", "C5"), ("End Date:", "E5", "F5"),
        ("Project Manager:", "B6", "C6"), ("Contract Amount:", "E6", "F6"),
    ]
    for lbl, lbl_cell, val_cell in info_fields:
        ws[lbl_cell] = lbl
        ws[lbl_cell].font = label_font
        ws[val_cell].font = dark_bold
        ws[val_cell].border = bottom_border
    ws["F4"].number_format = DATE_FMT
    ws["F5"].number_format = DATE_FMT
    ws["F6"].number_format = USD

    # Separator
    for c in range(2, 8):
        ws.cell(row=7, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))
    ws.row_dimensions[7].height = 8

    # Summary Table
    ws.merge_cells("B8:G8")
    ws.cell(row=8, column=2, value="COST SUMMARY").font = orange_bold_14
    ws.row_dimensions[8].height = 26

    sum_headers = ["Category", "Estimated", "Actual", "Variance", "% of Budget", "Status"]
    sum_cols = list(range(2, 8))
    hr = 9
    for c, h in zip(sum_cols, sum_headers):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[hr].height = 28

    # Category rows - referencing other sheets
    cats = [
        ("Labor", "Labor"),
        ("Materials", "Materials"),
        ("Subcontractors", "Subs & Equipment"),
        ("Equipment", "Subs & Equipment"),
        ("Change Orders", "Change Orders"),
    ]

    for i, (cat_name, sheet_name) in enumerate(cats):
        r = hr + 1 + i
        fill = light_gray_fill if i % 2 == 0 else white_fill
        for c in sum_cols:
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = dark_font
        ws.cell(row=r, column=2, value=cat_name).font = dark_bold
        ws.cell(row=r, column=2).alignment = left_align

        # Estimated column - user enters manually on dashboard
        ws.cell(row=r, column=3).number_format = USD
        ws.cell(row=r, column=3).alignment = right_align

        # Actual column - pull from sheets
        if cat_name == "Labor":
            ws.cell(row=r, column=4).value = "=Labor!F53"
        elif cat_name == "Materials":
            ws.cell(row=r, column=4).value = "=Materials!F53"
        elif cat_name == "Subcontractors":
            ws.cell(row=r, column=4).value = "='Subs & Equipment'!F13"
        elif cat_name == "Equipment":
            ws.cell(row=r, column=4).value = "='Subs & Equipment'!L13"
        elif cat_name == "Change Orders":
            ws.cell(row=r, column=4).value = "='Change Orders'!D23"
        ws.cell(row=r, column=4).number_format = USD
        ws.cell(row=r, column=4).alignment = right_align

        # Variance
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        ws.cell(row=r, column=5).number_format = USD
        ws.cell(row=r, column=5).alignment = right_align

        # % of Budget
        ws.cell(row=r, column=6).value = f'=IF(C{r}=0,"",D{r}/C{r})'
        ws.cell(row=r, column=6).number_format = PCT
        ws.cell(row=r, column=6).alignment = center

        # Status
        ws.cell(row=r, column=7).value = f'=IF(C{r}=0,"",IF(E{r}>=0,"✓ Under Budget","⚠ Over Budget"))'
        ws.cell(row=r, column=7).alignment = center

    # Total row
    total_r = hr + 1 + len(cats)
    ws.cell(row=total_r, column=2, value="TOTAL").font = white_bold_12
    ws.cell(row=total_r, column=2).fill = navy_fill
    ws.cell(row=total_r, column=2).alignment = left_align
    for c in sum_cols[1:]:
        ws.cell(row=total_r, column=c).fill = navy_fill
        ws.cell(row=total_r, column=c).font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        ws.cell(row=total_r, column=c).border = thin_border
    ws.cell(row=total_r, column=3).value = f"=SUM(C{hr+1}:C{total_r-1})"
    ws.cell(row=total_r, column=3).number_format = USD
    ws.cell(row=total_r, column=4).value = f"=SUM(D{hr+1}:D{total_r-1})"
    ws.cell(row=total_r, column=4).number_format = USD
    ws.cell(row=total_r, column=5).value = f"=C{total_r}-D{total_r}"
    ws.cell(row=total_r, column=5).number_format = USD
    ws.cell(row=total_r, column=6).value = f'=IF(C{total_r}=0,"",D{total_r}/C{total_r})'
    ws.cell(row=total_r, column=6).number_format = PCT
    ws.row_dimensions[total_r].height = 30

    # Conditional formatting on Variance column
    var_range = f"E{hr+1}:E{total_r}"
    ws.conditional_formatting.add(var_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=Font(color="22C55E", bold=True), fill=light_green_fill))
    ws.conditional_formatting.add(var_range,
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="EF4444", bold=True), fill=light_red_fill))

    # Profit/Loss section
    pl_row = total_r + 2
    ws.merge_cells(f"B{pl_row}:G{pl_row}")
    ws.cell(row=pl_row, column=2, value="PROFIT / LOSS ANALYSIS").font = orange_bold_14
    for c in range(2, 8):
        ws.cell(row=pl_row, column=c).border = Border(bottom=Side(style="medium", color=ORANGE))

    pl_items = [
        ("Contract Amount", "=F6", USD),
        ("Total Costs", f"=D{total_r}", USD),
        ("Gross Profit", f"=E{pl_row+1}-E{pl_row+2}", USD),
        ("Profit Margin", f'=IF(E{pl_row+1}=0,"",E{pl_row+3}/E{pl_row+1})', PCT),
    ]
    for i, (lbl, formula, fmt) in enumerate(pl_items):
        r = pl_row + 1 + i
        ws.merge_cells(f"C{r}:D{r}")
        ws.cell(row=r, column=3, value=lbl).font = dark_bold
        ws.cell(row=r, column=3).alignment = right_align
        ws.cell(row=r, column=5, value=formula).number_format = fmt
        ws.cell(row=r, column=5).font = dark_bold if i < 3 else navy_bold_14
        ws.cell(row=r, column=5).alignment = right_align
        ws.cell(row=r, column=5).border = thin_border

    # Conditional formatting on profit
    profit_cell = f"E{pl_row+3}"
    ws.conditional_formatting.add(profit_cell,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=Font(color="22C55E", bold=True, size=14), fill=light_green_fill))
    ws.conditional_formatting.add(profit_cell,
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="EF4444", bold=True, size=14), fill=light_red_fill))

    # ═══ SHEET 2: LABOR ═══
    ws2 = wb.create_sheet("Labor")
    set_col_widths(ws2, {"A": 3, "B": 14, "C": 18, "D": 16, "E": 10, "F": 12, "G": 14, "H": 20, "I": 20, "J": 3})

    for c in range(1, 11):
        ws2.cell(row=1, column=c).fill = orange_fill
    ws2.row_dimensions[1].height = 6

    ws2.merge_cells("B2:I2")
    ws2.cell(row=2, column=2, value="LABOR TRACKING").font = navy_bold_16
    ws2.row_dimensions[2].height = 30

    labor_headers = ["Date", "Employee", "Role", "Hours", "Rate", "Total", "Category", "Notes"]
    labor_cols = list(range(2, 10))
    hr = 3
    for c, h in zip(labor_cols, labor_headers):
        cell = ws2.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws2.row_dimensions[hr].height = 28

    for r in range(hr + 1, hr + 51):
        fill = light_gray_fill if (r - hr - 1) % 2 == 0 else white_fill
        for c in labor_cols:
            cell = ws2.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws2.cell(row=r, column=2).number_format = DATE_FMT  # Date
        ws2.cell(row=r, column=2).alignment = center
        ws2.cell(row=r, column=5).alignment = center  # Hours
        ws2.cell(row=r, column=6).number_format = USD  # Rate
        ws2.cell(row=r, column=6).alignment = right_align
        ws2.cell(row=r, column=7).value = f'=IF(E{r}="","",E{r}*F{r})'  # Total
        ws2.cell(row=r, column=7).number_format = USD
        ws2.cell(row=r, column=7).alignment = right_align
        ws2.cell(row=r, column=9).alignment = wrap  # Notes

    # Category dropdown
    dv_labor = DataValidation(type="list", formula1='"Demo,Rough-in,Finish,Punch list,Travel"', allow_blank=True)
    dv_labor.prompt = "Select category"
    ws2.add_data_validation(dv_labor)
    dv_labor.add(f"H{hr+1}:H{hr+50}")

    # Running total
    rt_row = hr + 51
    ws2.merge_cells(f"B{rt_row}:F{rt_row}")
    ws2.cell(row=rt_row, column=2, value="TOTAL LABOR COST").font = white_bold_12
    ws2.cell(row=rt_row, column=2).fill = navy_fill
    ws2.cell(row=rt_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(3, 8):
        ws2.cell(row=rt_row, column=c).fill = navy_fill
    ws2.cell(row=rt_row, column=7).value = f"=SUM(G{hr+1}:G{hr+50})"
    ws2.cell(row=rt_row, column=7).number_format = USD
    ws2.cell(row=rt_row, column=7).font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws2.cell(row=rt_row, column=7).fill = navy_fill
    ws2.cell(row=rt_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws2.row_dimensions[rt_row].height = 30

    # ═══ SHEET 3: MATERIALS ═══
    ws3 = wb.create_sheet("Materials")
    set_col_widths(ws3, {"A": 3, "B": 14, "C": 18, "D": 22, "E": 10, "F": 12, "G": 14, "H": 16, "I": 18, "J": 3})

    for c in range(1, 11):
        ws3.cell(row=1, column=c).fill = orange_fill
    ws3.row_dimensions[1].height = 6

    ws3.merge_cells("B2:I2")
    ws3.cell(row=2, column=2, value="MATERIALS TRACKING").font = navy_bold_16
    ws3.row_dimensions[2].height = 30

    mat_headers = ["Date", "Supplier", "Description", "Qty", "Unit Cost", "Total", "Receipt #", "Category"]
    mat_cols = list(range(2, 10))
    hr = 3
    for c, h in zip(mat_cols, mat_headers):
        cell = ws3.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws3.row_dimensions[hr].height = 28

    for r in range(hr + 1, hr + 51):
        fill = light_gray_fill if (r - hr - 1) % 2 == 0 else white_fill
        for c in mat_cols:
            cell = ws3.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws3.cell(row=r, column=2).number_format = DATE_FMT
        ws3.cell(row=r, column=2).alignment = center
        ws3.cell(row=r, column=5).alignment = center  # Qty
        ws3.cell(row=r, column=6).number_format = USD  # Unit Cost
        ws3.cell(row=r, column=6).alignment = right_align
        ws3.cell(row=r, column=7).value = f'=IF(E{r}="","",E{r}*F{r})'  # Total
        ws3.cell(row=r, column=7).number_format = USD
        ws3.cell(row=r, column=7).alignment = right_align

    # Category dropdown for materials
    dv_mat = DataValidation(type="list", formula1='"Lumber,Electrical,Plumbing,Drywall,Paint,Hardware,Fixtures,Concrete,Roofing,Other"', allow_blank=True)
    ws3.add_data_validation(dv_mat)
    dv_mat.add(f"I{hr+1}:I{hr+50}")

    # Running total
    rt_row = hr + 51
    ws3.merge_cells(f"B{rt_row}:F{rt_row}")
    ws3.cell(row=rt_row, column=2, value="TOTAL MATERIALS COST").font = white_bold_12
    ws3.cell(row=rt_row, column=2).fill = navy_fill
    ws3.cell(row=rt_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(3, 8):
        ws3.cell(row=rt_row, column=c).fill = navy_fill
    ws3.cell(row=rt_row, column=7).value = f"=SUM(G{hr+1}:G{hr+50})"
    ws3.cell(row=rt_row, column=7).number_format = USD
    ws3.cell(row=rt_row, column=7).font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws3.cell(row=rt_row, column=7).fill = navy_fill
    ws3.cell(row=rt_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws3.row_dimensions[rt_row].height = 30

    # ═══ SHEET 4: SUBS & EQUIPMENT ═══
    ws4 = wb.create_sheet("Subs & Equipment")
    set_col_widths(ws4, {"A": 3, "B": 14, "C": 18, "D": 22, "E": 16, "F": 16, "G": 16,
                         "H": 3, "I": 14, "J": 18, "K": 16, "L": 10, "M": 12, "N": 14, "O": 3})

    for c in range(1, 16):
        ws4.cell(row=1, column=c).fill = orange_fill
    ws4.row_dimensions[1].height = 6

    # ── Subcontractors section ──
    ws4.merge_cells("B2:G2")
    ws4.cell(row=2, column=2, value="SUBCONTRACTORS").font = navy_bold_16
    ws4.row_dimensions[2].height = 30

    sub_headers = ["Date", "Name", "Scope of Work", "Quote", "Actual", "Variance"]
    sub_cols = list(range(2, 8))
    hr = 3
    for c, h in zip(sub_cols, sub_headers):
        cell = ws4.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws4.row_dimensions[hr].height = 28

    for r in range(hr + 1, hr + 11):
        fill = light_gray_fill if (r - hr - 1) % 2 == 0 else white_fill
        for c in sub_cols:
            cell = ws4.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws4.cell(row=r, column=2).number_format = DATE_FMT
        ws4.cell(row=r, column=2).alignment = center
        ws4.cell(row=r, column=4).alignment = wrap
        ws4.cell(row=r, column=5).number_format = USD
        ws4.cell(row=r, column=5).alignment = right_align
        ws4.cell(row=r, column=6).number_format = USD
        ws4.cell(row=r, column=6).alignment = right_align
        ws4.cell(row=r, column=7).value = f'=IF(E{r}="","",E{r}-F{r})'
        ws4.cell(row=r, column=7).number_format = USD
        ws4.cell(row=r, column=7).alignment = right_align

    # Sub total
    sub_total_row = hr + 11
    ws4.merge_cells(f"B{sub_total_row}:E{sub_total_row}")
    ws4.cell(row=sub_total_row, column=2, value="SUBCONTRACTOR TOTAL").font = white_bold
    ws4.cell(row=sub_total_row, column=2).fill = navy_fill
    ws4.cell(row=sub_total_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(3, 8):
        ws4.cell(row=sub_total_row, column=c).fill = navy_fill
    ws4.cell(row=sub_total_row, column=6).value = f"=SUM(F{hr+1}:F{hr+10})"
    ws4.cell(row=sub_total_row, column=6).number_format = USD
    ws4.cell(row=sub_total_row, column=6).font = Font(name="Calibri", bold=True, color=WHITE)
    ws4.cell(row=sub_total_row, column=6).fill = navy_fill
    ws4.cell(row=sub_total_row, column=6).alignment = right_align
    ws4.row_dimensions[sub_total_row].height = 28

    # Conditional formatting on variance
    ws4.conditional_formatting.add(f"G{hr+1}:G{hr+10}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=Font(color="22C55E", bold=True), fill=light_green_fill))
    ws4.conditional_formatting.add(f"G{hr+1}:G{hr+10}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="EF4444", bold=True), fill=light_red_fill))

    # ── Equipment section ──
    ws4.merge_cells("I2:N2")
    ws4.cell(row=2, column=9, value="EQUIPMENT RENTALS").font = navy_bold_16

    equip_headers = ["Date", "Item", "Vendor", "Days", "Rate", "Total"]
    equip_cols = list(range(9, 15))
    for c, h in zip(equip_cols, equip_headers):
        cell = ws4.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border

    for r in range(hr + 1, hr + 11):
        fill = light_gray_fill if (r - hr - 1) % 2 == 0 else white_fill
        for c in equip_cols:
            cell = ws4.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws4.cell(row=r, column=9).number_format = DATE_FMT
        ws4.cell(row=r, column=9).alignment = center
        ws4.cell(row=r, column=12).alignment = center  # Days
        ws4.cell(row=r, column=13).number_format = USD  # Rate
        ws4.cell(row=r, column=13).alignment = right_align
        ws4.cell(row=r, column=14).value = f'=IF(L{r}="","",L{r}*M{r})'  # Total
        ws4.cell(row=r, column=14).number_format = USD
        ws4.cell(row=r, column=14).alignment = right_align

    # Equipment total
    eq_total_row = hr + 11
    ws4.merge_cells(f"I{eq_total_row}:K{eq_total_row}")
    ws4.cell(row=eq_total_row, column=9, value="EQUIPMENT TOTAL").font = white_bold
    ws4.cell(row=eq_total_row, column=9).fill = navy_fill
    ws4.cell(row=eq_total_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(10, 15):
        ws4.cell(row=eq_total_row, column=c).fill = navy_fill
    # Fix: Equipment total should reference column N (14), not L (12)
    ws4.cell(row=eq_total_row, column=14).value = f"=SUM(N{hr+1}:N{hr+10})"
    ws4.cell(row=eq_total_row, column=14).number_format = USD
    ws4.cell(row=eq_total_row, column=14).font = Font(name="Calibri", bold=True, color=WHITE)
    ws4.cell(row=eq_total_row, column=14).fill = navy_fill
    ws4.cell(row=eq_total_row, column=14).alignment = right_align
    ws4.row_dimensions[eq_total_row].height = 28

    # ═══ SHEET 5: CHANGE ORDERS ═══
    ws5 = wb.create_sheet("Change Orders")
    set_col_widths(ws5, {"A": 3, "B": 10, "C": 14, "D": 36, "E": 16, "F": 18, "G": 16, "H": 3})

    for c in range(1, 9):
        ws5.cell(row=1, column=c).fill = orange_fill
    ws5.row_dimensions[1].height = 6

    ws5.merge_cells("B2:G2")
    ws5.cell(row=2, column=2, value="CHANGE ORDERS").font = navy_bold_16
    ws5.row_dimensions[2].height = 30

    co_headers = ["CO #", "Date", "Description", "Amount", "Approved By", "Status"]
    co_cols = list(range(2, 8))
    hr = 3
    for c, h in zip(co_cols, co_headers):
        cell = ws5.cell(row=hr, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
    ws5.row_dimensions[hr].height = 28

    for r in range(hr + 1, hr + 21):
        fill = light_gray_fill if (r - hr - 1) % 2 == 0 else white_fill
        for c in co_cols:
            cell = ws5.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = dark_font
        ws5.cell(row=r, column=2).alignment = center  # CO#
        ws5.cell(row=r, column=3).number_format = DATE_FMT
        ws5.cell(row=r, column=3).alignment = center
        ws5.cell(row=r, column=4).alignment = wrap  # Description
        ws5.cell(row=r, column=5).number_format = USD
        ws5.cell(row=r, column=5).alignment = right_align

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Pending,Approved,Rejected"', allow_blank=True)
    dv_status.prompt = "Select status"
    ws5.add_data_validation(dv_status)
    dv_status.add(f"G{hr+1}:G{hr+20}")

    # Total approved changes
    co_total_row = hr + 21
    ws5.merge_cells(f"B{co_total_row}:D{co_total_row}")
    ws5.cell(row=co_total_row, column=2, value='TOTAL APPROVED CHANGE ORDERS').font = white_bold
    ws5.cell(row=co_total_row, column=2).fill = navy_fill
    ws5.cell(row=co_total_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(3, 8):
        ws5.cell(row=co_total_row, column=c).fill = navy_fill
    # SUMIF for approved only
    ws5.cell(row=co_total_row, column=5).value = f'=SUMIF(G{hr+1}:G{hr+20},"Approved",E{hr+1}:E{hr+20})'
    ws5.cell(row=co_total_row, column=5).number_format = USD
    ws5.cell(row=co_total_row, column=5).font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    ws5.cell(row=co_total_row, column=5).fill = navy_fill
    ws5.cell(row=co_total_row, column=5).alignment = right_align
    ws5.row_dimensions[co_total_row].height = 30

    # Now fix Dashboard references to match actual total rows
    # Labor total is at row 53 (hr=3, 50 rows, total at 54) -> G54
    # Materials total is at row 54 -> G54
    # Subs actual total is at F13 (hr=3, 10 rows, total at 14) -> F14
    # Equipment total at L14 -> but we use N14 for total
    # Change Orders total at E24

    # Fix dashboard formulas
    dash = wb["Dashboard"]
    dash.cell(row=10, column=4).value = "=Labor!G54"      # Labor actual
    dash.cell(row=11, column=4).value = "=Materials!G54"   # Materials actual
    dash.cell(row=12, column=4).value = "='Subs & Equipment'!F14"  # Subs actual
    dash.cell(row=13, column=4).value = "='Subs & Equipment'!N14"  # Equipment actual
    dash.cell(row=14, column=4).value = f"='Change Orders'!E{co_total_row}"  # Change orders

    wb.save("/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/job-costing-tracker.xlsx")
    print("✅ Job costing tracker built")


if __name__ == "__main__":
    build_invoice()
    build_estimate()
    build_job_costing()
    print("\n🎉 All 3 templates built successfully!")
