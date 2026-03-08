#!/usr/bin/env python3
"""Template 2: Job Estimate Template"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.page import PageMargins

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estimate"

NAVY = "1A1A2E"
ORANGE = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
DARK_GRAY = "6B7280"
MEDIUM_GRAY = "D1D5DB"

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

white_font = Font(color=WHITE, bold=True, size=11)
navy_font = Font(color=NAVY, bold=True, size=11)
orange_font = Font(color=ORANGE, bold=True, size=11)
normal_font = Font(size=10, color="333333")
small_font = Font(size=9, color=DARK_GRAY)

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)
bottom_border = Border(bottom=Side(style='thin', color=MEDIUM_GRAY))

widths = {'A': 3, 'B': 35, 'C': 10, 'D': 10, 'E': 14, 'F': 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.4, bottom=0.6)
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# Header
for col in range(1, 7):
    for row in range(1, 4):
        ws.cell(row=row, column=col).fill = navy_fill

ws.merge_cells('B1:C3')
ws.cell(row=1, column=2, value="JOB ESTIMATE").font = Font(color=WHITE, bold=True, size=22)
ws.cell(row=1, column=2).alignment = Alignment(vertical='center')

ws.merge_cells('E1:F1')
ws.cell(row=1, column=5, value="[YOUR LOGO]").font = Font(color="888888", italic=True, size=9)
ws.cell(row=1, column=5).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=1, column=5).fill = navy_fill

# Company info
ws.merge_cells('B5:D5')
ws.cell(row=5, column=2, value="Your Company Name").font = Font(color=NAVY, bold=True, size=14)
ws.merge_cells('B6:D6')
ws.cell(row=6, column=2, value="123 Business St, City, State ZIP | (555) 123-4567").font = small_font
ws.merge_cells('B7:D7')
ws.cell(row=7, column=2, value="License #: _____________ | info@company.com").font = small_font

details = [("Estimate #:", "EST-0001"), ("Date:", "MM/DD/YYYY"), ("Valid Until:", "MM/DD/YYYY")]
for i, (label, val) in enumerate(details):
    r = 5 + i
    ws.cell(row=r, column=5, value=label).font = Font(color=NAVY, bold=True, size=9)
    ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6, value=val).font = normal_font
    ws.cell(row=r, column=6).border = bottom_border

for col in range(2, 7):
    ws.cell(row=9, column=col).fill = orange_fill

ws.cell(row=11, column=2, value="CLIENT INFORMATION").font = orange_font
ws.cell(row=11, column=5, value="PROJECT DETAILS").font = orange_font

for i, field in enumerate(["Client Name", "Company", "Address", "Phone / Email"]):
    r = 12 + i
    ws.cell(row=r, column=2, value=field).font = Font(color=DARK_GRAY, italic=True, size=9)
    ws.cell(row=r, column=2).border = bottom_border
    ws.cell(row=r, column=3).border = bottom_border

for i, field in enumerate(["Project Name", "Location"]):
    r = 12 + i
    ws.cell(row=r, column=5, value=field).font = Font(color=DARK_GRAY, italic=True, size=9)
    ws.cell(row=r, column=5).border = bottom_border
    ws.cell(row=r, column=6).border = bottom_border

ws.merge_cells('B17:F17')
ws.cell(row=17, column=2, value="SCOPE OF WORK").font = orange_font
ws.merge_cells('B18:F19')
ws.cell(row=18, column=2, value="[Describe the scope of work here...]").font = Font(color=DARK_GRAY, italic=True, size=9)
ws.cell(row=18, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws.cell(row=18, column=2).border = thin_border

dv = DataValidation(type="list", formula1='"Hours,Days,Sq Ft,Linear Ft,Each,Lot,Job,Set,Roll,Bundle"', allow_blank=True)
ws.add_data_validation(dv)

def add_category(ws, start_row, name, num_rows=10):
    for col in range(2, 7):
        ws.cell(row=start_row, column=col).fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    ws.cell(row=start_row, column=2, value=name).font = Font(color=NAVY, bold=True, size=11)
    
    for col_idx, header in enumerate(["Description", "Qty", "Unit", "Unit Cost", "Total"]):
        cell = ws.cell(row=start_row + 1, column=col_idx + 2, value=header)
        cell.font = Font(color=WHITE, bold=True, size=9)
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for i in range(num_rows):
        r = start_row + 2 + i
        fill = light_gray_fill if i % 2 == 0 else white_fill
        for col in range(2, 7):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(size=9, color="333333")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        dv.add(ws.cell(row=r, column=4))
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=6).value = f'=IF(AND(C{r}<>"",E{r}<>""),C{r}*E{r},"")'
        ws.cell(row=r, column=6).number_format = '$#,##0.00'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=6).font = Font(size=9, color=NAVY, bold=True)
    
    sub_row = start_row + 2 + num_rows
    ws.cell(row=sub_row, column=5, value=f"{name} Subtotal:").font = Font(color=NAVY, bold=True, size=9)
    ws.cell(row=sub_row, column=5).alignment = Alignment(horizontal='right')
    first_r = start_row + 2
    last_r = start_row + 1 + num_rows
    ws.cell(row=sub_row, column=6).value = f'=SUM(F{first_r}:F{last_r})'
    ws.cell(row=sub_row, column=6).font = Font(color=NAVY, bold=True, size=10)
    ws.cell(row=sub_row, column=6).number_format = '$#,##0.00'
    ws.cell(row=sub_row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=sub_row, column=6).border = Border(top=Side(style='thin', color=NAVY), bottom=Side(style='double', color=NAVY))
    return sub_row

row = 21
cat_rows = {}
for cat in ["LABOR", "MATERIALS", "EQUIPMENT", "OTHER COSTS"]:
    sub_row = add_category(ws, row, cat, num_rows=10)
    cat_rows[cat] = sub_row
    row = sub_row + 2

# Summary
summary_start = row + 1
for col in range(2, 7):
    ws.cell(row=summary_start, column=col).fill = orange_fill
ws.merge_cells(f'B{summary_start}:F{summary_start}')
ws.cell(row=summary_start, column=2, value="ESTIMATE SUMMARY").font = Font(color=WHITE, bold=True, size=12)
ws.cell(row=summary_start, column=2).alignment = Alignment(horizontal='center')
ws.cell(row=summary_start, column=2).fill = orange_fill

r = summary_start + 1
for cat_name, cat_sub_row in cat_rows.items():
    ws.cell(row=r, column=4, value=cat_name.title()).font = Font(color="333333", size=10)
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).value = f'=F{cat_sub_row}'
    ws.cell(row=r, column=6).font = normal_font
    ws.cell(row=r, column=6).number_format = '$#,##0.00'
    ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).border = bottom_border
    r += 1

# Direct costs subtotal
ws.cell(row=r, column=4, value="Direct Costs Subtotal:").font = navy_font
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
direct_first = summary_start + 1
direct_last = r - 1
ws.cell(row=r, column=6).value = f'=SUM(F{direct_first}:F{direct_last})'
ws.cell(row=r, column=6).font = Font(color=NAVY, bold=True, size=10)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = Border(top=Side(style='thin', color=NAVY))
direct_subtotal_row = r
r += 1

# Overhead
ws.cell(row=r, column=3).value = 0.20
ws.cell(row=r, column=3).number_format = '0%'
ws.cell(row=r, column=3).font = Font(color=ORANGE, bold=True, size=10)
ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=4, value="Overhead").font = Font(color="333333", size=10)
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=F{direct_subtotal_row}*C{r}'
ws.cell(row=r, column=6).font = normal_font
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = bottom_border
overhead_row = r
r += 1

# Contingency
ws.cell(row=r, column=3).value = 0.10
ws.cell(row=r, column=3).number_format = '0%'
ws.cell(row=r, column=3).font = Font(color=ORANGE, bold=True, size=10)
ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=4, value="Contingency").font = Font(color="333333", size=10)
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=F{direct_subtotal_row}*C{r}'
ws.cell(row=r, column=6).font = normal_font
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = bottom_border
contingency_row = r
r += 1

# Subtotal before tax
ws.cell(row=r, column=4, value="Subtotal:").font = navy_font
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=F{direct_subtotal_row}+F{overhead_row}+F{contingency_row}'
ws.cell(row=r, column=6).font = Font(color=NAVY, bold=True, size=10)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
subtotal_row = r
r += 1

# Tax
ws.cell(row=r, column=3).value = 0.0
ws.cell(row=r, column=3).number_format = '0.00%'
ws.cell(row=r, column=3).font = Font(color=ORANGE, bold=True, size=10)
ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
ws.cell(row=r, column=4, value="Tax").font = Font(color="333333", size=10)
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=F{subtotal_row}*C{r}'
ws.cell(row=r, column=6).font = normal_font
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = bottom_border
tax_row = r
r += 1

# Grand Total
for col in range(4, 7):
    ws.cell(row=r, column=col).fill = navy_fill
ws.cell(row=r, column=4, value="GRAND TOTAL").font = Font(color=WHITE, bold=True, size=14)
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=r, column=4).fill = navy_fill
ws.cell(row=r, column=6).value = f'=F{subtotal_row}+F{tax_row}'
ws.cell(row=r, column=6).font = Font(color=WHITE, bold=True, size=14)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=r, column=6).fill = navy_fill
ws.row_dimensions[r].height = 30
r += 2

# Timeline
ws.cell(row=r, column=2, value="PROJECT TIMELINE").font = orange_font
r += 1
for label in ["Estimated Start Date:", "Estimated End Date:", "Estimated Working Days:"]:
    ws.cell(row=r, column=2, value=label).font = Font(color=NAVY, bold=True, size=9)
    ws.cell(row=r, column=3).border = bottom_border
    ws.cell(row=r, column=4).border = bottom_border
    r += 1

r += 1
ws.cell(row=r, column=2, value="TERMS & CONDITIONS").font = orange_font
r += 1
terms = [
    "1. This estimate is valid for 30 days from the date above.",
    "2. Any changes to the scope of work may result in additional charges.",
    "3. Payment schedule: 50% deposit, 40% at rough-in, 10% at completion.",
    "4. All work performed in accordance with local building codes.",
    "5. Warranty: 1 year on workmanship from date of completion.",
    "6. Client responsible for obtaining necessary permits unless otherwise stated.",
]
for term in terms:
    ws.cell(row=r, column=2, value=term).font = Font(size=8, color=DARK_GRAY)
    ws.merge_cells(f'B{r}:F{r}')
    r += 1

r += 1
ws.cell(row=r, column=2, value="ACCEPTANCE").font = orange_font
r += 1
ws.cell(row=r, column=2, value="I accept this estimate and authorize the work to begin.").font = Font(size=9, color="333333")
ws.merge_cells(f'B{r}:F{r}')
r += 2

for label in ["Client Signature:", "Printed Name:", "Date:"]:
    ws.cell(row=r, column=2, value=label).font = Font(color=NAVY, bold=True, size=9)
    ws.merge_cells(f'C{r}:D{r}')
    ws.cell(row=r, column=3).border = Border(bottom=Side(style='thin', color=NAVY))
    ws.cell(row=r, column=4).border = Border(bottom=Side(style='thin', color=NAVY))
    r += 1

ws.print_area = f'A1:F{r}'
ws.oddFooter.center.text = "Page &P of &N"

# === SHEET 2: Estimate vs Actual ===
ws2 = wb.create_sheet("Estimate vs Actual")

for col in range(1, 9):
    ws2.cell(row=1, column=col).fill = navy_fill
ws2.merge_cells('A1:H1')
ws2.cell(row=1, column=1, value="ESTIMATE vs ACTUAL COMPARISON").font = Font(color=WHITE, bold=True, size=16)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

ws2.cell(row=3, column=1, value="Project:").font = navy_font
ws2.merge_cells('B3:D3')
ws2.cell(row=3, column=2).border = bottom_border

ea_headers = ["Category", "Item", "Estimated", "Actual", "Variance ($)", "Variance (%)", "Status", "Notes"]
col_widths = [16, 25, 16, 16, 16, 14, 12, 25]
for col_idx, (header, width) in enumerate(zip(ea_headers, col_widths), 1):
    cell = ws2.cell(row=5, column=col_idx, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

dv_cat = DataValidation(type="list", formula1='"Labor,Materials,Equipment,Other,Overhead,Contingency"', allow_blank=True)
ws2.add_data_validation(dv_cat)
dv_status = DataValidation(type="list", formula1='"Under Budget,On Budget,Over Budget,Pending"', allow_blank=True)
ws2.add_data_validation(dv_status)

for i in range(40):
    r = 6 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    for col in range(1, 9):
        cell = ws2.cell(row=r, column=col)
        cell.fill = fill
        cell.border = thin_border
        cell.font = Font(size=9, color="333333")
    dv_cat.add(ws2.cell(row=r, column=1))
    ws2.cell(row=r, column=3).number_format = '$#,##0.00'
    ws2.cell(row=r, column=4).number_format = '$#,##0.00'
    ws2.cell(row=r, column=5).value = f'=IF(AND(C{r}<>"",D{r}<>""),C{r}-D{r},"")'
    ws2.cell(row=r, column=5).number_format = '$#,##0.00'
    ws2.cell(row=r, column=6).value = f'=IF(AND(C{r}<>"",D{r}<>"",C{r}<>0),(C{r}-D{r})/C{r},"")'
    ws2.cell(row=r, column=6).number_format = '0.0%'
    dv_status.add(ws2.cell(row=r, column=7))

ws2.conditional_formatting.add('E6:E45',
    CellIsRule(operator='lessThan', formula=['0'],
               fill=PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
               font=Font(color="CC0000", bold=True)))
ws2.conditional_formatting.add('E6:E45',
    CellIsRule(operator='greaterThan', formula=['0'],
               fill=PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
               font=Font(color="008800", bold=True)))

r = 46
ws2.cell(row=r, column=2, value="TOTALS").font = navy_font
for col in [3, 4, 5]:
    ws2.cell(row=r, column=col).value = f'=SUM({get_column_letter(col)}6:{get_column_letter(col)}45)'
    ws2.cell(row=r, column=col).font = Font(color=NAVY, bold=True)
    ws2.cell(row=r, column=col).number_format = '$#,##0.00'

ws2.print_area = 'A1:H46'
ws2.page_setup.orientation = 'landscape'

output_path = "/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/job-estimate-template.xlsx"
wb.save(output_path)
print(f"DONE: job-estimate-template.xlsx")
