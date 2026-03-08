#!/usr/bin/env python3
"""Template 1: Professional Contractor Invoice"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.page import PageMargins

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Invoice"

# Colors
NAVY = "1A1A2E"
ORANGE = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
DARK_GRAY = "6B7280"
MEDIUM_GRAY = "D1D5DB"

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

white_font = Font(color=WHITE, bold=True, size=11)
white_font_lg = Font(color=WHITE, bold=True, size=24)
navy_font = Font(color=NAVY, bold=True, size=11)
orange_font = Font(color=ORANGE, bold=True, size=11)
normal_font = Font(size=11, color="333333")
small_font = Font(size=9, color=DARK_GRAY)

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)
bottom_border = Border(bottom=Side(style='thin', color=MEDIUM_GRAY))

# Column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 3
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 20

# Page setup
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.75)
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

# === HEADER SECTION ===
for col in range(1, 10):
    for row in range(1, 4):
        ws.cell(row=row, column=col).fill = navy_fill

ws.merge_cells('B1:D3')
cell = ws.cell(row=1, column=2, value="INVOICE")
cell.font = white_font_lg
cell.alignment = Alignment(vertical='center', horizontal='left')

ws.merge_cells('E1:F1')
ws.cell(row=1, column=5, value="[YOUR LOGO HERE]").font = Font(color="888888", italic=True, size=10)
ws.cell(row=1, column=5).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=1, column=5).fill = navy_fill

# Company info area
ws.merge_cells('B5:D5')
ws.cell(row=5, column=2, value="Your Company Name").font = Font(color=NAVY, bold=True, size=16)
ws.merge_cells('B6:D6')
ws.cell(row=6, column=2, value="123 Business Street, City, State ZIP").font = small_font
ws.merge_cells('B7:D7')
ws.cell(row=7, column=2, value="Phone: (555) 123-4567 | Email: info@company.com").font = small_font

# Invoice details (right side)
labels_r = [("Invoice #:", "INV-0001"), ("Date:", "MM/DD/YYYY"), ("Due Date:", "MM/DD/YYYY"), ("PO #:", "")]
for i, (label, val) in enumerate(labels_r):
    r = 5 + i
    ws.cell(row=r, column=5, value=label).font = Font(color=NAVY, bold=True, size=10)
    ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6, value=val).font = normal_font
    ws.cell(row=r, column=6).alignment = Alignment(horizontal='left')
    ws.cell(row=r, column=6).border = bottom_border

# Orange accent line
for col in range(2, 7):
    ws.cell(row=9, column=col).fill = orange_fill

# Bill To Section
ws.cell(row=11, column=2, value="BILL TO").font = Font(color=ORANGE, bold=True, size=11)
ws.cell(row=11, column=5, value="PROJECT / JOB SITE").font = Font(color=ORANGE, bold=True, size=11)

bill_to_fields = ["Client Name", "Company Name", "Address", "City, State ZIP", "Phone / Email"]
for i, field in enumerate(bill_to_fields):
    r = 12 + i
    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(row=r, column=2, value=field).font = Font(color=DARK_GRAY, italic=True, size=10)
    ws.cell(row=r, column=2).border = bottom_border
    ws.cell(row=r, column=3).border = bottom_border

project_fields = ["Project Name", "Job Site Address", "City, State ZIP"]
for i, field in enumerate(project_fields):
    r = 12 + i
    ws.merge_cells(f'E{r}:F{r}')
    ws.cell(row=r, column=5, value=field).font = Font(color=DARK_GRAY, italic=True, size=10)
    ws.cell(row=r, column=5).border = bottom_border
    ws.cell(row=r, column=6).border = bottom_border

# === LINE ITEMS TABLE ===
header_row = 18
col_headers = ["Description", "Quantity", "Unit", "Unit Price", "Amount"]
for col_idx, header in enumerate(col_headers, 2):
    cell = ws.cell(row=header_row, column=col_idx, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data validation for Unit column
dv = DataValidation(type="list", formula1='"Hours,Days,Sq Ft,Linear Ft,Each,Lot,Job"', allow_blank=True)
dv.error = "Please select a valid unit type"
dv.errorTitle = "Invalid Unit"
ws.add_data_validation(dv)

# Line item rows (20 rows)
first_data_row = 19
for i in range(20):
    r = first_data_row + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    
    ws.cell(row=r, column=1).value = i + 1
    ws.cell(row=r, column=1).font = Font(size=8, color=DARK_GRAY)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    
    for col in range(2, 7):
        cell = ws.cell(row=r, column=col)
        cell.fill = fill
        cell.border = thin_border
        cell.font = normal_font
    
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3).number_format = '#,##0.00'
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    dv.add(ws.cell(row=r, column=4))
    ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=5).number_format = '$#,##0.00'
    
    ws.cell(row=r, column=6).value = f'=IF(AND(C{r}<>"",E{r}<>""),C{r}*E{r},"")'
    ws.cell(row=r, column=6).font = Font(size=11, color=NAVY, bold=True)
    ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=6).number_format = '$#,##0.00'

last_data_row = first_data_row + 19

# Conditional formatting: highlight amount > $1000
ws.conditional_formatting.add(f'F{first_data_row}:F{last_data_row}',
    CellIsRule(operator='greaterThan', formula=['1000'],
               fill=PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
               font=Font(color=ORANGE, bold=True)))

# === TOTALS SECTION ===
totals_start = last_data_row + 1
for col in range(2, 7):
    ws.cell(row=totals_start, column=col).fill = orange_fill

# Subtotal
r = totals_start + 1
ws.cell(row=r, column=5, value="Subtotal").font = navy_font
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=SUM(F{first_data_row}:F{last_data_row})'
ws.cell(row=r, column=6).font = Font(size=11, color=NAVY, bold=True)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = bottom_border

# Tax Rate (configurable in H6)
ws.cell(row=5, column=8, value="Tax Rate:").font = Font(size=9, color=DARK_GRAY, bold=True)
ws.cell(row=6, column=8, value=0.0).font = Font(size=12, color=ORANGE, bold=True)
ws.cell(row=6, column=8).number_format = '0.00%'
ws.cell(row=6, column=8).border = Border(bottom=Side(style='double', color=ORANGE))
ws.cell(row=6, column=8).alignment = Alignment(horizontal='center')

r = totals_start + 2
ws.cell(row=r, column=5, value="Tax").font = navy_font
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=F{totals_start+1}*H6'
ws.cell(row=r, column=6).font = Font(size=11, color=NAVY)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = bottom_border

# Total
r = totals_start + 3
for col in range(5, 7):
    ws.cell(row=r, column=col).fill = navy_fill
ws.cell(row=r, column=5, value="TOTAL").font = Font(color=WHITE, bold=True, size=14)
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=r, column=5).fill = navy_fill
ws.cell(row=r, column=6).value = f'=F{totals_start+1}+F{totals_start+2}'
ws.cell(row=r, column=6).font = Font(color=WHITE, bold=True, size=14)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=r, column=6).fill = navy_fill

# Amount Paid
r = totals_start + 4
ws.cell(row=r, column=5, value="Amount Paid").font = Font(color=DARK_GRAY, size=10)
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = 0
ws.cell(row=r, column=6).font = normal_font
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).border = bottom_border

# Balance Due
r = totals_start + 5
ws.cell(row=r, column=5, value="BALANCE DUE").font = Font(color=ORANGE, bold=True, size=12)
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6).value = f'=F{totals_start+3}-F{totals_start+4}'
ws.cell(row=r, column=6).font = Font(color=ORANGE, bold=True, size=12)
ws.cell(row=r, column=6).number_format = '$#,##0.00'
ws.cell(row=r, column=6).alignment = Alignment(horizontal='right')

# === PAYMENT TERMS ===
terms_row = totals_start + 7
ws.merge_cells(f'B{terms_row}:C{terms_row}')
ws.cell(row=terms_row, column=2, value="PAYMENT TERMS").font = Font(color=ORANGE, bold=True, size=11)

terms = [
    "Payment is due within 30 days of invoice date.",
    "Late payments subject to 1.5% monthly interest.",
    "Make checks payable to: [Your Company Name]",
    "Bank details for wire transfer: [Add bank details]"
]
for i, term in enumerate(terms):
    r = terms_row + 1 + i
    ws.merge_cells(f'B{r}:F{r}')
    ws.cell(row=r, column=2, value=term).font = Font(size=9, color=DARK_GRAY)

# Notes
notes_row = terms_row + 6
ws.merge_cells(f'B{notes_row}:C{notes_row}')
ws.cell(row=notes_row, column=2, value="NOTES").font = Font(color=ORANGE, bold=True, size=11)
ws.merge_cells(f'B{notes_row+1}:F{notes_row+2}')
ws.cell(row=notes_row+1, column=2).border = thin_border
ws.cell(row=notes_row+1, column=2).alignment = Alignment(wrap_text=True, vertical='top')

# Thank you footer
footer_row = notes_row + 4
for col in range(2, 7):
    ws.cell(row=footer_row, column=col).fill = orange_fill
ws.merge_cells(f'B{footer_row}:F{footer_row}')
ws.cell(row=footer_row, column=2, value="Thank you for your business!").font = Font(color=WHITE, bold=True, size=12, italic=True)
ws.cell(row=footer_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=footer_row, column=2).fill = orange_fill
ws.row_dimensions[footer_row].height = 30

ws.print_area = f'A1:G{footer_row}'
ws.oddFooter.center.text = "Page &P of &N"

# === SHEET 2: PAYMENT LOG ===
ws2 = wb.create_sheet("Payment Log")

for col in range(1, 9):
    ws2.cell(row=1, column=col).fill = navy_fill
ws2.merge_cells('A1:H1')
ws2.cell(row=1, column=1, value="PAYMENT LOG").font = Font(color=WHITE, bold=True, size=16)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

pay_headers = ["Date", "Invoice #", "Payment Method", "Check/Ref #", "Amount Received", "Applied To", "Balance Remaining", "Notes"]
col_widths = [14, 14, 18, 16, 18, 16, 18, 30]
for col_idx, (header, width) in enumerate(zip(pay_headers, col_widths), 1):
    cell = ws2.cell(row=3, column=col_idx, value=header)
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

dv_pay = DataValidation(type="list", formula1='"Check,Cash,Credit Card,Bank Transfer,ACH,Zelle,Venmo,Other"', allow_blank=True)
ws2.add_data_validation(dv_pay)

for i in range(30):
    r = 4 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    for col in range(1, 9):
        cell = ws2.cell(row=r, column=col)
        cell.fill = fill
        cell.border = thin_border
        cell.font = normal_font
    ws2.cell(row=r, column=1).number_format = 'MM/DD/YYYY'
    ws2.cell(row=r, column=5).number_format = '$#,##0.00'
    ws2.cell(row=r, column=7).number_format = '$#,##0.00'
    dv_pay.add(ws2.cell(row=r, column=3))

r = 34
ws2.cell(row=r, column=4, value="Total Received:").font = navy_font
ws2.cell(row=r, column=4).alignment = Alignment(horizontal='right')
ws2.cell(row=r, column=5).value = '=SUM(E4:E33)'
ws2.cell(row=r, column=5).font = Font(color=NAVY, bold=True, size=12)
ws2.cell(row=r, column=5).number_format = '$#,##0.00'

ws2.print_area = 'A1:H34'
ws2.page_setup.orientation = 'landscape'

output_path = "/home/openclaw/.openclaw/workspaces/agent4/builtright-academy/assets/templates/contractor-invoice-template.xlsx"
wb.save(output_path)
print(f"DONE: contractor-invoice-template.xlsx")
