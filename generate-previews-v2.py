#!/usr/bin/env python3
"""Generate better template preview images - focus on the data grid that sells."""

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os

DOWNLOADS = 'downloads'
ASSETS = 'assets/previews'
os.makedirs(ASSETS, exist_ok=True)

# Colors
BG = (15, 23, 42)
HEADER_BG = (30, 41, 59)
ROW1 = (22, 33, 54)
ROW2 = (17, 27, 46)
BORDER = (51, 65, 85)
TEXT = (248, 250, 252)
TEXT_DIM = (148, 163, 184)
ORANGE = (249, 115, 22)
GREEN = (34, 197, 94)

def load_font(name, size):
    try:
        return ImageFont.truetype(f"/usr/share/fonts/truetype/dejavu/{name}.ttf", size)
    except:
        return ImageFont.load_default()

font_title = load_font("DejaVuSans-Bold", 18)
font_bold = load_font("DejaVuSans-Bold", 13)
font = load_font("DejaVuSans", 12)
font_sm = load_font("DejaVuSans", 10)
font_xs = load_font("DejaVuSans", 9)

def format_val(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value) < 1 and value != 0:
            return f"{value:.1%}"
        return f"${value:,.2f}" if abs(value) >= 1 else f"{value:.2f}"
    if isinstance(value, int):
        if value > 100:
            return f"${value:,}"
        return str(value)
    return str(value)[:25]

def generate_preview(xlsx_path, output_name, title, subtitle, highlights):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    W, H = 800, 520
    img = Image.new('RGB', (W, H), BG)
    draw = ImageDraw.Draw(img)
    
    # Top bar with title
    draw.rectangle([(0, 0), (W, 55)], fill=HEADER_BG)
    draw.rectangle([(0, 52), (W, 55)], fill=ORANGE)  # Orange accent line
    draw.text((24, 10), title, fill=TEXT, font=font_title)
    draw.text((24, 33), subtitle, fill=TEXT_DIM, font=font_sm)
    draw.text((W - 160, 18), "BuiltRight Academy", fill=ORANGE, font=font_bold)
    
    # Spreadsheet area
    y_start = 65
    row_h = 26
    x_start = 20
    max_rows = min(14, ws.max_row or 10)
    max_cols = min(7, ws.max_column or 4)
    
    # Determine column widths based on content
    col_widths = []
    for c in range(1, max_cols + 1):
        max_len = 0
        for r in range(1, max_rows + 1):
            val = format_val(ws.cell(row=r, column=c).value)
            max_len = max(max_len, len(val))
        col_widths.append(max(70, min(160, max_len * 9 + 20)))
    
    # Scale to fit
    total_w = sum(col_widths)
    available = W - 40
    if total_w > available:
        ratio = available / total_w
        col_widths = [int(w * ratio) for w in col_widths]
    
    # Draw rows
    for row_idx in range(max_rows):
        r = row_idx + 1
        y = y_start + row_idx * row_h
        
        is_header = row_idx == 0
        row_bg = HEADER_BG if is_header else (ROW1 if row_idx % 2 == 0 else ROW2)
        
        draw.rectangle([(x_start, y), (x_start + sum(col_widths), y + row_h)], fill=row_bg)
        
        x = x_start
        for col_idx in range(max_cols):
            c = col_idx + 1
            value = format_val(ws.cell(row=r, column=c).value)
            
            text_color = ORANGE if is_header else TEXT
            f = font_bold if is_header else font
            
            # Highlight dollar values in green
            if '$' in value and not is_header:
                text_color = GREEN
            
            # Truncate if too wide
            max_chars = col_widths[col_idx] // 8
            if len(value) > max_chars:
                value = value[:max_chars-1] + "…"
            
            draw.text((x + 8, y + 5), value, fill=text_color, font=f)
            x += col_widths[col_idx]
            
            # Vertical grid line
            if col_idx < max_cols - 1:
                draw.line([(x, y), (x, y + row_h)], fill=BORDER, width=1)
        
        # Horizontal grid line
        draw.line([(x_start, y + row_h), (x_start + sum(col_widths), y + row_h)], fill=BORDER, width=1)
    
    # Bottom section - highlights
    bottom_y = y_start + max_rows * row_h + 15
    
    # Feature highlights
    for i, highlight in enumerate(highlights):
        hx = 30 + (i * (W - 60) // len(highlights))
        draw.text((hx, bottom_y), f"✓ {highlight}", fill=GREEN, font=font_sm)
    
    # CTA bar
    cta_y = H - 50
    draw.rectangle([(0, cta_y), (W, H)], fill=HEADER_BG)
    draw.rectangle([(0, cta_y), (W, cta_y + 2)], fill=ORANGE)
    draw.text((24, cta_y + 15), "📥 Instant Download  •  Works with Excel & Google Sheets  •  30-Day Money-Back Guarantee", fill=TEXT_DIM, font=font_sm)
    draw.text((W - 150, cta_y + 12), "builtrighthq.com", fill=ORANGE, font=font_bold)
    
    output_path = os.path.join(ASSETS, f"{output_name}.png")
    img.save(output_path, quality=95)
    print(f"  ✅ {output_path}")

# Generate each preview
templates = [
    ("contractor-invoice.xlsx", "invoice-preview", "Contractor Invoice Template",
     "Professional invoicing with auto-calculations",
     ["Auto-calc totals", "Tax handling", "Print-ready"]),
    
    ("contractor-estimate.xlsx", "estimate-preview", "Estimate & Bid Template",
     "Win more jobs with professional estimates",
     ["Materials + labor", "Markup calculator", "Terms included"]),
    
    ("job-costing-tracker.xlsx", "job-costing-preview", "Job Costing Tracker",
     "Know exactly which jobs make money",
     ["Estimated vs actual", "Per-job profit/loss", "6 cost categories"]),
    
    ("profit-loss-tracker.xlsx", "pnl-preview", "Profit & Loss Statement",
     "Track your business health month by month",
     ["12-month view", "Income & expenses", "Net profit formula"]),
    
    ("client-tracker.xlsx", "crm-preview", "Client Tracker / CRM",
     "Never lose a lead or forget a follow-up",
     ["Lead tracking", "Job status", "Revenue per client"]),
    
    ("contractor-proposal.xlsx", "proposal-preview", "Professional Proposal",
     "Close bigger jobs with polished proposals",
     ["Scope of work", "Payment schedule", "Signature ready"]),
    
    ("cash-flow-forecast.xlsx", "cashflow-preview", "12-Month Cash Flow Forecast",
     "See your financial future before it happens",
     ["Monthly projections", "Running balance", "Plan ahead"]),
]

for xlsx_name, output_name, title, subtitle, highlights in templates:
    xlsx_path = os.path.join(DOWNLOADS, xlsx_name)
    if os.path.exists(xlsx_path):
        generate_preview(xlsx_path, output_name, title, subtitle, highlights)

print(f"\n🔥 Generated {len(templates)} preview images")
