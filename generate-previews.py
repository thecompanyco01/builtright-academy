#!/usr/bin/env python3
"""Generate template preview images from Excel files using openpyxl + Pillow."""

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os

DOWNLOADS = 'downloads'
ASSETS = 'assets/previews'
os.makedirs(ASSETS, exist_ok=True)

# Colors
BG = (15, 23, 42)  # #0F172A
HEADER_BG = (30, 41, 59)  # #1E293B
CELL_BG = (30, 41, 59)
CELL_ALT = (22, 33, 50)
BORDER = (51, 65, 85)  # #334155
TEXT = (248, 250, 252)  # #f8fafc
TEXT_DIM = (148, 163, 184)  # #94a3b8
ORANGE = (249, 115, 22)  # #F97316
GREEN = (34, 197, 94)
RED = (239, 68, 68)

def generate_preview(xlsx_path, output_name, title):
    """Generate a preview image from an Excel file."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    # Image dimensions
    W, H = 800, 500
    img = Image.new('RGB', (W, H), BG)
    draw = ImageDraw.Draw(img)
    
    # Try to load a font
    try:
        font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
        font_sm = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
        font_title = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
    except:
        font_bold = ImageFont.load_default()
        font = ImageFont.load_default()
        font_sm = ImageFont.load_default()
        font_title = ImageFont.load_default()
    
    # Title bar
    draw.rectangle([(0, 0), (W, 50)], fill=HEADER_BG)
    draw.text((20, 14), f"📊 {title}", fill=ORANGE, font=font_title)
    draw.text((W-180, 18), "BuiltRight Academy", fill=TEXT_DIM, font=font_sm)
    
    # Draw spreadsheet grid from actual data
    y_start = 60
    row_height = 28
    max_rows = min(15, ws.max_row or 10)
    max_cols = min(6, ws.max_column or 4)
    
    # Calculate column widths
    col_widths = []
    available_width = W - 40
    for c in range(1, max_cols + 1):
        col_width = max(80, available_width // max_cols)
        col_widths.append(col_width)
    
    # Adjust to fit
    total = sum(col_widths)
    if total > available_width:
        ratio = available_width / total
        col_widths = [int(w * ratio) for w in col_widths]
    
    for row_idx in range(max_rows):
        r = row_idx + 1
        y = y_start + row_idx * row_height
        
        # Alternating row colors
        row_bg = CELL_ALT if row_idx % 2 == 0 else CELL_BG
        if row_idx == 0:
            row_bg = (39, 55, 77)  # Header row
        
        draw.rectangle([(20, y), (W - 20, y + row_height)], fill=row_bg)
        
        x = 20
        for col_idx in range(max_cols):
            c = col_idx + 1
            cell = ws.cell(row=r, column=c)
            value = cell.value
            
            if value is None:
                value = ""
            elif isinstance(value, (int, float)):
                if abs(value) > 100:
                    value = f"${value:,.0f}"
                elif abs(value) < 1 and value != 0:
                    value = f"{value:.1%}"
                else:
                    value = f"{value:,.2f}" if isinstance(value, float) else str(value)
            else:
                value = str(value)[:20]
            
            # Color coding
            text_color = TEXT
            if row_idx == 0:
                text_color = ORANGE
                current_font = font_bold
            elif '$' in str(value):
                if '-' in str(value):
                    text_color = RED
                else:
                    text_color = GREEN
                current_font = font
            else:
                current_font = font
            
            draw.text((x + 6, y + 6), value, fill=text_color, font=current_font)
            
            # Column border
            x += col_widths[col_idx] if col_idx < len(col_widths) else 100
            draw.line([(x, y), (x, y + row_height)], fill=BORDER, width=1)
        
        # Row border
        draw.line([(20, y + row_height), (W - 20, y + row_height)], fill=BORDER, width=1)
    
    # Bottom fade effect
    for i in range(60):
        alpha = int(255 * (i / 60))
        draw.rectangle([(0, H - 60 + i), (W, H - 60 + i + 1)], fill=(BG[0], BG[1], BG[2]))
    
    # Bottom text
    draw.text((W//2 - 100, H - 40), "📥 Instant Download — $29", fill=ORANGE, font=font_bold)
    
    # Save
    output_path = os.path.join(ASSETS, f"{output_name}.png")
    img.save(output_path, quality=90)
    print(f"  ✅ {output_path} ({W}x{H})")

# Generate previews for each template
templates = [
    ("contractor-invoice.xlsx", "invoice-preview", "Contractor Invoice Template"),
    ("contractor-estimate.xlsx", "estimate-preview", "Estimate & Bid Template"),
    ("job-costing-tracker.xlsx", "job-costing-preview", "Job Costing Tracker"),
    ("profit-loss-tracker.xlsx", "pnl-preview", "Profit & Loss Tracker"),
    ("client-tracker.xlsx", "crm-preview", "Client Tracker / CRM"),
    ("contractor-proposal.xlsx", "proposal-preview", "Contractor Proposal"),
    ("cash-flow-forecast.xlsx", "cashflow-preview", "Cash Flow Forecast"),
]

for xlsx_name, output_name, title in templates:
    xlsx_path = os.path.join(DOWNLOADS, xlsx_name)
    if os.path.exists(xlsx_path):
        generate_preview(xlsx_path, output_name, title)
    else:
        print(f"  ❌ Missing: {xlsx_path}")

print(f"\n🔥 Done! Generated {len(templates)} preview images")
