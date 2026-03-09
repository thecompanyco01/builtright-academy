#!/usr/bin/env python3
"""Generate all contractor Excel templates for BuiltRight Academy."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

DOWNLOADS = os.path.join(os.path.dirname(__file__), 'downloads')
os.makedirs(DOWNLOADS, exist_ok=True)

# Common styles
ORANGE = "F97316"
DARK_BG = "0F172A"
HEADER_FONT = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
TITLE_FONT = Font(name='Calibri', bold=True, size=24, color=ORANGE)
LABEL_FONT = Font(name='Calibri', bold=True, size=11, color="333333")
DATA_FONT = Font(name='Calibri', size=11)
CURRENCY_FMT = '"$"#,##0.00'
PERCENT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
WHITE_FONT = Font(name='Calibri', bold=True, size=11, color="FFFFFF")

def style_header_row(ws, row, cols, fill=HEADER_FILL):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL

# ============================================================
# 1. CONTRACTOR INVOICE
# ============================================================
def create_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.sheet_properties.pageSetUpPr = None

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    # Title
    ws['A1'] = "INVOICE"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    # Company info
    ws['A3'] = "YOUR COMPANY NAME"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A4'] = "123 Main Street, Suite 100"
    ws['A5'] = "Your City, ST 12345"
    ws['A6'] = "Phone: (555) 123-4567"
    ws['A7'] = "Email: info@yourcompany.com"

    # Invoice details
    ws['C3'] = "Invoice #:"
    ws['C3'].font = LABEL_FONT
    ws['D3'] = "INV-2026-001"
    ws['C4'] = "Date:"
    ws['C4'].font = LABEL_FONT
    ws['D4'] = "03/09/2026"
    ws['C5'] = "Due Date:"
    ws['C5'].font = LABEL_FONT
    ws['D5'] = "04/08/2026"

    # Bill To
    ws['A9'] = "BILL TO:"
    ws['A9'].font = Font(bold=True, size=12, color=ORANGE)
    ws['A10'] = "John & Sarah Mitchell"
    ws['A11'] = "456 Oak Avenue"
    ws['A12'] = "Springfield, IL 62704"
    ws['A13'] = "Phone: (555) 987-6543"
    ws['A14'] = "Email: mitchell.family@email.com"

    # Table header
    row = 16
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 4, fill=ORANGE_FILL)

    # Example line items (kitchen remodel)
    items = [
        ("Demo - Remove existing cabinets & countertops", 1, 450),
        ("Haul-away & dumpster rental (20 yd)", 1, 375),
        ("Install new cabinets (12 linear ft)", 12, 185),
        ("Countertop - Quartz fabrication & install", 1, 3200),
        ("Plumbing rough-in (sink relocation)", 1, 850),
        ("Electrical - Add 3 under-cabinet outlets", 3, 175),
        ("Tile backsplash - 30 sq ft subway tile", 30, 18),
        ("Backsplash labor - install", 1, 480),
        ("Paint - Kitchen walls & ceiling (250 sq ft)", 1, 650),
        ("New sink + faucet install", 1, 425),
        ("Lighting - 4 recessed LED can lights", 4, 95),
        ("Trim & molding - crown molding install", 1, 380),
        ("Hardware - Cabinet pulls & knobs (24 pc)", 24, 12),
        ("Final clean-up & debris removal", 1, 200),
        ("Permit filing fee", 1, 150),
    ]

    for i, (desc, qty, price) in enumerate(items):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=desc)
        ws.cell(row=r, column=2, value=qty).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=price).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
        ws.cell(row=r, column=4).number_format = CURRENCY_FMT
        style_data_row(ws, r, 4, alt=(i % 2 == 1))

    # Subtotal section
    last_item_row = row + len(items)
    sr = last_item_row + 2
    ws.cell(row=sr, column=3, value="Subtotal:").font = LABEL_FONT
    ws.cell(row=sr, column=4).value = f"=SUM(D{row+1}:D{last_item_row})"
    ws.cell(row=sr, column=4).number_format = CURRENCY_FMT
    ws.cell(row=sr, column=4).font = LABEL_FONT

    ws.cell(row=sr+1, column=3, value="Tax Rate:").font = LABEL_FONT
    ws.cell(row=sr+1, column=4, value=0.0825).number_format = PERCENT_FMT

    ws.cell(row=sr+2, column=3, value="Tax Amount:").font = LABEL_FONT
    ws.cell(row=sr+2, column=4).value = f"=D{sr}*D{sr+1}"
    ws.cell(row=sr+2, column=4).number_format = CURRENCY_FMT

    ws.cell(row=sr+3, column=3, value="GRAND TOTAL:").font = Font(bold=True, size=13, color=ORANGE)
    ws.cell(row=sr+3, column=4).value = f"=D{sr}+D{sr+2}"
    ws.cell(row=sr+3, column=4).number_format = CURRENCY_FMT
    ws.cell(row=sr+3, column=4).font = Font(bold=True, size=13, color=ORANGE)

    # Payment terms
    ptr = sr + 5
    ws.cell(row=ptr, column=1, value="PAYMENT TERMS").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=ptr+1, column=1, value="• Payment due within 30 days of invoice date")
    ws.cell(row=ptr+2, column=1, value="• Late payments subject to 1.5% monthly interest")
    ws.cell(row=ptr+3, column=1, value="• Make checks payable to: YOUR COMPANY NAME")
    ws.cell(row=ptr+4, column=1, value="• Accepted: Check, Zelle, Venmo, Credit Card (+3% processing fee)")

    ws.cell(row=ptr+6, column=1, value="NOTES").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=ptr+7, column=1, value="Thank you for choosing our services. Please contact us with any questions.")

    # Print area
    ws.print_area = f"A1:D{ptr+8}"
    wb.save(os.path.join(DOWNLOADS, 'contractor-invoice.xlsx'))
    print("✅ contractor-invoice.xlsx")


# ============================================================
# 2. CONTRACTOR ESTIMATE
# ============================================================
def create_estimate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    ws['A1'] = "ESTIMATE"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    ws['A3'] = "YOUR COMPANY NAME"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A4'] = "123 Main Street, Suite 100"
    ws['A5'] = "Your City, ST 12345"
    ws['A6'] = "Phone: (555) 123-4567"
    ws['A7'] = "Email: info@yourcompany.com"

    ws['C3'] = "Estimate #:"
    ws['C3'].font = LABEL_FONT
    ws['D3'] = "EST-2026-042"
    ws['C4'] = "Date:"
    ws['C4'].font = LABEL_FONT
    ws['D4'] = "03/09/2026"
    ws['C5'] = "Valid Until:"
    ws['C5'].font = LABEL_FONT
    ws['D5'] = "04/08/2026"

    ws['A9'] = "PREPARED FOR:"
    ws['A9'].font = Font(bold=True, size=12, color=ORANGE)
    ws['A10'] = "Robert & Linda Chen"
    ws['A11'] = "789 Maple Drive"
    ws['A12'] = "Springfield, IL 62704"

    ws['A14'] = "PROJECT: Master Bathroom Remodel"
    ws['A14'].font = Font(bold=True, size=13)

    # MATERIALS section
    row = 16
    ws.cell(row=row, column=1, value="MATERIALS").font = Font(bold=True, size=12, color=ORANGE)
    row += 1
    for i, h in enumerate(["Item", "Qty", "Unit Cost", "Total"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 4)

    materials = [
        ("Porcelain floor tile (12×24) - 80 sq ft", 80, 4.50),
        ("Shower wall tile (subway) - 120 sq ft", 120, 3.75),
        ("Vanity - 48\" double sink", 1, 850),
        ("Toilet - Kohler Highline comfort height", 1, 285),
        ("Shower valve + trim kit (Moen)", 1, 320),
        ("Frameless glass shower door (48\")", 1, 1100),
        ("Exhaust fan - 110 CFM with light", 1, 175),
        ("Thinset, grout, backer board, supplies", 1, 280),
    ]
    mat_start = row + 1
    for i, (item, qty, cost) in enumerate(materials):
        r = mat_start + i
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=qty).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=cost).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
        ws.cell(row=r, column=4).number_format = CURRENCY_FMT
        style_data_row(ws, r, 4, alt=(i % 2 == 1))
    mat_end = mat_start + len(materials) - 1

    # LABOR section
    lr = mat_end + 2
    ws.cell(row=lr, column=1, value="LABOR").font = Font(bold=True, size=12, color=ORANGE)
    lr += 1
    for i, h in enumerate(["Task", "Hours", "Rate", "Total"], 1):
        ws.cell(row=lr, column=i, value=h)
    style_header_row(ws, lr, 4)

    labor = [
        ("Demo & haul-away", 8, 65),
        ("Plumbing rough-in", 12, 85),
        ("Electrical rough-in", 6, 80),
        ("Tile installation (floor + shower)", 20, 70),
        ("Vanity & fixture install", 6, 75),
        ("Painting & trim", 8, 55),
        ("Final clean & punch list", 4, 55),
    ]
    lab_start = lr + 1
    for i, (task, hrs, rate) in enumerate(labor):
        r = lab_start + i
        ws.cell(row=r, column=1, value=task)
        ws.cell(row=r, column=2, value=hrs).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=rate).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
        ws.cell(row=r, column=4).number_format = CURRENCY_FMT
        style_data_row(ws, r, 4, alt=(i % 2 == 1))
    lab_end = lab_start + len(labor) - 1

    # Totals
    tr = lab_end + 2
    ws.cell(row=tr, column=3, value="Materials Subtotal:").font = LABEL_FONT
    ws.cell(row=tr, column=4).value = f"=SUM(D{mat_start}:D{mat_end})"
    ws.cell(row=tr, column=4).number_format = CURRENCY_FMT

    ws.cell(row=tr+1, column=3, value="Labor Subtotal:").font = LABEL_FONT
    ws.cell(row=tr+1, column=4).value = f"=SUM(D{lab_start}:D{lab_end})"
    ws.cell(row=tr+1, column=4).number_format = CURRENCY_FMT

    ws.cell(row=tr+2, column=3, value="Subtotal:").font = LABEL_FONT
    ws.cell(row=tr+2, column=4).value = f"=D{tr}+D{tr+1}"
    ws.cell(row=tr+2, column=4).number_format = CURRENCY_FMT

    ws.cell(row=tr+3, column=3, value="Markup (15%):").font = LABEL_FONT
    ws.cell(row=tr+3, column=4).value = f"=D{tr+2}*0.15"
    ws.cell(row=tr+3, column=4).number_format = CURRENCY_FMT

    ws.cell(row=tr+4, column=3, value="Tax (8.25%):").font = LABEL_FONT
    ws.cell(row=tr+4, column=4).value = f"=(D{tr+2}+D{tr+3})*0.0825"
    ws.cell(row=tr+4, column=4).number_format = CURRENCY_FMT

    ws.cell(row=tr+5, column=3, value="GRAND TOTAL:").font = Font(bold=True, size=13, color=ORANGE)
    ws.cell(row=tr+5, column=4).value = f"=D{tr+2}+D{tr+3}+D{tr+4}"
    ws.cell(row=tr+5, column=4).number_format = CURRENCY_FMT
    ws.cell(row=tr+5, column=4).font = Font(bold=True, size=13, color=ORANGE)

    # Terms
    ttr = tr + 7
    ws.cell(row=ttr, column=1, value="TERMS & CONDITIONS").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=ttr+1, column=1, value="1. This estimate is valid for 30 days from the date above.")
    ws.cell(row=ttr+2, column=1, value="2. 50% deposit required to schedule work. Balance due upon completion.")
    ws.cell(row=ttr+3, column=1, value="3. Any changes to scope will be documented via Change Order.")
    ws.cell(row=ttr+4, column=1, value="4. Unforeseen conditions (rot, mold, code issues) billed at T&M.")
    ws.cell(row=ttr+5, column=1, value="5. All work guaranteed for 1 year from completion date.")

    ws.cell(row=ttr+7, column=1, value="ACCEPTANCE").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=ttr+8, column=1, value="Client Signature: ____________________________    Date: ____________")
    ws.cell(row=ttr+9, column=1, value="Contractor Signature: ________________________    Date: ____________")

    wb.save(os.path.join(DOWNLOADS, 'contractor-estimate.xlsx'))
    print("✅ contractor-estimate.xlsx")


# ============================================================
# 3. JOB COSTING TRACKER
# ============================================================
def create_job_costing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Costing"

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    ws['A1'] = "JOB COSTING TRACKER"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    ws['A3'] = "Job Name:"
    ws['A3'].font = LABEL_FONT
    ws['B3'] = "Miller Kitchen & Bath Remodel"
    ws['A4'] = "Client:"
    ws['A4'].font = LABEL_FONT
    ws['B4'] = "David & Karen Miller"
    ws['A5'] = "Start Date:"
    ws['A5'].font = LABEL_FONT
    ws['B5'] = "02/15/2026"
    ws['A6'] = "Estimated Total:"
    ws['A6'].font = LABEL_FONT
    ws['B6'] = 28500
    ws['B6'].number_format = CURRENCY_FMT

    row = 8
    headers = ["Category", "Description", "Estimated Cost", "Actual Cost", "Variance"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 5, fill=ORANGE_FILL)

    items = [
        ("Materials", "Cabinets - KraftMaid 12 linear ft", 4200, 4350),
        ("Materials", "Quartz countertops - Caesarstone", 3200, 3400),
        ("Materials", "Floor tile - porcelain 200 sq ft", 900, 880),
        ("Materials", "Plumbing fixtures (faucets, sinks)", 650, 720),
        ("Materials", "Electrical supplies & lighting", 480, 510),
        ("Materials", "Paint, primer, caulk, supplies", 320, 295),
        ("Materials", "Shower door - frameless glass", 1100, 1100),
        ("Labor", "Demo crew - 2 days", 1200, 1200),
        ("Labor", "Lead carpenter - 80 hrs", 4800, 5200),
        ("Labor", "Tile installer - 32 hrs", 2240, 2240),
        ("Labor", "Plumber - rough-in + finals", 1800, 2100),
        ("Labor", "Electrician - rough-in + finals", 1200, 1200),
        ("Labor", "Painter - 16 hrs", 880, 880),
        ("Equipment Rental", "Dumpster rental (20 yd)", 450, 450),
        ("Equipment Rental", "Tile saw rental - 1 week", 175, 175),
        ("Permits", "Building permit", 350, 350),
        ("Permits", "Plumbing permit", 150, 150),
        ("Subcontractors", "HVAC duct modification", 800, 950),
        ("Subcontractors", "Granite fabricator (template + cut)", 600, 600),
        ("Overhead", "Project management (5%)", 1250, 1250),
        ("Overhead", "Insurance allocation", 400, 400),
        ("Overhead", "Vehicle / fuel for job", 280, 310),
        ("Overhead", "Warranty reserve (2%)", 570, 570),
    ]

    for i, (cat, desc, est, act) in enumerate(items):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=est).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=act).number_format = CURRENCY_FMT
        ws.cell(row=r, column=5).value = f"=D{r}-C{r}"
        ws.cell(row=r, column=5).number_format = CURRENCY_FMT
        style_data_row(ws, r, 5, alt=(i % 2 == 1))

    # Summary
    last = row + len(items)
    sr = last + 2
    ws.cell(row=sr, column=2, value="TOTAL ESTIMATED:").font = Font(bold=True, size=12)
    ws.cell(row=sr, column=3).value = f"=SUM(C{row+1}:C{last})"
    ws.cell(row=sr, column=3).number_format = CURRENCY_FMT
    ws.cell(row=sr, column=3).font = Font(bold=True, size=12)

    ws.cell(row=sr+1, column=2, value="TOTAL ACTUAL:").font = Font(bold=True, size=12)
    ws.cell(row=sr+1, column=4).value = f"=SUM(D{row+1}:D{last})"
    ws.cell(row=sr+1, column=4).number_format = CURRENCY_FMT
    ws.cell(row=sr+1, column=4).font = Font(bold=True, size=12)

    ws.cell(row=sr+2, column=2, value="TOTAL VARIANCE:").font = Font(bold=True, size=12, color="DC2626")
    ws.cell(row=sr+2, column=5).value = f"=SUM(E{row+1}:E{last})"
    ws.cell(row=sr+2, column=5).number_format = CURRENCY_FMT
    ws.cell(row=sr+2, column=5).font = Font(bold=True, size=12, color="DC2626")

    ws.cell(row=sr+3, column=2, value="CONTRACT PRICE:").font = Font(bold=True, size=12)
    ws.cell(row=sr+3, column=4, value=28500).number_format = CURRENCY_FMT

    ws.cell(row=sr+4, column=2, value="PROFIT / (LOSS):").font = Font(bold=True, size=13, color=ORANGE)
    ws.cell(row=sr+4, column=4).value = f"=D{sr+3}-D{sr+1}"
    ws.cell(row=sr+4, column=4).number_format = CURRENCY_FMT
    ws.cell(row=sr+4, column=4).font = Font(bold=True, size=13, color=ORANGE)

    wb.save(os.path.join(DOWNLOADS, 'job-costing-tracker.xlsx'))
    print("✅ job-costing-tracker.xlsx")


# ============================================================
# 4. PROFIT & LOSS TRACKER
# ============================================================
def create_pnl():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws.column_dimensions['A'].width = 22
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws['A1'] = "PROFIT & LOSS STATEMENT"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:N1')

    ws['A2'] = "YOUR COMPANY NAME — 2026"
    ws['A2'].font = Font(bold=True, size=12, color="666666")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "YTD Total"]
    row = 4
    ws.cell(row=row, column=1, value="")
    for i, m in enumerate(months, 2):
        ws.cell(row=row, column=i, value=m)
    style_header_row(ws, row, 14, fill=ORANGE_FILL)

    # INCOME
    r = 5
    ws.cell(row=r, column=1, value="INCOME").font = Font(bold=True, size=12, color=ORANGE)

    income_items = ["Service Revenue", "Material Markups", "Other Income"]
    example_data = {
        "Service Revenue": [18500, 22000, 19800, 0,0,0,0,0,0,0,0,0],
        "Material Markups": [2800, 3400, 3100, 0,0,0,0,0,0,0,0,0],
        "Other Income": [500, 0, 250, 0,0,0,0,0,0,0,0,0],
    }
    income_start = r + 1
    for idx, item in enumerate(income_items):
        cr = income_start + idx
        ws.cell(row=cr, column=1, value=item)
        for m in range(12):
            ws.cell(row=cr, column=m+2, value=example_data[item][m]).number_format = CURRENCY_FMT
        # YTD formula
        ws.cell(row=cr, column=14).value = f"=SUM(B{cr}:M{cr})"
        ws.cell(row=cr, column=14).number_format = CURRENCY_FMT
        style_data_row(ws, cr, 14, alt=(idx % 2 == 1))
    income_end = income_start + len(income_items) - 1

    ti_row = income_end + 1
    ws.cell(row=ti_row, column=1, value="TOTAL INCOME").font = Font(bold=True, size=11)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=ti_row, column=c).value = f"=SUM({col_letter}{income_start}:{col_letter}{income_end})"
        ws.cell(row=ti_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=ti_row, column=c).font = Font(bold=True)
    style_header_row(ws, ti_row, 14, fill=PatternFill(start_color="334155", end_color="334155", fill_type="solid"))

    # EXPENSES
    er = ti_row + 2
    ws.cell(row=er, column=1, value="EXPENSES").font = Font(bold=True, size=12, color=ORANGE)

    expense_items = ["Materials", "Labor (wages/subs)", "Vehicle / Fuel", "Insurance",
                     "Tools & Equipment", "Marketing / Ads", "Office / Admin",
                     "Permits & Licenses", "Subcontractors", "Other Expenses"]
    exp_data = {
        "Materials": [8200, 9500, 8800, 0,0,0,0,0,0,0,0,0],
        "Labor (wages/subs)": [6500, 7800, 7200, 0,0,0,0,0,0,0,0,0],
        "Vehicle / Fuel": [650, 700, 680, 0,0,0,0,0,0,0,0,0],
        "Insurance": [450, 450, 450, 0,0,0,0,0,0,0,0,0],
        "Tools & Equipment": [300, 150, 500, 0,0,0,0,0,0,0,0,0],
        "Marketing / Ads": [200, 250, 200, 0,0,0,0,0,0,0,0,0],
        "Office / Admin": [150, 150, 150, 0,0,0,0,0,0,0,0,0],
        "Permits & Licenses": [350, 200, 150, 0,0,0,0,0,0,0,0,0],
        "Subcontractors": [1200, 2000, 800, 0,0,0,0,0,0,0,0,0],
        "Other Expenses": [100, 75, 120, 0,0,0,0,0,0,0,0,0],
    }
    exp_start = er + 1
    for idx, item in enumerate(expense_items):
        cr = exp_start + idx
        ws.cell(row=cr, column=1, value=item)
        for m in range(12):
            ws.cell(row=cr, column=m+2, value=exp_data[item][m]).number_format = CURRENCY_FMT
        ws.cell(row=cr, column=14).value = f"=SUM(B{cr}:M{cr})"
        ws.cell(row=cr, column=14).number_format = CURRENCY_FMT
        style_data_row(ws, cr, 14, alt=(idx % 2 == 1))
    exp_end = exp_start + len(expense_items) - 1

    te_row = exp_end + 1
    ws.cell(row=te_row, column=1, value="TOTAL EXPENSES").font = Font(bold=True, size=11)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=te_row, column=c).value = f"=SUM({col_letter}{exp_start}:{col_letter}{exp_end})"
        ws.cell(row=te_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=te_row, column=c).font = Font(bold=True)
    style_header_row(ws, te_row, 14, fill=PatternFill(start_color="334155", end_color="334155", fill_type="solid"))

    # NET PROFIT
    np_row = te_row + 2
    ws.cell(row=np_row, column=1, value="NET PROFIT").font = Font(bold=True, size=13, color=ORANGE)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=np_row, column=c).value = f"={col_letter}{ti_row}-{col_letter}{te_row}"
        ws.cell(row=np_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=np_row, column=c).font = Font(bold=True, size=12, color=ORANGE)

    wb.save(os.path.join(DOWNLOADS, 'profit-loss-tracker.xlsx'))
    print("✅ profit-loss-tracker.xlsx")


# ============================================================
# 5. CLIENT TRACKER
# ============================================================
def create_client_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Tracker"

    cols = ["Client Name", "Phone", "Email", "Address", "Lead Source", "Job Type",
            "Status", "Quoted Amount", "Paid Amount", "Follow-up Date", "Notes"]
    widths = [22, 16, 26, 30, 14, 16, 12, 15, 15, 14, 30]

    ws['A1'] = "CLIENT & LEAD TRACKER"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')

    row = 3
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        ws.cell(row=row, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header_row(ws, row, 11, fill=ORANGE_FILL)

    clients = [
        ("Tom Richardson", "(555) 234-5678", "tom.r@email.com", "123 Oak St, Austin TX", "Google Ads", "Kitchen Remodel", "Active", 32000, 16000, "03/15/2026", "Started demo, Phase 1 underway"),
        ("Sarah Martinez", "(555) 345-6789", "sarah.m@email.com", "456 Pine Ave, Austin TX", "Referral", "Bathroom Remodel", "Quoted", 18500, 0, "03/12/2026", "Sent estimate, follow up Thurs"),
        ("Mike Johnson", "(555) 456-7890", "mike.j@email.com", "789 Elm Blvd, Round Rock TX", "Angi", "Deck Build", "Lead", 0, 0, "03/10/2026", "Called, wants free estimate"),
        ("Jennifer Park", "(555) 567-8901", "j.park@email.com", "321 Birch Ln, Cedar Park TX", "Website", "Whole Home Reno", "Active", 85000, 42500, "03/20/2026", "Phase 2 - electrical rough-in"),
        ("David Thompson", "(555) 678-9012", "david.t@email.com", "654 Maple Dr, Georgetown TX", "Nextdoor", "Fence Install", "Complete", 4800, 4800, "N/A", "Job complete, asked for review"),
        ("Lisa Wong", "(555) 789-0123", "lisa.w@email.com", "987 Cedar Ct, Pflugerville TX", "Referral", "Flooring", "Quoted", 12000, 0, "03/14/2026", "Choosing between LVP and hardwood"),
        ("Robert Garcia", "(555) 890-1234", "r.garcia@email.com", "246 Walnut St, Austin TX", "Google Organic", "Garage Conversion", "Lead", 0, 0, "03/11/2026", "Wants ADU info, send brochure"),
        ("Amanda Lewis", "(555) 901-2345", "amanda.l@email.com", "135 Spruce Way, Lakeway TX", "Home Show", "Patio Cover", "Active", 9200, 4600, "03/18/2026", "Permit approved, starting Mon"),
        ("Chris Baker", "(555) 012-3456", "chris.b@email.com", "864 Ash Blvd, Bee Cave TX", "Thumbtack", "Roof Repair", "Complete", 2800, 2800, "N/A", "Warranty card sent"),
        ("Karen Foster", "(555) 123-7890", "karen.f@email.com", "579 Hickory Pl, Leander TX", "Yard Sign", "Siding Replace", "Quoted", 14500, 0, "03/16/2026", "Getting 2nd quote, price sensitive"),
    ]

    # Status colors
    status_colors = {
        "Lead": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "Quoted": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Active": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Complete": PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"),
    }

    for i, c in enumerate(clients):
        r = row + 1 + i
        for j, val in enumerate(c, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = THIN_BORDER
            if j == 8 or j == 9:
                cell.number_format = CURRENCY_FMT
            if j == 7 and val in status_colors:
                cell.fill = status_colors[val]
                cell.font = Font(bold=True, size=11)

    wb.save(os.path.join(DOWNLOADS, 'client-tracker.xlsx'))
    print("✅ client-tracker.xlsx")


# ============================================================
# 6. CONTRACTOR PROPOSAL
# ============================================================
def create_proposal():
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposal"

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws['A1'] = "PROJECT PROPOSAL"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    ws['A3'] = "YOUR COMPANY NAME"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A4'] = "Lic #: XXXX-XXXXX | Insured & Bonded"
    ws['A5'] = "123 Main Street, Your City, ST 12345"
    ws['A6'] = "(555) 123-4567 | info@yourcompany.com"

    ws['A8'] = "PREPARED FOR:"
    ws['A8'].font = Font(bold=True, color=ORANGE)
    ws['A9'] = "Client Name: ________________________________"
    ws['A10'] = "Address: ____________________________________"
    ws['A11'] = "Phone: ____________  Email: _________________"

    ws['C8'] = "Proposal #:"
    ws['C8'].font = LABEL_FONT
    ws['D8'] = "PROP-2026-015"
    ws['C9'] = "Date:"
    ws['C9'].font = LABEL_FONT
    ws['D9'] = "03/09/2026"

    ws['A13'] = "PROJECT: Basement Finishing — 800 sq ft"
    ws['A13'].font = Font(bold=True, size=14)
    ws.merge_cells('A13:D13')

    # Scope of work
    ws['A15'] = "SCOPE OF WORK"
    ws['A15'].font = Font(bold=True, size=12, color=ORANGE)
    scope = [
        "Frame interior walls per approved floor plan (2 bedrooms, 1 bath, living area)",
        "Install electrical wiring — 12 outlets, 8 recessed lights, 2 switches, panel sub-feed",
        "Plumbing rough-in for full bathroom (shower, toilet, vanity)",
        "Install R-19 insulation in all exterior walls and rim joist",
        "Hang and finish drywall throughout (approx. 2,400 sq ft of board)",
        "Install LVP flooring — 800 sq ft (waterproof, click-lock)",
        "Tile bathroom floor and shower (60 sq ft tile)",
        "Install interior doors (4), trim, and baseboards",
        "Paint all walls and ceilings — 2 coats (customer selects colors)",
        "Final electrical — fixtures, outlets, switches, smoke detectors",
    ]
    for i, s in enumerate(scope, 1):
        ws.cell(row=15+i, column=1, value=i)
        ws.cell(row=15+i, column=2, value=s)
        ws.merge_cells(f'B{15+i}:D{15+i}')

    # Timeline
    tl_row = 27
    ws.cell(row=tl_row, column=1, value="PROJECT TIMELINE").font = Font(bold=True, size=12, color=ORANGE)
    tl_row += 1
    for i, h in enumerate(["#", "Milestone", "Duration", "Target Date"], 1):
        ws.cell(row=tl_row, column=i, value=h)
    style_header_row(ws, tl_row, 4)

    milestones = [
        (1, "Permits & material ordering", "1 week", "Week 1"),
        (2, "Framing & rough-in (plumb/elec/HVAC)", "2 weeks", "Week 2-3"),
        (3, "Insulation & drywall", "1.5 weeks", "Week 4-5"),
        (4, "Flooring & tile", "1 week", "Week 5-6"),
        (5, "Trim, doors, paint", "1 week", "Week 6-7"),
        (6, "Final fixtures & punch list", "3 days", "Week 7"),
        (7, "Final inspection & walkthrough", "1 day", "Week 8"),
    ]
    for i, (n, desc, dur, target) in enumerate(milestones):
        r = tl_row + 1 + i
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=dur)
        ws.cell(row=r, column=4, value=target)
        style_data_row(ws, r, 4, alt=(i % 2 == 1))

    # Pricing
    pr = tl_row + 1 + len(milestones) + 1
    ws.cell(row=pr, column=1, value="PRICING").font = Font(bold=True, size=12, color=ORANGE)
    pr += 1
    for i, h in enumerate(["#", "Item", "Cost", ""], 1):
        ws.cell(row=pr, column=i, value=h)
    style_header_row(ws, pr, 3)

    pricing = [
        (1, "Framing & structural", 4800),
        (2, "Electrical (rough + finish)", 5200),
        (3, "Plumbing (rough + finish)", 3800),
        (4, "Insulation", 1600),
        (5, "Drywall (hang, tape, finish)", 4200),
        (6, "Flooring - LVP install", 3600),
        (7, "Tile - bathroom", 2200),
        (8, "Trim, doors, paint", 3800),
        (9, "Permits & inspections", 650),
        (10, "Project management & cleanup", 1500),
    ]
    pr_start = pr + 1
    for i, (n, item, cost) in enumerate(pricing):
        r = pr_start + i
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=cost).number_format = CURRENCY_FMT
        style_data_row(ws, r, 3, alt=(i % 2 == 1))
    pr_end = pr_start + len(pricing) - 1

    tot_r = pr_end + 1
    ws.cell(row=tot_r, column=2, value="TOTAL PROJECT COST").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=tot_r, column=3).value = f"=SUM(C{pr_start}:C{pr_end})"
    ws.cell(row=tot_r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=tot_r, column=3).font = Font(bold=True, size=12, color=ORANGE)

    # Payment schedule
    ps = tot_r + 2
    ws.cell(row=ps, column=1, value="PAYMENT SCHEDULE").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=ps+1, column=1, value="1.")
    ws.cell(row=ps+1, column=2, value=f"Deposit (33%) — Due upon signing: =C{tot_r}*0.33")
    ws.cell(row=ps+2, column=1, value="2.")
    ws.cell(row=ps+2, column=2, value=f"Progress payment (33%) — Due at drywall completion")
    ws.cell(row=ps+3, column=1, value="3.")
    ws.cell(row=ps+3, column=2, value=f"Final payment (34%) — Due upon completion & walkthrough")

    # Terms
    tc = ps + 5
    ws.cell(row=tc, column=1, value="TERMS & CONDITIONS").font = Font(bold=True, size=12, color=ORANGE)
    terms = [
        "1. This proposal is valid for 30 days.",
        "2. Price does not include unforeseen structural, mold, or code issues.",
        "3. Change orders will be documented and priced separately.",
        "4. Contractor carries general liability and workers comp insurance.",
        "5. All work to meet local building codes; inspections included.",
        "6. 1-year workmanship warranty from date of completion.",
    ]
    for i, t in enumerate(terms):
        ws.cell(row=tc+1+i, column=1, value=t)
        ws.merge_cells(f'A{tc+1+i}:D{tc+1+i}')

    sig = tc + len(terms) + 2
    ws.cell(row=sig, column=1, value="ACCEPTANCE").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=sig+1, column=1, value="I accept this proposal and authorize work to begin.")
    ws.cell(row=sig+3, column=1, value="Client Signature: ____________________________    Date: ____________")
    ws.cell(row=sig+4, column=1, value="Contractor Signature: ________________________    Date: ____________")

    wb.save(os.path.join(DOWNLOADS, 'contractor-proposal.xlsx'))
    print("✅ contractor-proposal.xlsx")


# ============================================================
# 7. CHANGE ORDER
# ============================================================
def create_change_order():
    wb = Workbook()
    ws = wb.active
    ws.title = "Change Order"

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws['A1'] = "CHANGE ORDER"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    ws['A3'] = "YOUR COMPANY NAME"
    ws['A3'].font = Font(bold=True, size=14)
    ws['A4'] = "(555) 123-4567 | info@yourcompany.com"

    ws['C3'] = "Change Order #:"
    ws['C3'].font = LABEL_FONT
    ws['D3'] = "CO-001"
    ws['C4'] = "Date:"
    ws['C4'].font = LABEL_FONT
    ws['D4'] = "03/09/2026"
    ws['C5'] = "Original Contract #:"
    ws['C5'].font = LABEL_FONT
    ws['D5'] = "PROP-2026-015"

    ws['A7'] = "CLIENT:"
    ws['A7'].font = LABEL_FONT
    ws['B7'] = "Client Name"
    ws['A8'] = "PROJECT:"
    ws['A8'].font = LABEL_FONT
    ws['B8'] = "Basement Finishing"

    ws['A10'] = "ORIGINAL CONTRACT AMOUNT:"
    ws['A10'].font = LABEL_FONT
    ws['C10'] = 31350
    ws['C10'].number_format = CURRENCY_FMT

    ws['A12'] = "DESCRIPTION OF CHANGES"
    ws['A12'].font = Font(bold=True, size=12, color=ORANGE)
    ws['A13'] = "Client requests the following additions/modifications to the original scope:"
    ws.merge_cells('A13:D13')

    # Cost impact table
    row = 15
    for i, h in enumerate(["#", "Item Added / Removed", "Add", "Deduct"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 4, fill=ORANGE_FILL)

    changes = [
        (1, "Add wet bar with sink in living area", 3200, 0),
        (2, "Upgrade flooring from LVP to engineered hardwood", 1800, 0),
        (3, "Add egress window to bedroom 2", 2400, 0),
        (4, "Remove built-in shelving from original scope", 0, 800),
        (5, "Add recessed lighting (4 additional cans)", 680, 0),
    ]
    ch_start = row + 1
    for i, (n, desc, add, ded) in enumerate(changes):
        r = ch_start + i
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=add if add else "").number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=ded if ded else "").number_format = CURRENCY_FMT
        style_data_row(ws, r, 4, alt=(i % 2 == 1))
    ch_end = ch_start + len(changes) - 1

    sr = ch_end + 2
    ws.cell(row=sr, column=2, value="Total Additions:").font = LABEL_FONT
    ws.cell(row=sr, column=3).value = f"=SUM(C{ch_start}:C{ch_end})"
    ws.cell(row=sr, column=3).number_format = CURRENCY_FMT

    ws.cell(row=sr+1, column=2, value="Total Deductions:").font = LABEL_FONT
    ws.cell(row=sr+1, column=4).value = f"=SUM(D{ch_start}:D{ch_end})"
    ws.cell(row=sr+1, column=4).number_format = CURRENCY_FMT

    ws.cell(row=sr+2, column=2, value="Net Change:").font = Font(bold=True, size=12)
    ws.cell(row=sr+2, column=3).value = f"=C{sr}-D{sr+1}"
    ws.cell(row=sr+2, column=3).number_format = CURRENCY_FMT
    ws.cell(row=sr+2, column=3).font = Font(bold=True, size=12)

    ws.cell(row=sr+4, column=2, value="NEW CONTRACT TOTAL:").font = Font(bold=True, size=13, color=ORANGE)
    ws.cell(row=sr+4, column=3).value = f"=C10+C{sr+2}"
    ws.cell(row=sr+4, column=3).number_format = CURRENCY_FMT
    ws.cell(row=sr+4, column=3).font = Font(bold=True, size=13, color=ORANGE)

    ws.cell(row=sr+5, column=2, value="Schedule Impact:").font = LABEL_FONT
    ws.cell(row=sr+5, column=3, value="Add 5 working days")

    sig = sr + 7
    ws.cell(row=sig, column=1, value="APPROVAL").font = Font(bold=True, size=12, color=ORANGE)
    ws.cell(row=sig+1, column=1, value="Both parties agree to the changes described above.")
    ws.cell(row=sig+3, column=1, value="Client Signature: ____________________________    Date: ____________")
    ws.cell(row=sig+4, column=1, value="Contractor Signature: ________________________    Date: ____________")

    wb.save(os.path.join(DOWNLOADS, 'change-order.xlsx'))
    print("✅ change-order.xlsx")


# ============================================================
# 8. WEEKLY TIMESHEET
# ============================================================
def create_timesheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 22

    ws['A1'] = "WEEKLY TIMESHEET"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')

    ws['A3'] = "Employee:"
    ws['A3'].font = LABEL_FONT
    ws['B3'] = "Mike Torres"
    ws['D3'] = "Employee ID:"
    ws['D3'].font = LABEL_FONT
    ws['E3'] = "EMP-042"
    ws['A4'] = "Position:"
    ws['A4'].font = LABEL_FONT
    ws['B4'] = "Lead Carpenter"
    ws['D4'] = "Pay Rate:"
    ws['D4'].font = LABEL_FONT
    ws['E4'] = 32.00
    ws['E4'].number_format = CURRENCY_FMT

    def write_week(ws, start_row, week_label, data):
        ws.cell(row=start_row, column=1, value=week_label).font = Font(bold=True, size=12, color=ORANGE)
        hr = start_row + 1
        headers = ["Day", "Time In", "Lunch Out", "Lunch In", "Time Out", "Regular Hrs", "OT Hrs", "Job / Project"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=hr, column=i, value=h)
        style_header_row(ws, hr, 8)

        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i, (day, d) in enumerate(zip(days, data)):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=day)
            ws.cell(row=r, column=2, value=d[0])
            ws.cell(row=r, column=3, value=d[1])
            ws.cell(row=r, column=4, value=d[2])
            ws.cell(row=r, column=5, value=d[3])
            ws.cell(row=r, column=6, value=d[4])
            ws.cell(row=r, column=7, value=d[5])
            ws.cell(row=r, column=8, value=d[6])
            style_data_row(ws, r, 8, alt=(i % 2 == 1))

        tot_r = hr + 8
        ws.cell(row=tot_r, column=5, value="TOTALS:").font = Font(bold=True)
        ws.cell(row=tot_r, column=6).value = f"=SUM(F{hr+1}:F{hr+7})"
        ws.cell(row=tot_r, column=6).font = Font(bold=True)
        ws.cell(row=tot_r, column=7).value = f"=SUM(G{hr+1}:G{hr+7})"
        ws.cell(row=tot_r, column=7).font = Font(bold=True)
        return tot_r + 1

    week1_data = [
        ("7:00 AM", "12:00 PM", "12:30 PM", "4:30 PM", 8, 1, "Miller Kitchen Remodel"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "4:30 PM", 8, 1, "Miller Kitchen Remodel"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "3:30 PM", 8, 0, "Miller Kitchen Remodel"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "4:30 PM", 8, 1, "Chen Bathroom"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "3:30 PM", 8, 0, "Chen Bathroom"),
        ("8:00 AM", "", "", "12:00 PM", 4, 0, "Punch list - Thompson fence"),
        ("", "", "", "", 0, 0, "OFF"),
    ]
    week2_data = [
        ("6:30 AM", "12:00 PM", "12:30 PM", "4:00 PM", 8, 1.5, "Chen Bathroom"),
        ("6:30 AM", "12:00 PM", "12:30 PM", "4:00 PM", 8, 1.5, "Chen Bathroom"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "3:30 PM", 8, 0, "Park Whole Home Reno"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "5:00 PM", 8, 2, "Park Whole Home Reno"),
        ("7:00 AM", "12:00 PM", "12:30 PM", "3:30 PM", 8, 0, "Park Whole Home Reno"),
        ("", "", "", "", 0, 0, "OFF"),
        ("", "", "", "", 0, 0, "OFF"),
    ]

    next_row = write_week(ws, 6, "WEEK 1: March 2-8, 2026", week1_data)
    write_week(ws, next_row + 1, "WEEK 2: March 9-15, 2026", week2_data)

    wb.save(os.path.join(DOWNLOADS, 'weekly-timesheet.xlsx'))
    print("✅ weekly-timesheet.xlsx")


# ============================================================
# 9. CASH FLOW FORECAST
# ============================================================
def create_cash_flow():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Forecast"

    ws.column_dimensions['A'].width = 22
    for i in range(2, 14):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws['A1'] = "CASH FLOW FORECAST"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:M1')
    ws['A2'] = "YOUR COMPANY NAME — 2026"
    ws['A2'].font = Font(bold=True, size=12, color="666666")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    row = 4
    ws.cell(row=row, column=1, value="")
    for i, m in enumerate(months, 2):
        ws.cell(row=row, column=i, value=m)
    style_header_row(ws, row, 13, fill=ORANGE_FILL)

    # Opening balance
    r = 5
    ws.cell(row=r, column=1, value="OPENING BALANCE").font = Font(bold=True, size=11, color=ORANGE)
    ws.cell(row=r, column=2, value=15000).number_format = CURRENCY_FMT
    for c in range(3, 14):
        # Opening = previous month's closing
        pass  # Will set after we know closing row

    # Cash In
    r = 7
    ws.cell(row=r, column=1, value="CASH IN").font = Font(bold=True, size=12, color=ORANGE)
    cash_in_items = ["Customer Payments", "Other Income"]
    cash_in_data = {
        "Customer Payments": [16000, 22000, 18000, 20000, 25000, 28000, 30000, 27000, 24000, 21000, 18000, 15000],
        "Other Income": [500, 0, 300, 0, 200, 0, 0, 500, 0, 200, 0, 0],
    }
    ci_start = r + 1
    for idx, item in enumerate(cash_in_items):
        cr = ci_start + idx
        ws.cell(row=cr, column=1, value=item)
        for m in range(12):
            ws.cell(row=cr, column=m+2, value=cash_in_data[item][m]).number_format = CURRENCY_FMT
        style_data_row(ws, cr, 13, alt=(idx % 2 == 1))
    ci_end = ci_start + len(cash_in_items) - 1

    ti_row = ci_end + 1
    ws.cell(row=ti_row, column=1, value="TOTAL CASH IN").font = Font(bold=True)
    for c in range(2, 14):
        cl = get_column_letter(c)
        ws.cell(row=ti_row, column=c).value = f"=SUM({cl}{ci_start}:{cl}{ci_end})"
        ws.cell(row=ti_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=ti_row, column=c).font = Font(bold=True)

    # Cash Out
    co_label = ti_row + 2
    ws.cell(row=co_label, column=1, value="CASH OUT").font = Font(bold=True, size=12, color=ORANGE)
    cash_out_items = ["Materials", "Labor / Payroll", "Overhead / Rent", "Equipment", "Marketing", "Insurance", "Taxes (quarterly)"]
    cash_out_data = {
        "Materials": [7000, 9500, 8000, 8500, 10000, 11000, 12000, 10500, 9500, 8500, 7500, 6000],
        "Labor / Payroll": [5500, 7000, 6500, 7000, 8500, 9000, 10000, 9000, 8000, 7000, 6000, 5000],
        "Overhead / Rent": [1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200],
        "Equipment": [500, 200, 800, 300, 600, 200, 400, 300, 200, 500, 200, 100],
        "Marketing": [300, 300, 300, 400, 500, 500, 500, 400, 300, 300, 200, 200],
        "Insurance": [450, 450, 450, 450, 450, 450, 450, 450, 450, 450, 450, 450],
        "Taxes (quarterly)": [0, 0, 3500, 0, 0, 4000, 0, 0, 4500, 0, 0, 4000],
    }
    co_start = co_label + 1
    for idx, item in enumerate(cash_out_items):
        cr = co_start + idx
        ws.cell(row=cr, column=1, value=item)
        for m in range(12):
            ws.cell(row=cr, column=m+2, value=cash_out_data[item][m]).number_format = CURRENCY_FMT
        style_data_row(ws, cr, 13, alt=(idx % 2 == 1))
    co_end = co_start + len(cash_out_items) - 1

    to_row = co_end + 1
    ws.cell(row=to_row, column=1, value="TOTAL CASH OUT").font = Font(bold=True)
    for c in range(2, 14):
        cl = get_column_letter(c)
        ws.cell(row=to_row, column=c).value = f"=SUM({cl}{co_start}:{cl}{co_end})"
        ws.cell(row=to_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=to_row, column=c).font = Font(bold=True)

    # Net Cash Flow
    nc_row = to_row + 2
    ws.cell(row=nc_row, column=1, value="NET CASH FLOW").font = Font(bold=True, size=11)
    for c in range(2, 14):
        cl = get_column_letter(c)
        ws.cell(row=nc_row, column=c).value = f"={cl}{ti_row}-{cl}{to_row}"
        ws.cell(row=nc_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=nc_row, column=c).font = Font(bold=True)

    # Closing Balance
    cb_row = nc_row + 1
    ws.cell(row=cb_row, column=1, value="CLOSING BALANCE").font = Font(bold=True, size=12, color=ORANGE)
    for c in range(2, 14):
        cl = get_column_letter(c)
        ws.cell(row=cb_row, column=c).value = f"={cl}5+{cl}{nc_row}"
        ws.cell(row=cb_row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=cb_row, column=c).font = Font(bold=True, size=12, color=ORANGE)

    # Now set opening balance formulas (month 2+ = previous closing)
    for c in range(3, 14):
        prev_cl = get_column_letter(c - 1)
        ws.cell(row=5, column=c).value = f"={prev_cl}{cb_row}"
        ws.cell(row=5, column=c).number_format = CURRENCY_FMT

    wb.save(os.path.join(DOWNLOADS, 'cash-flow-forecast.xlsx'))
    print("✅ cash-flow-forecast.xlsx")


# ============================================================
# RUN ALL
# ============================================================
if __name__ == "__main__":
    create_invoice()
    create_estimate()
    create_job_costing()
    create_pnl()
    create_client_tracker()
    create_proposal()
    create_change_order()
    create_timesheet()
    create_cash_flow()
    print("\n🎉 All templates generated!")
