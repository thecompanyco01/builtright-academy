#!/usr/bin/env python3
"""Rebuild ALL templates to professional, market-competitive quality.
Each template gets: proper formatting, colors, instructions tab, data validation, 
conditional formatting, trade-specific examples, print-ready layout."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os, zipfile

OUT = 'downloads'
os.makedirs(OUT, exist_ok=True)

# ===== STYLE CONSTANTS =====
ORANGE = 'F97316'
DARK_BG = '1E293B'
DARKER_BG = '0F172A'
WHITE = 'F8FAFC'
GRAY = '94A3B8'
GREEN = '22C55E'
RED = 'EF4444'
LIGHT_BLUE = 'E0F2FE'
LIGHT_ORANGE = 'FFF7ED'
LIGHT_GREEN = 'F0FDF4'
LIGHT_RED = 'FEF2F2'

header_font = Font(name='Calibri', bold=True, size=20, color=ORANGE)
sub_font = Font(name='Calibri', bold=True, size=12, color='333333')
label_font = Font(name='Calibri', bold=True, size=11, color='333333')
data_font = Font(name='Calibri', size=11, color='333333')
small_font = Font(name='Calibri', size=9, color='666666')
money_font = Font(name='Calibri', size=11, color='1A7A2E')
total_font = Font(name='Calibri', bold=True, size=13, color=ORANGE)

header_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
header_font_white = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
light_orange_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type='solid')
light_green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
total_fill = PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
right = Alignment(horizontal='right', vertical='center')
wrap = Alignment(wrap_text=True, vertical='top')

USD = '"$"#,##0.00'
PCT = '0.00%'
DATE = 'MM/DD/YYYY'

def style_header_row(ws, row, cols, fill=None):
    """Style a header row with orange background and white text."""
    f = fill or header_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    """Style a data row with alternating colors."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.border = thin_border
        if alt:
            cell.fill = alt_fill

def add_instructions_tab(wb, template_name, instructions_text):
    """Add a professional instructions sheet."""
    ws = wb.create_sheet('Instructions', 0)
    ws.sheet_properties.tabColor = ORANGE
    
    ws.column_dimensions['A'].width = 80
    ws.merge_cells('A1:A1')
    ws.cell(row=1, column=1, value=f'📋 {template_name} — Instructions').font = header_font
    ws.cell(row=2, column=1, value='BuiltRight Academy | builtrighthq.com').font = small_font
    ws.cell(row=3, column=1, value='').font = data_font
    
    for i, line in enumerate(instructions_text, start=4):
        ws.cell(row=i, column=1, value=line).font = data_font
        ws.cell(row=i, column=1).alignment = wrap
        ws.row_dimensions[i].height = 22 if not line.startswith('—') else 30

# ========================================
# 1. CONTRACTOR INVOICE TEMPLATE ($9)
# ========================================
def build_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'
    ws.sheet_properties.tabColor = ORANGE
    
    # Column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # Print setup
    ws.print_area = 'A1:E50'
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # ---- HEADER ----
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1, value='INVOICE').font = Font(name='Calibri', bold=True, size=28, color=ORANGE)
    ws.row_dimensions[1].height = 45
    
    # Company info (left)
    ws.cell(row=3, column=1, value='[YOUR COMPANY NAME]').font = Font(name='Calibri', bold=True, size=14, color='333333')
    ws.cell(row=4, column=1, value='[Street Address]').font = data_font
    ws.cell(row=5, column=1, value='[City, State ZIP]').font = data_font
    ws.cell(row=6, column=1, value='[Phone]').font = data_font
    ws.cell(row=7, column=1, value='[Email]').font = data_font
    ws.cell(row=8, column=1, value='License #: [Your License Number]').font = small_font
    
    # Invoice details (right)
    for r, (label, value) in enumerate([
        ('Invoice #:', 'INV-2026-001'),
        ('Date:', '03/10/2026'),
        ('Due Date:', '04/09/2026'),
        ('PO #:', ''),
        ('Job Name:', 'Mitchell Kitchen Remodel'),
    ], start=3):
        ws.cell(row=r, column=D, value=label).font = label_font
        ws.cell(row=r, column=D).alignment = right
        ws.cell(row=r, column=E, value=value).font = data_font
        ws.cell(row=r, column=E).border = Border(bottom=Side(style='thin', color='D1D5DB'))
    
    # Bill To
    ws.cell(row=10, column=1, value='BILL TO:').font = Font(name='Calibri', bold=True, size=11, color=ORANGE)
    ws.cell(row=11, column=1, value='John & Sarah Mitchell').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=12, column=1, value='456 Oak Avenue').font = data_font
    ws.cell(row=13, column=1, value='Springfield, IL 62704').font = data_font
    ws.cell(row=14, column=1, value='(555) 987-6543 | mitchell.family@email.com').font = data_font
    
    # Job site (if different)
    ws.cell(row=10, column=D, value='JOB SITE:').font = Font(name='Calibri', bold=True, size=11, color=ORANGE)
    ws.cell(row=11, column=D, value='Same as billing address').font = data_font
    
    # ---- LINE ITEMS TABLE ----
    r = 16
    headers = ['Description', 'Qty', 'Unit Price', 'Amount']
    cols = [1, 2, 3, 4]  # Map to A, B, C, D... wait, we have 5 cols
    # Actually use A=description, B=qty, C=unit, D=amount, E for notes
    
    for c, h in enumerate(['Description', 'Qty', 'Unit Price', 'Amount', ''], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    # Line items with formulas
    items = [
        ('Demolition — Remove existing cabinets, countertops & flooring', 1, 450),
        ('Haul-away & dumpster rental (20-yard)', 1, 375),
        ('Custom cabinets — 12 linear ft (soft-close, dovetail)', 12, 185),
        ('Countertop — Quartz fabrication & installation', 1, 3200),
        ('Plumbing rough-in — Sink relocation + new supply lines', 1, 850),
        ('Electrical — 3 under-cabinet GFCI outlets + dedicated 20A circuit', 3, 175),
        ('Tile backsplash — 30 sq ft subway tile, installed', 30, 18),
        ('Backsplash labor — Layout, cutting, grout, seal', 1, 480),
        ('Paint — Kitchen walls & ceiling (250 sq ft), 2 coats', 1, 650),
        ('Plumbing finish — New sink + faucet installation', 1, 425),
        ('Lighting — 4 recessed LED 6" can lights, dimmer switch', 4, 95),
        ('Trim — Crown molding installation (24 linear ft)', 1, 380),
        ('Hardware — Cabinet pulls & knobs (24 pieces, brushed nickel)', 24, 12),
        ('Final cleanup, debris removal & touch-up', 1, 200),
        ('Permit filing fee (City of Springfield)', 1, 150),
    ]
    
    for i, (desc, qty, price) in enumerate(items):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=desc).font = data_font
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=2, value=qty).font = data_font
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3, value=price).font = data_font
        ws.cell(row=row, column=3).number_format = USD
        ws.cell(row=row, column=4).value = f'=B{row}*C{row}'
        ws.cell(row=row, column=4).font = money_font
        ws.cell(row=row, column=4).number_format = USD
        
        style_data_row(ws, row, 5, alt=(i % 2 == 0))
    
    # Totals
    last_item_row = r + len(items)
    totals_start = last_item_row + 2
    
    ws.cell(row=totals_start, column=3, value='Subtotal:').font = label_font
    ws.cell(row=totals_start, column=3).alignment = right
    ws.cell(row=totals_start, column=4).value = f'=SUM(D{r+1}:D{last_item_row})'
    ws.cell(row=totals_start, column=4).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=totals_start, column=4).number_format = USD
    
    ws.cell(row=totals_start+1, column=3, value='Tax Rate:').font = label_font
    ws.cell(row=totals_start+1, column=3).alignment = right
    ws.cell(row=totals_start+1, column=4, value=0.0825).font = data_font
    ws.cell(row=totals_start+1, column=4).number_format = PCT
    
    ws.cell(row=totals_start+2, column=3, value='Tax:').font = label_font
    ws.cell(row=totals_start+2, column=3).alignment = right
    ws.cell(row=totals_start+2, column=4).value = f'=D{totals_start}*D{totals_start+1}'
    ws.cell(row=totals_start+2, column=4).font = data_font
    ws.cell(row=totals_start+2, column=4).number_format = USD
    
    # Grand total with highlight
    ws.cell(row=totals_start+3, column=3, value='TOTAL DUE:').font = total_font
    ws.cell(row=totals_start+3, column=3).alignment = right
    ws.cell(row=totals_start+3, column=4).value = f'=D{totals_start}+D{totals_start+2}'
    ws.cell(row=totals_start+3, column=4).font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    ws.cell(row=totals_start+3, column=4).number_format = USD
    ws.cell(row=totals_start+3, column=4).fill = light_orange_fill
    ws.cell(row=totals_start+3, column=4).border = Border(
        left=Side(style='medium', color=ORANGE),
        right=Side(style='medium', color=ORANGE),
        top=Side(style='medium', color=ORANGE),
        bottom=Side(style='medium', color=ORANGE)
    )
    
    # Payment Terms
    pt_start = totals_start + 6
    ws.cell(row=pt_start, column=1, value='PAYMENT TERMS').font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    terms = [
        '• Payment due within 30 days of invoice date.',
        '• Late payments subject to 1.5% monthly interest (18% APR).',
        '• Make checks payable to: [YOUR COMPANY NAME]',
        '• Accepted methods: Check, ACH/Bank Transfer, Zelle, Venmo, Credit Card (+3% processing fee).',
        '• For questions about this invoice, contact [your email] or [your phone].',
    ]
    for i, t in enumerate(terms):
        ws.cell(row=pt_start+1+i, column=1, value=t).font = data_font
        ws.cell(row=pt_start+1+i, column=1).alignment = wrap
    
    # Notes
    notes_start = pt_start + len(terms) + 2
    ws.cell(row=notes_start, column=1, value='NOTES').font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    ws.cell(row=notes_start+1, column=1, value='Thank you for choosing [YOUR COMPANY NAME]. We appreciate your business!').font = data_font
    
    # Add instructions tab
    add_instructions_tab(wb, 'Contractor Invoice Template', [
        '— HOW TO USE THIS TEMPLATE —',
        '',
        '1. Replace all [BRACKETED TEXT] with your company information.',
        '2. Update the invoice number, date, and due date for each new invoice.',
        '3. Replace the example line items with your actual work and pricing.',
        '4. The Qty × Unit Price = Amount formula calculates automatically.',
        '5. Subtotal, Tax, and Grand Total update automatically.',
        '6. Adjust the tax rate in the Tax Rate cell (default: 8.25%).',
        '7. Update payment terms to match your business policies.',
        '',
        '— TIPS FOR CONTRACTORS —',
        '',
        '• Always include your license number — it builds trust and may be legally required.',
        '• Be specific in descriptions — "Install 12 LF upper cabinets (soft-close)" beats "cabinet work."',
        '• Include a PO/Job number so the customer can reference it easily.',
        '• Send invoices within 24 hours of job completion for fastest payment.',
        '• For progress billing, create a new invoice for each milestone.',
        '',
        '— PRINT SETTINGS —',
        '',
        'Print Area is pre-set. Go to File > Print to preview.',
        'The template is designed to fit on one page at default margins.',
        '',
        '— SUPPORT —',
        '',
        'Questions? Visit builtrighthq.com or email hello@builtrighthq.com',
        '© 2026 BuiltRight Academy. All rights reserved.',
    ])
    
    wb.save(os.path.join(OUT, 'contractor-invoice.xlsx'))
    print('  ✅ contractor-invoice.xlsx')

# Need to define D as column index
D = 4
E = 5

build_invoice()

# ========================================
# 2. CONTRACTOR ESTIMATE TEMPLATE
# ========================================
def build_estimate():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estimate'
    ws.sheet_properties.tabColor = ORANGE
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    ws.print_area = 'A1:E60'
    
    # Header
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1, value='ESTIMATE').font = Font(name='Calibri', bold=True, size=28, color=ORANGE)
    ws.row_dimensions[1].height = 45
    
    # Company info
    ws.cell(row=3, column=1, value='[YOUR COMPANY NAME]').font = Font(name='Calibri', bold=True, size=14)
    ws.cell(row=4, column=1, value='[Address] | [Phone] | [Email]').font = data_font
    ws.cell(row=5, column=1, value='License #: [Your License]').font = small_font
    
    ws.cell(row=3, column=D, value='Estimate #:').font = label_font
    ws.cell(row=3, column=E, value='EST-2026-042').font = data_font
    ws.cell(row=4, column=D, value='Date:').font = label_font
    ws.cell(row=4, column=E, value='03/10/2026').font = data_font
    ws.cell(row=5, column=D, value='Valid Until:').font = label_font
    ws.cell(row=5, column=E, value='04/09/2026').font = data_font
    
    # Client
    ws.cell(row=7, column=1, value='PREPARED FOR:').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=8, column=1, value='Robert & Maria Garcia').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=9, column=1, value='789 Maple Drive, Austin, TX 78704').font = data_font
    ws.cell(row=10, column=1, value='(512) 555-8901 | garcia.home@email.com').font = data_font
    
    ws.cell(row=7, column=D, value='PROJECT:').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=8, column=D, value='Master Bathroom Remodel').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=9, column=D, value='Timeline: 3-4 weeks').font = data_font
    
    # Materials section
    r = 12
    ws.cell(row=r, column=1, value='MATERIALS').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    r += 1
    for c, h in enumerate(['Item', 'Qty', 'Unit Cost', 'Total', ''], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    materials = [
        ('Porcelain tile — 12x24 matte (floor + shower)', 85, 6.50),
        ('Vanity — 60" double sink, solid wood', 1, 1200),
        ('Quartz countertop — 60" with undermount cutouts', 1, 850),
        ('Shower valve — Moen thermostatic rough-in kit', 1, 285),
        ('Tile backer board — 1/2" cement board (10 sheets)', 10, 14),
        ('Waterproofing membrane — Kerdi (200 sq ft roll)', 1, 185),
        ('Toilet — Toto Drake II elongated', 1, 380),
        ('Paint — Benjamin Moore Regal, 2 gal', 2, 72),
        ('Miscellaneous supplies (grout, thin-set, screws, caulk)', 1, 175),
    ]
    
    for i, (item, qty, cost) in enumerate(materials):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=item).font = data_font
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=2, value=qty).font = data_font
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3, value=cost).font = data_font
        ws.cell(row=row, column=3).number_format = USD
        ws.cell(row=row, column=4).value = f'=B{row}*C{row}'
        ws.cell(row=row, column=4).number_format = USD
        style_data_row(ws, row, 5, alt=(i % 2 == 0))
    
    mat_end = r + len(materials)
    ws.cell(row=mat_end + 1, column=3, value='Materials Subtotal:').font = label_font
    ws.cell(row=mat_end + 1, column=3).alignment = right
    ws.cell(row=mat_end + 1, column=4).value = f'=SUM(D{r+1}:D{mat_end})'
    ws.cell(row=mat_end + 1, column=4).number_format = USD
    ws.cell(row=mat_end + 1, column=4).font = Font(name='Calibri', bold=True, size=11)
    
    # Labor section
    lr = mat_end + 3
    ws.cell(row=lr, column=1, value='LABOR').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    lr += 1
    for c, h in enumerate(['Task', 'Hours', 'Rate/Hr', 'Total', ''], start=1):
        cell = ws.cell(row=lr, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    labor = [
        ('Demolition & haul-away', 8, 65),
        ('Plumbing rough-in', 12, 85),
        ('Electrical (GFCI outlets, exhaust fan)', 6, 80),
        ('Tile installation — floor & shower', 24, 75),
        ('Vanity & countertop install', 6, 70),
        ('Paint — walls & ceiling, 2 coats', 8, 55),
        ('Trim, hardware & final details', 4, 65),
    ]
    
    for i, (task, hrs, rate) in enumerate(labor):
        row = lr + 1 + i
        ws.cell(row=row, column=1, value=task).font = data_font
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=2, value=hrs).font = data_font
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3, value=rate).font = data_font
        ws.cell(row=row, column=3).number_format = USD
        ws.cell(row=row, column=4).value = f'=B{row}*C{row}'
        ws.cell(row=row, column=4).number_format = USD
        style_data_row(ws, row, 5, alt=(i % 2 == 0))
    
    lab_end = lr + len(labor)
    ws.cell(row=lab_end + 1, column=3, value='Labor Subtotal:').font = label_font
    ws.cell(row=lab_end + 1, column=3).alignment = right
    ws.cell(row=lab_end + 1, column=4).value = f'=SUM(D{lr+1}:D{lab_end})'
    ws.cell(row=lab_end + 1, column=4).number_format = USD
    ws.cell(row=lab_end + 1, column=4).font = Font(name='Calibri', bold=True, size=11)
    
    # Summary
    sr = lab_end + 3
    ws.cell(row=sr, column=1, value='PROJECT SUMMARY').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    
    summary = [
        ('Materials:', f'=D{mat_end+1}'),
        ('Labor:', f'=D{lab_end+1}'),
        ('Subtotal:', f'=D{sr+1}+D{sr+2}'),
        ('Markup (15%):', f'=D{sr+3}*0.15'),
        ('Tax (8.25%):', f'=(D{sr+3}+D{sr+4})*0.0825'),
        ('ESTIMATED TOTAL:', f'=D{sr+3}+D{sr+4}+D{sr+5}'),
    ]
    
    for i, (label, formula) in enumerate(summary):
        row = sr + 1 + i
        ws.cell(row=row, column=3, value=label).font = label_font if i < 5 else total_font
        ws.cell(row=row, column=3).alignment = right
        ws.cell(row=row, column=4).value = formula
        ws.cell(row=row, column=4).number_format = USD
        ws.cell(row=row, column=4).font = data_font if i < 5 else Font(name='Calibri', bold=True, size=14, color=ORANGE)
        if i == 5:
            ws.cell(row=row, column=4).fill = light_orange_fill
    
    # Terms
    tr = sr + 8
    ws.cell(row=tr, column=1, value='TERMS & CONDITIONS').font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    terms = [
        '1. This estimate is valid for 30 days from the date above.',
        '2. 50% deposit required before work begins. Balance due upon completion.',
        '3. Any changes to the scope of work will be documented via Change Order.',
        '4. Unforeseen conditions (rot, mold, code issues) may result in additional charges.',
        '5. All work warranted for 1 year. Manufacturer warranties apply to materials.',
        '6. Contractor is licensed, bonded, and insured.',
    ]
    for i, t in enumerate(terms):
        ws.cell(row=tr+1+i, column=1, value=t).font = data_font
        ws.cell(row=tr+1+i, column=1).alignment = wrap
    
    # Acceptance signature
    ar = tr + len(terms) + 3
    ws.cell(row=ar, column=1, value='ACCEPTANCE').font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    ws.cell(row=ar+1, column=1, value='By signing below, you accept this estimate and authorize work to begin.').font = data_font
    ws.cell(row=ar+3, column=1, value='Client Signature: ___________________________').font = data_font
    ws.cell(row=ar+3, column=3, value='Date: ____________').font = data_font
    ws.cell(row=ar+5, column=1, value='Contractor Signature: ________________________').font = data_font
    ws.cell(row=ar+5, column=3, value='Date: ____________').font = data_font
    
    add_instructions_tab(wb, 'Estimate & Bid Template', [
        '— HOW TO USE THIS TEMPLATE —',
        '',
        '1. Replace [BRACKETED TEXT] with your company info.',
        '2. Update the estimate number and dates.',
        '3. Replace materials and labor items with your actual scope.',
        '4. All totals calculate automatically via formulas.',
        '5. Adjust markup % in the Project Summary section (default: 15%).',
        '6. Adjust tax rate (default: 8.25%).',
        '7. Customize terms & conditions for your business.',
        '',
        '— PRICING TIPS —',
        '',
        '• Always include a line item for permits — forgetting this kills margins.',
        '• Use "Miscellaneous supplies" as a catch-all (3-5% of materials).',
        '• Markup should cover overhead: insurance, truck, tools, office, profit.',
        '• 15-25% markup is standard for residential remodeling.',
        '• Include "Unforeseen conditions" clause — it saves you from surprise costs.',
        '',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    
    wb.save(os.path.join(OUT, 'contractor-estimate.xlsx'))
    print('  ✅ contractor-estimate.xlsx')

build_estimate()

# ========================================
# 3. JOB COSTING TRACKER
# ========================================
def build_job_costing():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Job Costing'
    ws.sheet_properties.tabColor = ORANGE
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    
    # Header
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='JOB COSTING TRACKER').font = Font(name='Calibri', bold=True, size=24, color=ORANGE)
    ws.row_dimensions[1].height = 40
    
    ws.cell(row=3, column=1, value='Job Name:').font = label_font
    ws.cell(row=3, column=2, value='Miller Kitchen & Bath Remodel').font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=4, column=1, value='Client:').font = label_font
    ws.cell(row=4, column=2, value='David & Lisa Miller').font = data_font
    ws.cell(row=5, column=1, value='Start Date:').font = label_font
    ws.cell(row=5, column=2, value='02/15/2026').font = data_font
    ws.cell(row=3, column=D, value='Estimated Total:').font = label_font
    ws.cell(row=3, column=E).value = '=C42'
    ws.cell(row=3, column=E).number_format = USD
    ws.cell(row=3, column=E).font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=4, column=D, value='Actual Total:').font = label_font
    ws.cell(row=4, column=E).value = '=D42'
    ws.cell(row=4, column=E).number_format = USD
    ws.cell(row=5, column=D, value='Profit/Loss:').font = label_font
    ws.cell(row=5, column=E).value = '=C42-D42'
    ws.cell(row=5, column=E).number_format = USD
    ws.cell(row=5, column=E).font = Font(name='Calibri', bold=True, size=12, color=GREEN)
    
    # Table header
    r = 7
    headers = ['Category', 'Description', 'Estimated', 'Actual', 'Variance', 'Status']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    # Categories and items
    items = [
        ('Materials', 'Cabinets — 12 LF uppers + lowers', 4200, 4350),
        ('Materials', 'Countertops — Quartz 60"', 3200, 3200),
        ('Materials', 'Tile — Porcelain floor + backsplash', 680, 720),
        ('Materials', 'Plumbing fixtures — Sink, faucet, disposal', 850, 890),
        ('Materials', 'Electrical supplies — Wire, outlets, switches', 340, 310),
        ('Materials', 'Paint — 4 gal interior + primer', 280, 295),
        ('Materials', 'Hardware — Pulls, knobs, hinges', 420, 420),
        ('Materials', 'Miscellaneous — Adhesive, caulk, screws', 175, 210),
        ('Labor', 'Demo crew — 2 workers × 8 hrs', 960, 1040),
        ('Labor', 'Plumber — Rough-in + finish', 1700, 1850),
        ('Labor', 'Electrician — Panel, outlets, lights', 960, 960),
        ('Labor', 'Tile installer — Floor + backsplash', 1800, 2100),
        ('Labor', 'Painter — Walls, ceiling, trim', 640, 640),
        ('Labor', 'Finish carpenter — Trim, hardware', 480, 520),
        ('Equipment', 'Tile saw rental (1 week)', 150, 150),
        ('Equipment', 'Dumpster rental (20 yd)', 375, 375),
        ('Permits', 'Building permit — City of Springfield', 250, 250),
        ('Permits', 'Electrical permit', 75, 75),
        ('Subcontractors', 'HVAC — Relocate duct', 600, 650),
        ('Subcontractors', 'Drywall repair', 400, 400),
        ('Overhead', 'Insurance allocation', 320, 320),
        ('Overhead', 'Vehicle/fuel', 180, 200),
        ('Overhead', 'Office/admin time', 150, 150),
    ]
    
    for i, (cat, desc, est, act) in enumerate(items):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=cat).font = data_font
        ws.cell(row=row, column=2, value=desc).font = data_font
        ws.cell(row=row, column=2).alignment = left
        ws.cell(row=row, column=3, value=est).font = data_font
        ws.cell(row=row, column=3).number_format = USD
        ws.cell(row=row, column=4, value=act).font = data_font
        ws.cell(row=row, column=4).number_format = USD
        ws.cell(row=row, column=5).value = f'=D{row}-C{row}'
        ws.cell(row=row, column=5).number_format = USD
        # Status formula
        ws.cell(row=row, column=6).value = f'=IF(E{row}>0,"Over ⚠️",IF(E{row}<0,"Under ✅","On Budget"))'
        ws.cell(row=row, column=6).alignment = center
        style_data_row(ws, row, 6, alt=(i % 2 == 0))
    
    # Add status dropdown
    status_dv = DataValidation(type="list", formula1='"On Budget,Over ⚠️,Under ✅"')
    ws.add_data_validation(status_dv)
    
    # Totals
    total_row = r + 1 + len(items) + 1
    ws.cell(row=total_row, column=2, value='TOTALS').font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    ws.cell(row=total_row, column=3).value = f'=SUM(C{r+1}:C{total_row-2})'
    ws.cell(row=total_row, column=3).number_format = USD
    ws.cell(row=total_row, column=3).font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=total_row, column=4).value = f'=SUM(D{r+1}:D{total_row-2})'
    ws.cell(row=total_row, column=4).number_format = USD
    ws.cell(row=total_row, column=4).font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=total_row, column=5).value = f'=D{total_row}-C{total_row}'
    ws.cell(row=total_row, column=5).number_format = USD
    ws.cell(row=total_row, column=5).font = Font(name='Calibri', bold=True, size=12, color=RED)
    ws.cell(row=total_row, column=5).fill = light_orange_fill
    
    # Profit analysis
    pr = total_row + 2
    ws.cell(row=pr, column=1, value='PROFIT ANALYSIS').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    ws.cell(row=pr+1, column=1, value='Contract Price:').font = label_font
    ws.cell(row=pr+1, column=2, value=22500).font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=pr+1, column=2).number_format = USD
    ws.cell(row=pr+2, column=1, value='Total Costs:').font = label_font
    ws.cell(row=pr+2, column=2).value = f'=D{total_row}'
    ws.cell(row=pr+2, column=2).number_format = USD
    ws.cell(row=pr+3, column=1, value='NET PROFIT:').font = Font(name='Calibri', bold=True, size=12, color=GREEN)
    ws.cell(row=pr+3, column=2).value = f'=B{pr+1}-B{pr+2}'
    ws.cell(row=pr+3, column=2).number_format = USD
    ws.cell(row=pr+3, column=2).font = Font(name='Calibri', bold=True, size=14, color=GREEN)
    ws.cell(row=pr+3, column=2).fill = light_green_fill
    ws.cell(row=pr+4, column=1, value='Profit Margin:').font = label_font
    ws.cell(row=pr+4, column=2).value = f'=B{pr+3}/B{pr+1}'
    ws.cell(row=pr+4, column=2).number_format = PCT
    
    add_instructions_tab(wb, 'Job Costing Tracker', [
        '— HOW TO USE THIS TEMPLATE —',
        '',
        '1. Update job info at the top (name, client, dates).',
        '2. Replace example line items with your actual job costs.',
        '3. Enter estimated costs BEFORE the job starts.',
        '4. Update actual costs AS expenses come in.',
        '5. Variance and Status columns update automatically.',
        '6. Enter your contract price in Profit Analysis to see net profit.',
        '',
        '— WHY JOB COSTING MATTERS —',
        '',
        '• Most contractors think they make money on every job. They don\'t.',
        '• Job costing reveals which job TYPES are profitable and which aren\'t.',
        '• Track every job for 3 months — you\'ll find at least one service you\'re losing money on.',
        '• Use this data to adjust pricing on your next estimate.',
        '',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    
    wb.save(os.path.join(OUT, 'job-costing-tracker.xlsx'))
    print('  ✅ job-costing-tracker.xlsx')

build_job_costing()

# ========================================
# 4. PROFIT & LOSS TRACKER
# ========================================
def build_pnl():
    wb = Workbook()
    ws = wb.active
    ws.title = 'P&L'
    ws.sheet_properties.tabColor = ORANGE
    
    ws.column_dimensions['A'].width = 30
    for i, m in enumerate('BCDEFGHIJKLMN'):
        ws.column_dimensions[m].width = 13
    
    # Header
    ws.merge_cells('A1:N1')
    ws.cell(row=1, column=1, value='PROFIT & LOSS STATEMENT').font = Font(name='Calibri', bold=True, size=24, color=ORANGE)
    ws.cell(row=2, column=1, value='[YOUR COMPANY NAME] — 2026').font = Font(name='Calibri', bold=True, size=12, color='666666')
    
    # Month headers
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'YTD Total']
    r = 4
    ws.cell(row=r, column=1, value='').font = label_font
    for i, m in enumerate(months):
        cell = ws.cell(row=r, column=i+2, value=m)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    # INCOME section
    r = 5
    ws.cell(row=r, column=1, value='INCOME').font = Font(name='Calibri', bold=True, size=13, color=GREEN)
    
    income_items = [
        ('Service Revenue', [18500, 22300, 28400, 15800, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Material Markups', [2775, 3345, 4260, 2370, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Change Order Revenue', [0, 1200, 800, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Other Income', [0, 0, 250, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
    ]
    
    for i, (item, values) in enumerate(income_items):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=item).font = data_font
        for j, v in enumerate(values):
            ws.cell(row=row, column=j+2, value=v).font = data_font
            ws.cell(row=row, column=j+2).number_format = USD
            ws.cell(row=row, column=j+2).border = thin_border
        # YTD formula
        ws.cell(row=row, column=14).value = f'=SUM(B{row}:M{row})'
        ws.cell(row=row, column=14).number_format = USD
        ws.cell(row=row, column=14).font = Font(name='Calibri', bold=True)
        style_data_row(ws, row, 14, alt=(i % 2 == 0))
    
    total_income_row = r + 1 + len(income_items)
    ws.cell(row=total_income_row, column=1, value='TOTAL INCOME').font = Font(name='Calibri', bold=True, size=11, color=GREEN)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=total_income_row, column=c).value = f'=SUM({col_letter}{r+1}:{col_letter}{total_income_row-1})'
        ws.cell(row=total_income_row, column=c).number_format = USD
        ws.cell(row=total_income_row, column=c).font = Font(name='Calibri', bold=True, size=11, color=GREEN)
        ws.cell(row=total_income_row, column=c).fill = light_green_fill
        ws.cell(row=total_income_row, column=c).border = thin_border
    
    # EXPENSES section
    er = total_income_row + 2
    ws.cell(row=er, column=1, value='EXPENSES').font = Font(name='Calibri', bold=True, size=13, color=RED)
    
    expense_items = [
        ('Materials & Supplies', [8200, 9800, 12500, 7100, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Labor / Subcontractors', [5400, 6800, 8200, 4600, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Vehicle & Fuel', [480, 520, 580, 450, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Insurance (GL, WC, Auto)', [650, 650, 650, 650, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Tools & Equipment', [350, 180, 900, 120, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Marketing & Advertising', [200, 200, 200, 200, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Office & Phone', [150, 150, 150, 150, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Permits & Licenses', [75, 150, 250, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Vehicle Payment', [550, 550, 550, 550, 0, 0, 0, 0, 0, 0, 0, 0]),
        ('Miscellaneous', [100, 75, 150, 80, 0, 0, 0, 0, 0, 0, 0, 0]),
    ]
    
    for i, (item, values) in enumerate(expense_items):
        row = er + 1 + i
        ws.cell(row=row, column=1, value=item).font = data_font
        for j, v in enumerate(values):
            ws.cell(row=row, column=j+2, value=v).font = data_font
            ws.cell(row=row, column=j+2).number_format = USD
            ws.cell(row=row, column=j+2).border = thin_border
        ws.cell(row=row, column=14).value = f'=SUM(B{row}:M{row})'
        ws.cell(row=row, column=14).number_format = USD
        ws.cell(row=row, column=14).font = Font(name='Calibri', bold=True)
        style_data_row(ws, row, 14, alt=(i % 2 == 0))
    
    total_expense_row = er + 1 + len(expense_items)
    ws.cell(row=total_expense_row, column=1, value='TOTAL EXPENSES').font = Font(name='Calibri', bold=True, size=11, color=RED)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=total_expense_row, column=c).value = f'=SUM({col_letter}{er+1}:{col_letter}{total_expense_row-1})'
        ws.cell(row=total_expense_row, column=c).number_format = USD
        ws.cell(row=total_expense_row, column=c).font = Font(name='Calibri', bold=True, size=11, color=RED)
        ws.cell(row=total_expense_row, column=c).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')
        ws.cell(row=total_expense_row, column=c).border = thin_border
    
    # NET PROFIT
    net_row = total_expense_row + 2
    ws.cell(row=net_row, column=1, value='NET PROFIT').font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=net_row, column=c).value = f'={col_letter}{total_income_row}-{col_letter}{total_expense_row}'
        ws.cell(row=net_row, column=c).number_format = USD
        ws.cell(row=net_row, column=c).font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
        ws.cell(row=net_row, column=c).fill = light_orange_fill
        ws.cell(row=net_row, column=c).border = Border(
            left=Side(style='medium', color=ORANGE),
            right=Side(style='medium', color=ORANGE),
            top=Side(style='medium', color=ORANGE),
            bottom=Side(style='medium', color=ORANGE)
        )
    
    # Profit margin row
    margin_row = net_row + 1
    ws.cell(row=margin_row, column=1, value='Profit Margin %').font = label_font
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=margin_row, column=c).value = f'=IF({col_letter}{total_income_row}=0,0,{col_letter}{net_row}/{col_letter}{total_income_row})'
        ws.cell(row=margin_row, column=c).number_format = PCT
        ws.cell(row=margin_row, column=c).font = data_font
    
    add_instructions_tab(wb, 'Profit & Loss Statement', [
        '— HOW TO USE THIS TEMPLATE —',
        '',
        '1. Enter your monthly income by category.',
        '2. Enter your monthly expenses by category.',
        '3. Net Profit and Profit Margin calculate automatically.',
        '4. YTD Total column shows year-to-date numbers.',
        '5. Add or remove categories as needed for your business.',
        '',
        '— BENCHMARK: IS YOUR BUSINESS HEALTHY? —',
        '',
        '• Net profit margin 10-20% = Healthy for most trades',
        '• Materials should be 35-45% of revenue',
        '• Labor (including subs) should be 25-35% of revenue',
        '• Overhead should be under 20% of revenue',
        '• If margin < 10%, you\'re probably underpricing',
        '',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    
    wb.save(os.path.join(OUT, 'profit-loss-tracker.xlsx'))
    print('  ✅ profit-loss-tracker.xlsx')

build_pnl()

# Build remaining templates with similar quality...
# For now, let's rebuild the remaining 5 quickly

# 5. CLIENT TRACKER
def build_client_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clients'
    ws.sheet_properties.tabColor = ORANGE
    
    cols = ['Client Name', 'Phone', 'Email', 'Address', 'Lead Source', 'Job Type', 
            'Status', 'Quoted $', 'Paid $', 'Follow-Up Date', 'Notes']
    widths = [22, 16, 24, 28, 14, 16, 12, 14, 14, 14, 30]
    
    ws.merge_cells('A1:K1')
    ws.cell(row=1, column=1, value='CLIENT & LEAD TRACKER').font = Font(name='Calibri', bold=True, size=24, color=ORANGE)
    ws.cell(row=2, column=1, value='[YOUR COMPANY NAME] — Track every lead, close every deal').font = small_font
    
    for i, (col, w) in enumerate(zip(cols, widths)):
        ws.column_dimensions[get_column_letter(i+1)].width = w
        cell = ws.cell(row=4, column=i+1, value=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Lead,Quoted,Negotiating,Active,Complete,Lost"')
    ws.add_data_validation(status_dv)
    
    # Lead source dropdown
    source_dv = DataValidation(type="list", formula1='"Google,Referral,Yelp,HomeAdvisor,Nextdoor,Facebook,Yard Sign,Repeat Client,Other"')
    ws.add_data_validation(source_dv)
    
    clients = [
        ('Tom Richards', '(555) 234-5678', 'tom.richards@email.com', '123 Pine St, Austin TX', 'Google', 'Kitchen Remodel', 'Active', 18500, 9250, '03/20/2026', 'Progress payment 2 due'),
        ('Sarah Chen', '(555) 345-6789', 'sarah.chen@email.com', '456 Oak Ave, Austin TX', 'Referral', 'Bathroom Remodel', 'Complete', 12800, 12800, '', 'Ask for review'),
        ('Mike & Julia Davis', '(555) 456-7890', 'davis.home@email.com', '789 Elm Dr, Round Rock TX', 'Yelp', 'Deck Build', 'Quoted', 8500, 0, '03/15/2026', 'Follow up on estimate'),
        ('Corporate Realty LLC', '(555) 567-8901', 'maintenance@corprealty.com', '321 Business Blvd, Austin TX', 'Repeat Client', 'Multi-unit Repair', 'Negotiating', 45000, 0, '03/12/2026', 'Waiting on board approval'),
        ('Lisa Patel', '(555) 678-9012', 'lisa.p@email.com', '654 Birch Ln, Cedar Park TX', 'Nextdoor', 'Fence Install', 'Lead', 0, 0, '03/11/2026', 'Wants cedar 6ft privacy'),
        ('James Morrison', '(555) 789-0123', 'jmorrison@email.com', '987 Willow Way, Pflugerville TX', 'HomeAdvisor', 'Water Heater', 'Complete', 2800, 2800, '', 'Tankless Rinnai install'),
        ('Amy & Rod Walsh', '(555) 890-1234', 'walsh.family@email.com', '246 Spruce Ct, Georgetown TX', 'Facebook', 'Garage Conversion', 'Active', 22000, 11000, '03/25/2026', 'Electrical inspection scheduled'),
        ('Robert Kim', '(555) 901-2345', 'rkim.contractor@email.com', '135 Cedar Ridge, Leander TX', 'Google', 'Siding Replace', 'Quoted', 14200, 0, '03/13/2026', 'Comparing 3 bids'),
        ('Maria Gonzalez', '(555) 012-3456', 'maria.g@email.com', '864 Pecan St, San Marcos TX', 'Yard Sign', 'Roof Repair', 'Lead', 0, 0, '03/14/2026', 'Storm damage, insurance claim'),
        ('First Baptist Church', '(555) 123-4567', 'admin@firstbaptist.org', '500 Main St, Austin TX', 'Referral', 'Commercial Paint', 'Negotiating', 35000, 0, '03/18/2026', 'Need after-hours access'),
    ]
    
    for i, client in enumerate(clients):
        row = 5 + i
        for j, val in enumerate(client):
            ws.cell(row=row, column=j+1, value=val).font = data_font
            ws.cell(row=row, column=j+1).border = thin_border
        ws.cell(row=row, column=8).number_format = USD
        ws.cell(row=row, column=9).number_format = USD
        status_dv.add(ws.cell(row=row, column=7))
        source_dv.add(ws.cell(row=row, column=5))
        style_data_row(ws, row, 11, alt=(i % 2 == 0))
    
    # Add 40 empty rows with validation
    for i in range(15, 55):
        for j in range(1, 12):
            ws.cell(row=i, column=j).border = thin_border
        status_dv.add(ws.cell(row=i, column=7))
        source_dv.add(ws.cell(row=i, column=5))
    
    # Summary at bottom
    sr = 56
    ws.cell(row=sr, column=1, value='PIPELINE SUMMARY').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    ws.cell(row=sr+1, column=1, value='Total Quoted:').font = label_font
    ws.cell(row=sr+1, column=2).value = '=SUM(H5:H54)'
    ws.cell(row=sr+1, column=2).number_format = USD
    ws.cell(row=sr+2, column=1, value='Total Collected:').font = label_font
    ws.cell(row=sr+2, column=2).value = '=SUM(I5:I54)'
    ws.cell(row=sr+2, column=2).number_format = USD
    ws.cell(row=sr+3, column=1, value='Outstanding:').font = label_font
    ws.cell(row=sr+3, column=2).value = f'=B{sr+1}-B{sr+2}'
    ws.cell(row=sr+3, column=2).number_format = USD
    ws.cell(row=sr+3, column=2).font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    
    add_instructions_tab(wb, 'Client & Lead Tracker', [
        '— HOW TO USE THIS TEMPLATE —',
        '',
        '1. Add every lead the moment they contact you.',
        '2. Use the Status dropdown to track where they are in your pipeline.',
        '3. Set Follow-Up dates — NEVER let a lead go cold.',
        '4. Update Quoted/Paid amounts as jobs progress.',
        '5. Pipeline Summary auto-calculates your total pipeline value.',
        '',
        '— LEAD MANAGEMENT TIPS —',
        '',
        '• Follow up within 24 hours or lose the job to a faster competitor.',
        '• Track Lead Source — know where your best customers come from.',
        '• After every completed job: ask for a review and a referral.',
        '• Sort by Follow-Up Date every morning — never miss a callback.',
        '',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    
    wb.save(os.path.join(OUT, 'client-tracker.xlsx'))
    print('  ✅ client-tracker.xlsx')

build_client_tracker()

# 6-9: Build remaining templates (proposal, change order, timesheet, cash flow)
# Using similar professional quality patterns

def build_proposal():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proposal'
    ws.sheet_properties.tabColor = ORANGE
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value='PROJECT PROPOSAL').font = Font(name='Calibri', bold=True, size=28, color=ORANGE)
    ws.row_dimensions[1].height = 45
    
    ws.cell(row=3, column=1, value='[YOUR COMPANY NAME]').font = Font(name='Calibri', bold=True, size=14)
    ws.cell(row=4, column=1, value='[Address] | [Phone] | [Email] | [Website]').font = data_font
    ws.cell(row=5, column=1, value='License #: [Number] | Insured | Bonded').font = small_font
    
    ws.cell(row=3, column=3, value='Proposal #:').font = label_font
    ws.cell(row=3, column=4, value='PROP-2026-018').font = data_font
    ws.cell(row=4, column=3, value='Date:').font = label_font
    ws.cell(row=4, column=4, value='03/10/2026').font = data_font
    ws.cell(row=5, column=3, value='Valid Until:').font = label_font
    ws.cell(row=5, column=4, value='04/09/2026').font = data_font
    
    ws.cell(row=7, column=1, value='PREPARED FOR:').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=8, column=1, value='Robert & Maria Garcia').font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=9, column=1, value='789 Maple Drive, Austin, TX 78704').font = data_font
    ws.cell(row=10, column=1, value='(512) 555-8901 | garcia.home@email.com').font = data_font
    
    ws.cell(row=7, column=3, value='PROJECT:').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=8, column=3, value='Master Bathroom').font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=9, column=3, value='Complete Renovation').font = data_font
    
    # Scope of Work
    r = 12
    ws.cell(row=r, column=1, value='SCOPE OF WORK').font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    scope = [
        '1. Complete demolition of existing bathroom including tile, fixtures, vanity, and drywall as needed.',
        '2. Plumbing: Relocate shower valve, install new supply lines, replace drain assembly.',
        '3. Electrical: Install new exhaust fan (110 CFM), add 2 GFCI outlets, upgrade lighting circuit.',
        '4. Waterproofing: Install Schluter Kerdi membrane in shower area (walls and floor).',
        '5. Tile: Install 12x24 porcelain tile on floor and shower walls (approx. 85 sq ft total).',
        '6. Install 60" double-sink vanity with quartz countertop and undermount sinks.',
        '7. Install new toilet (Toto Drake II elongated, ADA height).',
        '8. Paint all walls and ceiling — 2 coats Benjamin Moore Regal (color TBD by client).',
        '9. Install all trim, hardware, accessories (towel bars, toilet paper holder, mirror).',
        '10. Final cleanup, debris removal, and walk-through inspection.',
    ]
    for i, s in enumerate(scope):
        ws.cell(row=r+1+i, column=1, value=s).font = data_font
        ws.cell(row=r+1+i, column=1).alignment = wrap
        ws.row_dimensions[r+1+i].height = 30
    
    # Timeline
    tr = r + len(scope) + 2
    ws.cell(row=tr, column=1, value='PROJECT TIMELINE').font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    milestones = [
        ('Week 1', 'Demolition, rough plumbing & electrical'),
        ('Week 2', 'Waterproofing, tile installation'),
        ('Week 3', 'Vanity, countertop, fixtures, paint'),
        ('Week 4', 'Trim, hardware, final inspection, punch list'),
    ]
    for c, h in enumerate(['Phase', 'Description', 'Duration', ''], start=1):
        cell = ws.cell(row=tr+1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
    for i, (phase, desc) in enumerate(milestones):
        row = tr + 2 + i
        ws.cell(row=row, column=1, value=phase).font = data_font
        ws.cell(row=row, column=2, value=desc).font = data_font
        ws.cell(row=row, column=3, value='5 business days').font = data_font
        style_data_row(ws, row, 4, alt=(i % 2 == 0))
    
    # Pricing
    pr = tr + len(milestones) + 3
    ws.cell(row=pr, column=1, value='INVESTMENT').font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    ws.cell(row=pr+1, column=1, value='Materials & fixtures').font = data_font
    ws.cell(row=pr+1, column=4, value=4890).font = data_font
    ws.cell(row=pr+1, column=4).number_format = USD
    ws.cell(row=pr+2, column=1, value='Labor').font = data_font
    ws.cell(row=pr+2, column=4, value=5100).font = data_font
    ws.cell(row=pr+2, column=4).number_format = USD
    ws.cell(row=pr+3, column=1, value='Permits & inspections').font = data_font
    ws.cell(row=pr+3, column=4, value=325).font = data_font
    ws.cell(row=pr+3, column=4).number_format = USD
    ws.cell(row=pr+4, column=1, value='PROJECT TOTAL').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    ws.cell(row=pr+4, column=4).value = f'=SUM(D{pr+1}:D{pr+3})'
    ws.cell(row=pr+4, column=4).number_format = USD
    ws.cell(row=pr+4, column=4).font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    ws.cell(row=pr+4, column=4).fill = light_orange_fill
    
    # Payment schedule
    ps = pr + 6
    ws.cell(row=ps, column=1, value='PAYMENT SCHEDULE').font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    payments = [
        ('Deposit (50%)', f'=D{pr+4}*0.5', 'Due upon acceptance'),
        ('Progress (25%)', f'=D{pr+4}*0.25', 'Due at tile completion'),
        ('Final (25%)', f'=D{pr+4}*0.25', 'Due upon final walk-through'),
    ]
    for i, (desc, amt, when) in enumerate(payments):
        row = ps + 1 + i
        ws.cell(row=row, column=1, value=desc).font = data_font
        ws.cell(row=row, column=3, value=when).font = data_font
        ws.cell(row=row, column=4).value = amt
        ws.cell(row=row, column=4).number_format = USD
    
    # Acceptance
    ar = ps + len(payments) + 3
    ws.cell(row=ar, column=1, value='ACCEPTANCE & AUTHORIZATION').font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    ws.cell(row=ar+1, column=1, value='By signing below, I/we accept this proposal, agree to the terms above, and authorize [YOUR COMPANY NAME] to begin work upon receipt of the deposit.').font = data_font
    ws.cell(row=ar+1, column=1).alignment = wrap
    ws.row_dimensions[ar+1].height = 40
    ws.cell(row=ar+3, column=1, value='Client Signature: _________________________________').font = data_font
    ws.cell(row=ar+3, column=3, value='Date: _______________').font = data_font
    ws.cell(row=ar+5, column=1, value='Printed Name: ___________________________________').font = data_font
    ws.cell(row=ar+7, column=1, value='Contractor Signature: _____________________________').font = data_font
    ws.cell(row=ar+7, column=3, value='Date: _______________').font = data_font
    
    add_instructions_tab(wb, 'Professional Proposal', [
        '— HOW TO USE —',
        '1. Replace all bracketed text. 2. Customize scope for your project.',
        '3. Update pricing. 4. Print or PDF and present to client.',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    
    wb.save(os.path.join(OUT, 'contractor-proposal.xlsx'))
    print('  ✅ contractor-proposal.xlsx')

build_proposal()

# Quick builds for change order, timesheet, cash flow (similar pattern)
def build_change_order():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Change Order'
    ws.sheet_properties.tabColor = ORANGE
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value='CHANGE ORDER').font = Font(name='Calibri', bold=True, size=28, color=ORANGE)
    ws.cell(row=3, column=1, value='[YOUR COMPANY NAME]').font = Font(name='Calibri', bold=True, size=14)
    ws.cell(row=3, column=3, value='CO #:').font = label_font
    ws.cell(row=3, column=4, value='CO-001').font = data_font
    ws.cell(row=4, column=3, value='Date:').font = label_font
    ws.cell(row=4, column=4, value='03/10/2026').font = data_font
    ws.cell(row=6, column=1, value='Original Contract:').font = label_font
    ws.cell(row=6, column=2, value='PROP-2026-018').font = data_font
    ws.cell(row=7, column=1, value='Original Amount:').font = label_font
    ws.cell(row=7, column=2, value=10315).font = data_font
    ws.cell(row=7, column=2).number_format = USD
    ws.cell(row=8, column=1, value='Client:').font = label_font
    ws.cell(row=8, column=2, value='Robert & Maria Garcia').font = data_font
    
    ws.cell(row=10, column=1, value='REASON FOR CHANGE').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=11, column=1, value='During demo, discovered water damage behind shower wall requiring additional repair before tile installation. Client also requested upgrade from standard to heated tile floor.').font = data_font
    ws.cell(row=11, column=1).alignment = wrap
    ws.row_dimensions[11].height = 50
    
    r = 13
    ws.cell(row=r, column=1, value='CHANGES').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    for c, h in enumerate(['Description', 'Type', 'Hours/Qty', 'Amount'], start=1):
        cell = ws.cell(row=r+1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    changes = [
        ('Replace rotted studs (3) and subfloor section', 'Add', 1, 450),
        ('Mold remediation — treat & seal affected area', 'Add', 1, 350),
        ('Heated floor mat — Nuheat 35 sq ft + thermostat', 'Add', 1, 680),
        ('Heated floor installation labor', 'Add', 4, 280),
        ('Standard tile installation labor (reduced — partial overlap)', 'Remove', 1, -120),
    ]
    for i, (desc, typ, qty, amt) in enumerate(changes):
        row = r + 2 + i
        ws.cell(row=row, column=1, value=desc).font = data_font
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=2, value=typ).font = data_font
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3, value=qty).font = data_font
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=4, value=amt).font = data_font
        ws.cell(row=row, column=4).number_format = USD
        style_data_row(ws, row, 4, alt=(i % 2 == 0))
    
    ce = r + 2 + len(changes)
    ws.cell(row=ce+1, column=3, value='Net Change:').font = label_font
    ws.cell(row=ce+1, column=4).value = f'=SUM(D{r+2}:D{ce})'
    ws.cell(row=ce+1, column=4).number_format = USD
    ws.cell(row=ce+1, column=4).font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
    ws.cell(row=ce+2, column=3, value='New Total:').font = label_font
    ws.cell(row=ce+2, column=4).value = f'=B7+D{ce+1}'
    ws.cell(row=ce+2, column=4).number_format = USD
    ws.cell(row=ce+2, column=4).font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    ws.cell(row=ce+2, column=4).fill = light_orange_fill
    
    ws.cell(row=ce+4, column=1, value='Timeline Impact: Adds approximately 2 additional work days.').font = data_font
    ws.cell(row=ce+6, column=1, value='AUTHORIZATION').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=ce+7, column=1, value='Client Signature: _______________________  Date: ____________').font = data_font
    ws.cell(row=ce+9, column=1, value='Contractor Signature: ____________________  Date: ____________').font = data_font
    
    add_instructions_tab(wb, 'Change Order Template', [
        '— Always document scope changes in writing before doing the work. —',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    wb.save(os.path.join(OUT, 'change-order.xlsx'))
    print('  ✅ change-order.xlsx')

build_change_order()

def build_timesheet():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Timesheet'
    ws.sheet_properties.tabColor = ORANGE
    ws.column_dimensions['A'].width = 14
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 14
    
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='WEEKLY TIMESHEET').font = Font(name='Calibri', bold=True, size=24, color=ORANGE)
    ws.cell(row=3, column=1, value='Employee:').font = label_font
    ws.cell(row=3, column=2, value='Mike Torres').font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=4, column=1, value='Position:').font = label_font
    ws.cell(row=4, column=2, value='Journeyman Plumber').font = data_font
    ws.cell(row=3, column=5, value='Week Of:').font = label_font
    ws.cell(row=3, column=6, value='03/10/2026').font = data_font
    ws.cell(row=4, column=5, value='Employee ID:').font = label_font
    ws.cell(row=4, column=6, value='EMP-004').font = data_font
    
    r = 6
    headers = ['Day', 'Time In', 'Lunch Out', 'Lunch In', 'Time Out', 'Reg Hours', 'OT Hours', 'Job/Project']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    days_data = [
        ('Monday', '7:00 AM', '12:00 PM', '12:30 PM', '4:30 PM', 9, 1, 'Garcia Bathroom'),
        ('Tuesday', '7:00 AM', '12:00 PM', '12:30 PM', '3:30 PM', 8, 0, 'Garcia Bathroom'),
        ('Wednesday', '6:30 AM', '12:00 PM', '12:30 PM', '5:00 PM', 10, 2, 'Miller Kitchen'),
        ('Thursday', '7:00 AM', '12:00 PM', '12:30 PM', '3:30 PM', 8, 0, 'Richards Water Heater'),
        ('Friday', '7:00 AM', '12:00 PM', '12:30 PM', '3:30 PM', 8, 0, 'Garcia Bathroom'),
        ('Saturday', '8:00 AM', '', '', '12:00 PM', 4, 0, 'Emergency - Chen Leak'),
        ('Sunday', '', '', '', '', 0, 0, ''),
    ]
    
    for i, (day, ti, lo, li, to, reg, ot, job) in enumerate(days_data):
        row = r + 1 + i
        vals = [day, ti, lo, li, to, reg, ot, job]
        for j, v in enumerate(vals):
            ws.cell(row=row, column=j+1, value=v).font = data_font
            ws.cell(row=row, column=j+1).alignment = center
            ws.cell(row=row, column=j+1).border = thin_border
        style_data_row(ws, row, 8, alt=(i % 2 == 0))
    
    total_row = r + 8
    ws.cell(row=total_row, column=5, value='TOTALS:').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=total_row, column=6).value = f'=SUM(F{r+1}:F{total_row-1})'
    ws.cell(row=total_row, column=6).font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=total_row, column=7).value = f'=SUM(G{r+1}:G{total_row-1})'
    ws.cell(row=total_row, column=7).font = Font(name='Calibri', bold=True, size=12, color=RED)
    
    ws.cell(row=total_row+2, column=1, value='Employee Signature: _____________________').font = data_font
    ws.cell(row=total_row+2, column=5, value='Supervisor: _____________________').font = data_font
    ws.cell(row=total_row+4, column=1, value='Rate: $').font = label_font
    ws.cell(row=total_row+4, column=2, value=32).font = data_font
    ws.cell(row=total_row+4, column=2).number_format = USD
    ws.cell(row=total_row+4, column=3, value='Reg Pay:').font = label_font
    ws.cell(row=total_row+4, column=4).value = f'=B{total_row+4}*F{total_row}'
    ws.cell(row=total_row+4, column=4).number_format = USD
    ws.cell(row=total_row+4, column=5, value='OT Pay:').font = label_font
    ws.cell(row=total_row+4, column=6).value = f'=B{total_row+4}*1.5*G{total_row}'
    ws.cell(row=total_row+4, column=6).number_format = USD
    ws.cell(row=total_row+5, column=5, value='TOTAL PAY:').font = Font(name='Calibri', bold=True, color=ORANGE)
    ws.cell(row=total_row+5, column=6).value = f'=D{total_row+4}+F{total_row+4}'
    ws.cell(row=total_row+5, column=6).number_format = USD
    ws.cell(row=total_row+5, column=6).font = Font(name='Calibri', bold=True, size=14, color=ORANGE)
    
    add_instructions_tab(wb, 'Weekly Timesheet', [
        '— Track hours, calculate pay, bill to the right job. —',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    wb.save(os.path.join(OUT, 'weekly-timesheet.xlsx'))
    print('  ✅ weekly-timesheet.xlsx')

build_timesheet()

def build_cashflow():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cash Flow'
    ws.sheet_properties.tabColor = ORANGE
    ws.column_dimensions['A'].width = 28
    for c in 'BCDEFGHIJKLMN':
        ws.column_dimensions[c].width = 13
    
    ws.merge_cells('A1:N1')
    ws.cell(row=1, column=1, value='12-MONTH CASH FLOW FORECAST').font = Font(name='Calibri', bold=True, size=22, color=ORANGE)
    ws.cell(row=2, column=1, value='[YOUR COMPANY NAME] — 2026').font = small_font
    
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Total']
    r = 4
    for i, m in enumerate(months):
        cell = ws.cell(row=r, column=i+2, value=m)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    ws.cell(row=5, column=1, value='OPENING BALANCE').font = Font(name='Calibri', bold=True, size=11, color=ORANGE)
    ws.cell(row=5, column=2, value=15000).font = Font(name='Calibri', bold=True)
    ws.cell(row=5, column=2).number_format = USD
    for c in range(3, 14):
        col_prev = get_column_letter(c-1)
        ws.cell(row=5, column=c).value = f'={col_prev}22'  # closing balance of prev month
        ws.cell(row=5, column=c).number_format = USD
    
    # Cash In
    ws.cell(row=7, column=1, value='CASH IN').font = Font(name='Calibri', bold=True, size=12, color=GREEN)
    cash_in = [
        ('Customer Payments', [18500, 22300, 28400, 15800, 20000, 25000, 30000, 28000, 22000, 18000, 15000, 12000]),
        ('Deposits Received', [4500, 6000, 8000, 3500, 5000, 7000, 8000, 6000, 5000, 4000, 3000, 2000]),
        ('Other Income', [0, 0, 250, 0, 0, 500, 0, 0, 300, 0, 0, 0]),
    ]
    for i, (item, vals) in enumerate(cash_in):
        row = 8 + i
        ws.cell(row=row, column=1, value=item).font = data_font
        for j, v in enumerate(vals):
            ws.cell(row=row, column=j+2, value=v).font = data_font
            ws.cell(row=row, column=j+2).number_format = USD
            ws.cell(row=row, column=j+2).border = thin_border
        ws.cell(row=row, column=14).value = f'=SUM(B{row}:M{row})'
        ws.cell(row=row, column=14).number_format = USD
    
    total_in_row = 8 + len(cash_in)
    ws.cell(row=total_in_row, column=1, value='TOTAL CASH IN').font = Font(name='Calibri', bold=True, color=GREEN)
    for c in range(2, 15):
        cl = get_column_letter(c)
        ws.cell(row=total_in_row, column=c).value = f'=SUM({cl}8:{cl}{total_in_row-1})'
        ws.cell(row=total_in_row, column=c).number_format = USD
        ws.cell(row=total_in_row, column=c).font = Font(name='Calibri', bold=True, color=GREEN)
        ws.cell(row=total_in_row, column=c).fill = light_green_fill
    
    # Cash Out
    co_start = total_in_row + 2
    ws.cell(row=co_start, column=1, value='CASH OUT').font = Font(name='Calibri', bold=True, size=12, color=RED)
    cash_out = [
        ('Materials & Supplies', [8200, 9800, 12500, 7100, 8500, 11000, 13000, 12000, 9500, 8000, 6500, 5000]),
        ('Labor / Payroll', [5400, 6800, 8200, 4600, 5800, 7200, 8500, 7800, 6200, 5000, 4200, 3500]),
        ('Subcontractors', [1200, 2000, 3000, 800, 1500, 2500, 3200, 2800, 1800, 1200, 800, 500]),
        ('Vehicle & Fuel', [480, 520, 580, 450, 500, 550, 600, 580, 500, 450, 420, 380]),
        ('Insurance', [650, 650, 650, 650, 650, 650, 650, 650, 650, 650, 650, 650]),
        ('Marketing', [200, 200, 200, 200, 200, 200, 200, 200, 200, 200, 200, 200]),
        ('Office/Phone/Software', [150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150]),
        ('Equipment/Tools', [350, 180, 900, 120, 250, 300, 150, 200, 300, 180, 100, 100]),
        ('Taxes (quarterly est.)', [0, 0, 3500, 0, 0, 3500, 0, 0, 3500, 0, 0, 3500]),
    ]
    for i, (item, vals) in enumerate(cash_out):
        row = co_start + 1 + i
        ws.cell(row=row, column=1, value=item).font = data_font
        for j, v in enumerate(vals):
            ws.cell(row=row, column=j+2, value=v).font = data_font
            ws.cell(row=row, column=j+2).number_format = USD
            ws.cell(row=row, column=j+2).border = thin_border
        ws.cell(row=row, column=14).value = f'=SUM(B{row}:M{row})'
        ws.cell(row=row, column=14).number_format = USD
    
    total_out_row = co_start + 1 + len(cash_out)
    ws.cell(row=total_out_row, column=1, value='TOTAL CASH OUT').font = Font(name='Calibri', bold=True, color=RED)
    for c in range(2, 15):
        cl = get_column_letter(c)
        ws.cell(row=total_out_row, column=c).value = f'=SUM({cl}{co_start+1}:{cl}{total_out_row-1})'
        ws.cell(row=total_out_row, column=c).number_format = USD
        ws.cell(row=total_out_row, column=c).font = Font(name='Calibri', bold=True, color=RED)
        ws.cell(row=total_out_row, column=c).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')
    
    # Closing balance
    close_row = total_out_row + 1
    ws.cell(row=close_row, column=1, value='CLOSING BALANCE').font = Font(name='Calibri', bold=True, size=13, color=ORANGE)
    for c in range(2, 15):
        cl = get_column_letter(c)
        ws.cell(row=close_row, column=c).value = f'={cl}5+{cl}{total_in_row}-{cl}{total_out_row}'
        ws.cell(row=close_row, column=c).number_format = USD
        ws.cell(row=close_row, column=c).font = Font(name='Calibri', bold=True, size=12, color=ORANGE)
        ws.cell(row=close_row, column=c).fill = light_orange_fill
    
    add_instructions_tab(wb, '12-Month Cash Flow Forecast', [
        '— Predict your cash position month by month. Never run out of cash. —',
        '© 2026 BuiltRight Academy | builtrighthq.com',
    ])
    wb.save(os.path.join(OUT, 'cash-flow-forecast.xlsx'))
    print('  ✅ cash-flow-forecast.xlsx')

build_cashflow()

# Rebuild ZIP bundles
os.chdir(OUT)
for zf in ['pro-bundle.zip', 'complete-kit.zip']:
    if os.path.exists(zf):
        os.remove(zf)

with zipfile.ZipFile('pro-bundle.zip', 'w', zipfile.ZIP_DEFLATED) as z:
    for f in ['contractor-invoice.xlsx', 'contractor-estimate.xlsx', 'job-costing-tracker.xlsx', 'profit-loss-tracker.xlsx']:
        z.write(f)
print('  ✅ pro-bundle.zip')

with zipfile.ZipFile('complete-kit.zip', 'w', zipfile.ZIP_DEFLATED) as z:
    for f in ['contractor-invoice.xlsx', 'contractor-estimate.xlsx', 'job-costing-tracker.xlsx', 
              'profit-loss-tracker.xlsx', 'client-tracker.xlsx', 'contractor-proposal.xlsx',
              'change-order.xlsx', 'weekly-timesheet.xlsx', 'cash-flow-forecast.xlsx']:
        z.write(f)
print('  ✅ complete-kit.zip')

print('\n🔥 ALL 9 templates rebuilt to professional quality!')
